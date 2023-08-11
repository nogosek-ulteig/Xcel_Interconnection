#Made by Joe Nogosek, joe.nogosek@ulteig.com
#10/26/2021
#NOTE: Tracker must be closed on your local computer!

import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
import shutil

# Now get the newly added reviews ready for me to complete
IA_num = []
case_num = []
size = []

wb = openpyxl.load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2022']

i=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num.append(ws[row][1].value)
        case_num.append(ws[row][8].value)
        size.append(ws[row][6].value)
        i=i+1
for value in range(i):
    name = str(IA_num[value])
    full_name = r'G:\2021\21.00016\Reviews\JN\{} - JN'.format(name)
    if not os.path.exists(full_name):
        os.makedirs(full_name)
    if size[value] <= 40:
        old_file = os.path.join(full_name, 'Xcel CR Checklist_under40kW_metering - IA.xlsm')
        new_file = os.path.join(full_name, 'Xcel CR Checklist_under40kW_metering - IA{}.xlsm'.format(name[-5:]))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2021\21.00016\Reviews\Xcel CR Checklist_under40kW_metering - IA.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['DER Study']
            ws['I1'].value = name
            ws['C1'].value = case_num[value]
            wb.save(new_file)
    else:
        old_file = os.path.join(full_name, 'Xcel Completeness Review Checklist_over40kW - IA.xlsm')
        new_file = os.path.join(full_name, 'Xcel Completeness Review Checklist_over40kW - IA{}.xlsm'.format(name[-5:]))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2021\21.00016\Reviews\Xcel Completeness Review Checklist_over40kW - IA.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['DER Study']
            ws['I1'].value = name
            ws['C1'].value = case_num[value]
            wb.save(new_file)
