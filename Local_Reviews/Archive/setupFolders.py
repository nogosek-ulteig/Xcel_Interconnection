#Made by Joe Nogosek, joe.nogosek@ulteig.com
#10/26/2021
#NOTE: Tracker must be closed on your local computer!

import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
import shutil
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import glob
from selenium.webdriver.chrome.options import Options

username = 'joseph.h.nogosek@xcelenergy.com'
password = 'airdoc4Ee'

warnings.filterwarnings("ignore", category=DeprecationWarning)

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe", options=options)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(10)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Now get the newly added reviews ready for me to complete
IA_num_Nick = []
case_num_Nick = []
size_Nick = []
program_Nick = []
type_Nick = []
installer_Nick = []
due_date_Nick = []
reviewer_Nick = []
meter_Nick = []

IA_num_Ethan = []
case_num_Ethan = []
size_Ethan = []
program_Ethan = []
type_Ethan = []
installer_Ethan = []
due_date_Ethan = []
reviewer_Ethan = []
meter_Ethan = []

IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
installer_Joe = []
due_date_Joe = []
reviewer_Joe = []
meter_Joe = []

IA_num_Jose = []
case_num_Jose = []
size_Jose = []
program_Jose = []
type_Jose = []
installer_Jose = []
due_date_Jose = []
reviewer_Jose = []
meter_Jose = []

IA_num_Ross = []
case_num_Ross = []
size_Ross = []
program_Ross = []
type_Ross = []
installer_Ross = []
due_date_Ross = []
reviewer_Ross = []
meter_Ross = []

IA_num_Josh = []
case_num_Josh = []
size_Josh = []
program_Josh = []
type_Josh = []
installer_Josh = []
due_date_Josh = []
reviewer_Josh = []
meter_Josh = []

IA_num_Adan = []
case_num_Adan = []
size_Adan = []
program_Adan = []
type_Adan = []
installer_Adan = []
due_date_Adan = []
reviewer_Adan = []
meter_Adan = []

wb = load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2023']

# Nick
i=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Nick C" and ws[row][10].value==None:
        IA_num_Nick.append(ws[row][1].value)
        case_num_Nick.append(ws[row][8].value)
        size_Nick.append(ws[row][6].value)
        program_Nick.append(ws[row][15].value)
        type_Nick.append(ws[row][4].value)
        installer_Nick.append(ws[row][9].value)
        due_date_Nick.append(ws[row][13].value)
        reviewer_Nick.append(ws[row][7].value)
        meter_Nick.append(ws[row][3].value)
        i=i+1

# Ethan
j=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ethan U" and ws[row][10].value==None:
        IA_num_Ethan.append(ws[row][1].value)
        case_num_Ethan.append(ws[row][8].value)
        size_Ethan.append(ws[row][6].value)
        program_Ethan.append(ws[row][15].value)
        type_Ethan.append(ws[row][4].value)
        installer_Ethan.append(ws[row][9].value)
        due_date_Ethan.append(ws[row][13].value)
        reviewer_Ethan.append(ws[row][7].value)
        meter_Ethan.append(ws[row][3].value)
        j=j+1

# Joe
k=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        installer_Joe.append(ws[row][9].value)
        due_date_Joe.append(ws[row][13].value)
        reviewer_Joe.append(ws[row][7].value)
        meter_Joe.append(ws[row][3].value)
        k=k+1

# Jose
l=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Jose CN" and ws[row][10].value==None:
        IA_num_Jose.append(ws[row][1].value)
        case_num_Jose.append(ws[row][8].value)
        size_Jose.append(ws[row][6].value)
        program_Jose.append(ws[row][15].value)
        type_Jose.append(ws[row][4].value)
        installer_Jose.append(ws[row][9].value)
        due_date_Jose.append(ws[row][13].value)
        reviewer_Jose.append(ws[row][7].value)
        meter_Jose.append(ws[row][3].value)
        l=l+1

# Ross
m=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ross K" and ws[row][10].value==None:
        IA_num_Ross.append(ws[row][1].value)
        case_num_Ross.append(ws[row][8].value)
        size_Ross.append(ws[row][6].value)
        program_Ross.append(ws[row][15].value)
        type_Ross.append(ws[row][4].value)
        installer_Ross.append(ws[row][9].value)
        due_date_Ross.append(ws[row][13].value)
        reviewer_Ross.append(ws[row][7].value)
        meter_Ross.append(ws[row][3].value)
        m=m+1

# Josh
p=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Josh B" and ws[row][10].value==None:
        IA_num_Josh.append(ws[row][1].value)
        case_num_Josh.append(ws[row][8].value)
        size_Josh.append(ws[row][6].value)
        program_Josh.append(ws[row][15].value)
        type_Josh.append(ws[row][4].value)
        installer_Josh.append(ws[row][9].value)
        due_date_Josh.append(ws[row][13].value)
        reviewer_Josh.append(ws[row][7].value)
        meter_Josh.append(ws[row][3].value)
        p=p+1

# Adan
q=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Adan A" and ws[row][10].value==None:
        IA_num_Adan.append(ws[row][1].value)
        case_num_Adan.append(ws[row][8].value)
        size_Adan.append(ws[row][6].value)
        program_Adan.append(ws[row][15].value)
        type_Adan.append(ws[row][4].value)
        installer_Adan.append(ws[row][9].value)
        due_date_Adan.append(ws[row][13].value)
        reviewer_Adan.append(ws[row][7].value)
        meter_Adan.append(ws[row][3].value)
        q=q+1

wb.close()
print("You can open the tracker now.")

# Count the total number of folders succesfully created
counter = 0
total = 0

print("Now populating Ethan's folder.")
# Ethan
for value in range(j):
    try:
        total += 1
        name = str(IA_num_Ethan[value])
        type = str(type_Ethan[value])
        full_name = r'G:\2021\21.00016\Reviews\EU\{}-EU'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ethan[value] < 40 and (program_Ethan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        elif size_Ethan[value] < 40 and (program_Ethan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Ethan[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Ethan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Ethan[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Ethan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue
print("And now Jose (if he has any).")
# Jose
for value in range(l):
    try:
        total += 1
        name = str(IA_num_Jose[value])
        type = str(type_Jose[value])
        full_name = r'G:\2021\21.00016\Reviews\JCN\{}-JCN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Jose[value] < 40 and (program_Jose[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        elif size_Jose[value] < 40 and (program_Jose[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Jose[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Jose[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Jose[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Jose[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

print("Now populating Nick's folder.")
# Nick
for value in range(i):
    try:
        total += 1
        name = str(IA_num_Nick[value])
        type = str(type_Nick[value])
        full_name = r'G:\2021\21.00016\Reviews\NC\{}-NC'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Nick[value] < 40 and (program_Nick[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        elif size_Nick[value] < 40 and (program_Nick[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Nick[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Nick[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Nick[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Nick[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

print("Now populating Ross's folder.")
# Ross
for value in range(m):
    try:
        total += 1
        name = str(IA_num_Ross[value])
        type = str(type_Ross[value])
        full_name = r'G:\2021\21.00016\Reviews\RK\{}-RK'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ross[value] < 40 and (program_Ross[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        elif size_Ross[value] < 40 and (program_Ross[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Ross[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Ross[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Ross[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Ross[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

print("Populating Joe's folder.")
# Joe
for value in range(k):
    no_DoS_flag = False
    try:
        total += 1
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2021\21.00016\Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Joe[value] < 40 and (program_Joe[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        elif size_Joe[value] < 40 and (program_Joe[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Joe[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Joe[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Joe[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Joe[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

print("Populating Josh's folder.")
# Josh
for value in range(p):
    no_DoS_flag = False
    try:
        total += 1
        name = str(IA_num_Josh[value])
        type = str(type_Josh[value])
        size = float(size_Josh[value])
        full_name = r'G:\2021\21.00016\Reviews\JB\{}-JB'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Josh[value] < 40 and (program_Josh[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        elif size_Josh[value] < 40 and (program_Josh[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Josh[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Josh[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Josh[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Josh[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

print("Populating Adan's folder.")
# Adan
for value in range(q):
    try:
        total += 1
        name = str(IA_num_Adan[value])
        type = str(type_Adan[value])
        full_name = r'G:\2021\21.00016\Reviews\AA\{}-AA'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Adan[value] < 40 and (program_Adan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        elif size_Adan[value] < 40 and (program_Adan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Adan[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Adan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Adan[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Adan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Existing DG
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "Address: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]") + "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                "\nCoordinates: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[3]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(5)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(5)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except Exception as e:
        print(f"Error {e} with {name}")
        continue

driver.close()
print("Folders are now all populated.")
print(f"{counter} folders successfully created out of a total of {total}!")
