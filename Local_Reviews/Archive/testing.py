# import calendar
# import glob
# import os
# import win32com.client as win32

# print(calendar.month_name[10])

# pathToDownloadsFolder = r'C:\Users\joe.nogosek\Documents\Local Reviews'
# pathToReportxls = os.path.join(pathToDownloadsFolder, str(glob.glob('report*')[0]))
# pathToReportxlsx = os.path.join(pathToReportxls, 'x')
#
# excel = win32.gencache.EnsureDispatch('Excel.Application')
# wb = excel.Workbooks.Open(pathToReportxls)
# wb.SaveAs(pathToReportxls+'x', FileFormat = 51) # 51 corresponds to the .xlsx extension file format. 56 would be for .xls
# wb.Close()
#
# os.remove(pathToReportxls)

# from simple_salesforce import Salesforce
# import requests
# import pandas as pd
# import csv
# from io import StringIO
# # Sign into Salesforce
# sf = Salesforce(username='joseph.h.nogosek@xcelenergy.com', password='airdoc@2Ee', security_token='ZCNHeFLqD2V0glkzD5CEvLpMO')
# # Set report details
# sf_org = 'https://xcelenergy.my.salesforce.com/'
# report_id = '00O4O0000043gZP'
# export_params = '?isdtp=p1&export=1&enc=UTF-8&xf=xls'
# # Download report
# sf_report_url = sf_org + report_id + export_params
# response = requests.get(sf_report_url, headers=sf.headers, cookies={'sid': sf.session_id})
# new_report = response.content.decode('utf-8')
# report_df = pd.read_csv(StringIO(new_report))

# from simple_salesforce import Salesforce, SFType, SalesforceLogin
# from pandas import DataFrame, read_csv
# import json
# from pprint import pprint as pp
#
# login = json.load(open('login.json'))
# username = login['joseph.h.nogosek@xcelenergy.com']
# password = login['airdoc@2Ee']
# token = login['ZCNHeFLqD2V0glkzD5CEvLpMO']
#
# session_id, isntance = SalesforceLogin(username=username, password=password, security_token=token, sandbox=False)
#
# print(session_id, instance)

# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 11/30/2021
# CLose tracker and make sure G:\ directory is green

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import glob
import win32com.client as win32
import calendar
import os
import warnings
import datetime
import numpy as np
import shutil
from copy import copy
import pandas as pd

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Define the path to the Completeness Review Tracker and your downloads folder
pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx'
# pathToDownloadsFolder = r'C:\Users\joe.nogosek\Downloads'
#
# # Define paths to report and convert to correct file format
# if not os.path.exists(os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))):
#     raise ValueError('No report file present in the Completeness Review Reports folder')
# else:
#     pathToReportxls = os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))
#     pathToReportxlsx = pathToReportxls + 'x'
#
#     excel = win32.gencache.EnsureDispatch('Excel.Application')
#     wb = excel.Workbooks.Open(pathToReportxls)
#     wb.SaveAs(pathToReportxlsx, FileFormat = 51) # 51 corresponds to the .xlsx extension file format. 56 would be for .xls
#     wb.Close()
#     os.remove(pathToReportxls)

pathToReportxlsx = r'C:\Users\joe.nogosek\Downloads\report1640254413564.xlsx'
# Load in the new report file just created and current tracker and set some values which will be useful
wbReport = load_workbook(pathToReportxlsx)
wbTracker = load_workbook(pathToTracker)

wsReport = wbReport.active
row_max_report = wsReport.max_row - 6

premise_report = []

for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=4, max_col=4, values_only=True):
    premise_report.append(row[0])

for premise in range(len(premise_report)):
    if type(premise_report[premise]) != type(None):
        premise_report[premise] = premise_report[premise][-11:]
        premise_report[premise] = premise_report[premise][:9]
        print(premise_report[premise])
