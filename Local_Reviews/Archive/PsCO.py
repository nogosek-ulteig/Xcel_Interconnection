from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import glob
import win32com.client as win32
import calendar
import os
import warnings
import datetime
import numpy as np
import shutil
from copy import copy
import pandas as pd
from decimal import Decimal
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Define the path to the Completeness Review Tracker and your downloads folder
pathToTracker = r'C:\Users\joe.nogosek\Documents\PsCO.xlsx'

wbTracker = load_workbook(pathToTracker)

wsTracker = wbTracker.active
row_max_a = 1075
row_max_k = 182

OID_a = []
OID_k = []
status = []
due_date_May = []
due_date_history = []
# Go through all the sheets in the tracker to get all the necessary information in lists
for row in wsTracker['A']:
    OID_a.append(row.value)
for row in wsTracker['K']:
    OID_k.append(row.value)
for row in wsTracker['B']:
    status.append(row.value)
for row in wsTracker['C']:
    due_date_history.append(row.value)
for row in wsTracker['L']:
    due_date_May.append(row.value)

# Number of times we have made it through
k = 1
for i in range(row_max_k-1):
    #if (OID_k[i] in OID_a) and (status[len(OID_a) - 1 - OID_a[::-1].index(OID_k[i])] == 'Rejected') and (due_date_May[i] != due_date_history[len(OID_a) - 1 - OID_a[::-1].index(OID_k[i])]):
        # Verify complete
        #print(f'Verify complete: {OID_k[i]}')
    #elif (OID_k[i] not in OID_a):
        # Brand new
        #print(f'Brand new: {OID_k[i]}')
    if (OID_k[i] in OID_a) and status[len(OID_a) - 1 - OID_a[::-1].index(OID_k[i])] != 'Rejected':
        print(f'Delete: {OID_k[i]}')
        wsTracker['M' + str(k)] = OID_k[i]
        k += 1

# Save the Excel file
wbTracker.save(pathToTracker)

wbTracker.close()
