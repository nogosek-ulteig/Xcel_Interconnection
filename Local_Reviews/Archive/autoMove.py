# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 12/3/2021

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import warnings

warnings.filterwarnings('ignore')

wb = load_workbook(r"G:\2021\21.00016\Reviews\Xcel CR Checklist_under40kW_metering - IA.xlsm", read_only=False, keep_vba=True)
ws = wb.active

ws['F146'] = 'Worked'

wb.save(r"G:\2021\21.00016\Reviews\Xcel CR Checklist_under40kW_metering - IA.xlsm")
