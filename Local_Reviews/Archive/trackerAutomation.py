# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 11/30/2021
# Close tracker and make sure G:\ directory is green

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import glob
import win32com.client as win32
import calendar
import os
import warnings
import datetime
import numpy as np
import shutil
from copy import copy
import pandas as pd
from decimal import Decimal
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys

twoFA = input("Enter six digit 2FA code: ")

driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe")
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element_by_xpath('//button[normalize-space()="Xcel Energy CORP credentials"]').click()

username = '239665'
user_box = driver.find_element_by_id('username')
user_box.send_keys(username)

password = 'airdoc1Ee'
pass_box = driver.find_element_by_id('password')
pass_box.send_keys(password)

sign_on_button = driver.find_element_by_css_selector("a[onclick^='postOk']").click()

passcode_box = driver.find_element_by_name('pf.pass')
passcode_box.send_keys('2846' + str(twoFA))

submit_button = driver.find_element_by_xpath("//button[contains(@onclick,'postOk')]").click()

reports_button = driver.find_element_by_css_selector("[title*='Reports Tab']").click()

time.sleep(5)
driver.find_element_by_css_selector("div[class='nameFieldContainer descrContainer']").click()

driver.find_element_by_css_selector("[title*='Export Details']").click()

driver.find_element_by_css_selector("[title*='Export']").click()
time.sleep(15)

# Please enter the ratio of reviews from this batch you want assigned to Nick (the rest will go to Ethan) as a percent (0.5 corresponding to 50%)
ratioNick = 0.5

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Define the path to the Completeness Review Tracker and your downloads folder
pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx'
pathToDownloadsFolder = r'C:\Users\joe.nogosek\Downloads'

# Define paths to report and convert to correct file format
if not os.path.exists(os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))):
    raise ValueError('No report file present in the Completeness Review Reports folder')
else:
    pathToReportxls = os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))
    pathToReportxlsx = pathToReportxls + 'x'

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(pathToReportxls)
    wb.SaveAs(pathToReportxlsx, FileFormat = 51) # 51 corresponds to the .xlsx extension file format. 56 would be for .xls
    wb.Close()
    os.remove(pathToReportxls)

# Load in the new report file just created and current tracker and set some values which will be useful
wbReport = load_workbook(pathToReportxlsx)
wbTracker = load_workbook(pathToTracker)

wsReport = wbReport.active
row_max_report = wsReport.max_row - 6

IAnums_tracker = []
existingMeters_tracker = []
dueDates_tracker = []
types_tracker = []
developers_tracker = []
history_tracker = []
lineDistances_tracker = []
sizes_tracker = []
# Go through all the sheets in the tracker to get all the necessary information in lists
for sheet in wbTracker.sheetnames:
    wsTracker = wbTracker[sheet]
    # Get all the IA numbers from this sheet in the book
    for row in wsTracker['B']:
        if row.row == 1:
            pass
        else:
            IAnums_tracker.append(row.value)
            #print(f"IA Num: {row.value}")
    # Get all the existing meter types from this sheet in the book
    for row in wsTracker['D']:
        if row.row == 1:
            pass
        else:
            existingMeters_tracker.append(row.value)
            #print(f"Existing Meter: {row.value}")
    # Get all the due dates from this sheet in the book
    for row in wsTracker['N']:
        if row.row == 1:
            pass
        else:
            dueDates_tracker.append(row.value)
            #print(f"Due date: {row.value}")
    # Get all the review types (ESS or not) from this sheet in the book
    for row in wsTracker['E']:
        if row.row == 1:
            pass
        else:
            types_tracker.append(row.value)
            #print(f"Type: {row.value}")
    # Get all the developers from this sheet in the book
    for row in wsTracker['J']:
        if row.row == 1:
            pass
        else:
            developers_tracker.append(row.value)
            #print(f"Developer: {row.value}")
    # Get all the approval/rejection from this sheet in the book
    for row in wsTracker['L']:
        if row.row == 1:
            pass
        else:
            history_tracker.append(row.value)
            #print(f"History: {row.value}")
    # Get all the line distances from this sheet in the book
    for row in wsTracker['F']:
        if row.row == 1:
            pass
        else:
            lineDistances_tracker.append(row.value)
            #print(f"Line Distance: {row.value}")
    # Get all the system sizes from this sheet in the book
    for row in wsTracker['G']:
        if row.row == 1:
            pass
        else:
            sizes_tracker.append(row.value)
            #print(f"Size: {row.value}")

# Get all the information from the report in lists
# Initialize necessary lists
IAnums_report = []
caseNums_report = []
developers_report = []
dueDates_report = []
statuses_report = []
sizes_report = []
types_report = []
premise_report = []
programType_report = []

# Get the IA numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=5, max_col=5, values_only=True):
    IAnums_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Get the premise numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=6, max_col=6, values_only=True):
    premise_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Extract just the premise number
for premise in range(len(premise_report)):
    if type(premise_report[premise]) != type(None):
        premise_report[premise] = premise_report[premise][-11:]
        premise_report[premise] = premise_report[premise][:9]
        if premise_report[premise] == 'der Premi':
            premise_report[premise] = 'New Construction'
# Get the case numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=7, max_col=7, values_only=True):
    caseNums_report.append(row[0])
    #print(f"Case num: {row[0]}")
# Get all the developers from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=8, max_col=8, values_only=True):
    developers_report.append(row[0])
    #print(f"Developer: {row[0]}")
# Get the due dates from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=9, max_col=9, values_only=True):
    dueDates_report.append(row[0])
    #print(f"Due Date: {row[0]}")
# Get all the application statuses from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=12, max_col=12, values_only=True):
    statuses_report.append(row[0])
    #print(f"Status: {row[0]}")
# Get the application sizes from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=13, max_col=13, values_only=True):
    sizes_report.append(row[0])
    #print(f"Size: {row[0]}")
# Get all the application types (ESS or not) from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=14, max_col=14, values_only=True):
    types_report.append(row[0])
    #print(f"Type: {row[0]}")
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=10, max_col=10, values_only=True):
    programType_report.append(row[0])
    #print(f"Type: {row[0]}")

# Create styles
redFill = PatternFill(start_color='FFFF0000', end_color='FF0000', fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF', end_color='000000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFF00', fill_type='solid')
greyFill = PatternFill(start_color='FFF2F2F2', end_color='F2F2F2', fill_type='solid')
greenFill = PatternFill(start_color = 'FFA9D08E', end_color = 'A9D08E', fill_type='solid')

row_max_tracker = wsTracker.max_row
wsTracker['O' + str(row_max_tracker)] = ''
next_row_tracker = row_max_tracker + 1

ratio = 1 - Decimal(ratioNick)

amountNewReviews = 0
# Count the total number of reviews to be added
for w in range(row_max_report-1):
    if statuses_report[w] == 'Initiate Application':
        if (IAnums_report[w] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])] == 'Rejected') and (dueDates_report[w] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])]):
            amountNewReviews += 1
        elif (IAnums_report[w] not in IAnums_tracker):
            amountNewReviews += 1

# Number of times we have made it through
k = 1
flipFlop = 0
for i in range(row_max_report-1):
    if statuses_report[i] == 'Initiate Application':
        if (IAnums_report[i] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Rejected') and (dueDates_report[i] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])]):
            wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)].fill = whiteFill
            wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
            wsTracker['C' + str(next_row_tracker)].fill = whiteFill
            wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
            wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
            wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
            if sizes_report[i] >= 20:
                 wsTracker['F' + str(next_row_tracker)].fill = redFill
            else:
                 wsTracker['F' + str(next_row_tracker)].fill = greyFill
            wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
            if sizes_report[i] < 40:
                 wsTracker['G' + str(next_row_tracker)].fill = yellowFill
            else:
                wsTracker['G' + str(next_row_tracker)].fill = whiteFill
            wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
            wsTracker['H' + str(next_row_tracker)].fill = whiteFill
            wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
            wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
            wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
            wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
            wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
            wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
            wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
            wsTracker['P' + str(next_row_tracker)]._style = copy(wsTracker['P' + str(row_max_tracker)]._style)
            dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
            wsTracker['A' + str(next_row_tracker)] = dueDate - datetime.timedelta(days=2)
            wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
            wsTracker['C' + str(next_row_tracker)] = premise_report[i]
            wsTracker['D' + str(next_row_tracker)] = 'Verify Complete'
            wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
            if types_report[i] == 'Yes':
                #print('Reassign ' + str(IAnums_report[i]) + ' to metering (ESS)')
                wsTracker['E' + str(next_row_tracker)] = 'ESS'
            if lineDistances_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Complete':
                wsTracker['F' + str(next_row_tracker)] = 'Complete'
            wsTracker['G' + str(next_row_tracker)] = sizes_report[i]
            if sizes_report[i] < 40:
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
            else:
                #print('Reassign ' + str(IAnums_report[i]) + ' to metering (Larger than 40 kW)')
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
            if amountNewReviews*ratio < k:
                wsTracker['H' + str(next_row_tracker)] = 'Nick C'
            else:
                wsTracker['H' + str(next_row_tracker)] = 'Ethan U'
            wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
            wsTracker['J' + str(next_row_tracker)] = developers_report[i]
            wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
            wsTracker['P' + str(next_row_tracker)] = programType_report[i]
            next_row_tracker = next_row_tracker + 1
            k += 1
        elif (IAnums_report[i] not in IAnums_tracker):
            wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)].fill = whiteFill
            wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
            wsTracker['C' + str(next_row_tracker)].fill = whiteFill
            wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
            wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
            wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
            if sizes_report[i] >= 20:
                 wsTracker['F' + str(next_row_tracker)].fill = redFill
            else:
                 wsTracker['F' + str(next_row_tracker)].fill = greyFill
            wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
            if sizes_report[i] < 40:
                 wsTracker['G' + str(next_row_tracker)].fill = yellowFill
            else:
                wsTracker['G' + str(next_row_tracker)].fill = whiteFill
            wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
            wsTracker['H' + str(next_row_tracker)].fill = whiteFill
            wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
            wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
            wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
            wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
            wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
            wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
            wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
            wsTracker['P' + str(next_row_tracker)]._style = copy(wsTracker['P' + str(row_max_tracker)]._style)
            dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
            wsTracker['A' + str(next_row_tracker)] = dueDate - datetime.timedelta(days=9)
            wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
            wsTracker['C' + str(next_row_tracker)] = premise_report[i]
            if types_report[i] == 'Yes' or sizes_report[i] >= 40:
                wsTracker['D' + str(next_row_tracker)] = 'N/A'
                wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
            else:
                wsTracker['D' + str(next_row_tracker)] = '2S'
                wsTracker['D' + str(next_row_tracker)].font = Font(bold=True)
            if types_report[i] == 'Yes':
                # Put the IA that is being processed here and pull it from the folder name
                driver.find_element_by_id('phSearchInput').send_keys(str(IAnums_report[i]))
                driver.find_element_by_id('phSearchButton').click()

                # Click the top result from case number selector
                driver.find_element_by_css_selector("a[target='_top']").click()
                time.sleep(5)

                # Change the name of Area and meter engineer (if applicable)
                driver.find_element_by_css_selector("[title='Edit']").click()
                driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                if flipFlop % 2 == 0:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Shaun Ly")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Shaun Ly (ESS)')
                else:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Chi Vang")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Chi Vang (ESS)')
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                time.sleep(5)
                flipFlop += 1
                wsTracker['E' + str(next_row_tracker)] = 'ESS'
            wsTracker['G' + str(next_row_tracker)] = sizes_report[i]
            if sizes_report[i] < 40:
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
            else:
                # Put the IA that is being processed here and pull it from the folder name
                driver.find_element_by_id('phSearchInput').send_keys(str(IAnums_report[i]))
                driver.find_element_by_id('phSearchButton').click()

                # Click the top result from case number selector
                driver.find_element_by_css_selector("a[target='_top']").click()
                time.sleep(5)

                # Change the name of Area and meter engineer (if applicable)
                driver.find_element_by_css_selector("[title='Edit']").click()
                driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                if flipFlop % 2 == 0:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Shaun Ly")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Shaun Ly (Larger than 40 kW)')
                else:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Chi Vang")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Chi Vang (Larger than 40 kW)')
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                time.sleep(5)
                flipFlop += 1
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
            if amountNewReviews*ratio < k:
                wsTracker['H' + str(next_row_tracker)] = 'Nick C'
            else:
                wsTracker['H' + str(next_row_tracker)] = 'Ethan U'
            wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
            wsTracker['J' + str(next_row_tracker)] = developers_report[i]
            wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
            wsTracker['P' + str(next_row_tracker)] = programType_report[i]
            next_row_tracker = next_row_tracker + 1
            k += 1

time_str = datetime.datetime.today().strftime("%I:%M %p")
date = datetime.date.strftime(datetime.date.today(), "%m/%d/%Y")
wsTracker['O' + str(next_row_tracker - 1)] = 'Tracker updated as of ' + str(time_str) + ' CST on ' + str(date)
# Save the Excel file
wbTracker.save(pathToTracker)

wbTracker.close()
wbReport.close()

# Delete the report so the script is ready for next time
os.remove(pathToReportxlsx)

# Now get the newly added reviews ready for me to complete
IA_num_Nick = []
case_num_Nick = []
size_Nick = []
program_Nick = []
type_Nick = []
verify_Nick = []

IA_num_Ethan = []
case_num_Ethan = []
size_Ethan = []
program_Ethan = []
type_Ethan = []
verify_Ethan = []

IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
verify_Joe = []

IA_num_Jose = []
case_num_Jose = []
size_Jose = []
program_Jose = []
type_Jose = []
verify_Jose = []

IA_num_Ross = []
case_num_Ross = []
size_Ross = []
program_Ross = []
type_Ross = []
verify_Ross = []

wb = load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2022']

# Nick
i=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Nick C" and ws[row][10].value==None:
        IA_num_Nick.append(ws[row][1].value)
        case_num_Nick.append(ws[row][8].value)
        size_Nick.append(ws[row][6].value)
        program_Nick.append(ws[row][15].value)
        type_Nick.append(ws[row][4].value)
        verify_Nick.append(ws[row][3].value)
        i=i+1

# Ethan
j=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ethan U" and ws[row][10].value==None:
        IA_num_Ethan.append(ws[row][1].value)
        case_num_Ethan.append(ws[row][8].value)
        size_Ethan.append(ws[row][6].value)
        program_Ethan.append(ws[row][15].value)
        type_Ethan.append(ws[row][4].value)
        verify_Ethan.append(ws[row][3].value)
        j=j+1

# Joe
k=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        verify_Joe.append(ws[row][3].value)
        k=k+1

# Jose
l=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Jose CN" and ws[row][10].value==None:
        IA_num_Jose.append(ws[row][1].value)
        case_num_Jose.append(ws[row][8].value)
        size_Jose.append(ws[row][6].value)
        program_Jose.append(ws[row][15].value)
        type_Jose.append(ws[row][4].value)
        verify_Jose.append(ws[row][3].value)
        l=l+1

# Ross
m=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ross K" and ws[row][10].value==None:
        IA_num_Ross.append(ws[row][1].value)
        case_num_Ross.append(ws[row][8].value)
        size_Ross.append(ws[row][6].value)
        program_Ross.append(ws[row][15].value)
        type_Ross.append(ws[row][4].value)
        verify_Ross.append(ws[row][3].value)
        m=m+1

wb.close()
print("You can open the tracker now.")

print("Populating Joe's folder.")
# Joe
for value in range(k):
    try:
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2021\21.00016\Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Joe[value] < 40 and (program_Joe[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        elif size_Joe[value] < 40 and (program_Joe[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # If it's a verify complete
            first_action = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
            if first_action == 'Reopen Action (Review Missing Info)':
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Not verify complete
            else:
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/th[1]/a[1]").click()
                time.sleep(1)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
    except:
            continue

print("Now populating Nick's folder.")
# Nick
for value in range(i):
    try:
        name = str(IA_num_Nick[value])
        type = str(type_Nick[value])
        full_name = r'G:\2021\21.00016\Reviews\NC\{}-NC'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Nick[value] < 40 and (program_Nick[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        elif size_Nick[value] < 40 and (program_Nick[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel Completeness Review Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # If it's a verify complete
            first_action = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
            if first_action == 'Reopen Action (Review Missing Info)':
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Not verify complete
            else:
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
    except:
        continue
print("Now populating Ethan's folder.")
# Ethan
for value in range(j):
    try:
        name = str(IA_num_Ethan[value])
        type = str(type_Ethan[value])
        full_name = r'G:\2021\21.00016\Reviews\EU\{}-EU'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ethan[value] < 40 and (program_Ethan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        elif size_Ethan[value] < 40 and (program_Ethan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # If it's a verify complete
            first_action = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
            if first_action == 'Reopen Action (Review Missing Info)':
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Not verify complete
            else:
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
    except:
        continue
print("And now Jose (if he has any).")
# Jose
for value in range(l):
    try:
        name = str(IA_num_Jose[value])
        type = str(type_Jose[value])
        full_name = r'G:\2021\21.00016\Reviews\JCN\{}-JCN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Jose[value] < 40 and (program_Jose[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        elif size_Jose[value] < 40 and (program_Jose[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # If it's a verify complete
            first_action = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
            if first_action == 'Reopen Action (Review Missing Info)':
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(1)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Not verify complete
            else:
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/th[1]/a[1]").click()
                time.sleep(1)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
    except:
        continue
print("Lastly populating Ross's folder.")
# Ross
for value in range(m):
    try:
        name = str(IA_num_Ross[value])
        type = str(type_Ross[value])
        full_name = r'G:\2021\21.00016\Reviews\RK\{}-RK'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ross[value] < 40 and (program_Ross[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        elif size_Ross[value] < 40 and (program_Ross[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # If it's a verify complete
            first_action = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
            if first_action == 'Reopen Action (Review Missing Info)':
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Not verify complete
            else:
                # Oneline first
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

                # Now site plan
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/th[1]/a[1]").click()
                time.sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                time.sleep(5)
                driver.back()
                time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
    except:
        continue

driver.close()
print("Folders are now all populated.")
