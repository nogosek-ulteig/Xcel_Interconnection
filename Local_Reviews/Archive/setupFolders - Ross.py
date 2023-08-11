#Made by Joe Nogosek, joe.nogosek@ulteig.com
#10/26/2021
#NOTE: Tracker must be closed on your local computer!

import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
import shutil
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import glob

username = ''
password = ''
pin = ''

twoFA = input("Enter six digit 2FA code: ")

driver = webdriver.Chrome(executable_path=r"C:\Users\ross.kirby\Documents\Python\chromedriver.exe")
driver.maximize_window()
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element_by_xpath('//button[normalize-space()="Xcel Energy CORP credentials"]').click()

user_box = driver.find_element_by_id('username')
user_box.send_keys(username)

pass_box = driver.find_element_by_id('password')
pass_box.send_keys(password)

sign_on_button = driver.find_element_by_css_selector("a[onclick^='postOk']").click()

passcode_box = driver.find_element_by_name('pf.pass')
passcode_box.send_keys(pin + str(twoFA))

submit_button = driver.find_element_by_xpath("//button[contains(@onclick,'postOk')]").click()

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Now get the newly added reviews ready for me to complete
IA_num_Nick = []
case_num_Nick = []
size_Nick = []
program_Nick = []
type_Nick = []
installer_Nick = []

IA_num_Ethan = []
case_num_Ethan = []
size_Ethan = []
program_Ethan = []
type_Ethan = []
installer_Ethan = []

IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
installer_Joe = []

IA_num_Jose = []
case_num_Jose = []
size_Jose = []
program_Jose = []
type_Jose = []
installer_Jose = []

IA_num_Ross = []
case_num_Ross = []
size_Ross = []
program_Ross = []
type_Ross = []
installer_Ross = []

IA_num_Josh = []
case_num_Josh = []
size_Josh = []
program_Josh = []
type_Josh = []
installer_Josh = []

IA_num_Adan = []
case_num_Adan = []
size_Adan = []
program_Adan = []
type_Adan = []
installer_Adan = []

wb = load_workbook(r'C:\Users\ross.kirby\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2022']

# Nick
i=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Nick C" and ws[row][10].value==None:
        IA_num_Nick.append(ws[row][1].value)
        case_num_Nick.append(ws[row][8].value)
        size_Nick.append(ws[row][6].value)
        program_Nick.append(ws[row][15].value)
        type_Nick.append(ws[row][4].value)
        installer_Nick.append(ws[row][9].value)
        i=i+1

# Ethan
j=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ethan U" and ws[row][10].value==None:
        IA_num_Ethan.append(ws[row][1].value)
        case_num_Ethan.append(ws[row][8].value)
        size_Ethan.append(ws[row][6].value)
        program_Ethan.append(ws[row][15].value)
        type_Ethan.append(ws[row][4].value)
        installer_Ethan.append(ws[row][9].value)
        j=j+1

# Joe
k=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        installer_Joe.append(ws[row][9].value)
        k=k+1

# Jose
l=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Jose CN" and ws[row][10].value==None:
        IA_num_Jose.append(ws[row][1].value)
        case_num_Jose.append(ws[row][8].value)
        size_Jose.append(ws[row][6].value)
        program_Jose.append(ws[row][15].value)
        type_Jose.append(ws[row][4].value)
        installer_Jose.append(ws[row][9].value)
        l=l+1

# Ross
m=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ross K" and ws[row][10].value==None:
        IA_num_Ross.append(ws[row][1].value)
        case_num_Ross.append(ws[row][8].value)
        size_Ross.append(ws[row][6].value)
        program_Ross.append(ws[row][15].value)
        type_Ross.append(ws[row][4].value)
        installer_Ross.append(ws[row][9].value)
        m=m+1

# Josh
p=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Josh B" and ws[row][10].value==None:
        IA_num_Josh.append(ws[row][1].value)
        case_num_Josh.append(ws[row][8].value)
        size_Josh.append(ws[row][6].value)
        program_Josh.append(ws[row][15].value)
        type_Josh.append(ws[row][4].value)
        installer_Josh.append(ws[row][9].value)
        p=p+1

# Adan
q=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Adan A" and ws[row][10].value==None:
        IA_num_Adan.append(ws[row][1].value)
        case_num_Adan.append(ws[row][8].value)
        size_Adan.append(ws[row][6].value)
        program_Adan.append(ws[row][15].value)
        type_Adan.append(ws[row][4].value)
        installer_Adan.append(ws[row][9].value)
        q=q+1

wb.close()
print("You can open the tracker now.")

# Count the total number of folders succesfully created
counter = 0
total = 0

print("Populating Josh's folder.")
# Josh
for value in range(p):
    try:
        total += 1
        name = str(IA_num_Josh[value])
        type = str(type_Josh[value])
        full_name = r'G:\2021\21.00016\Reviews\JB\{}-JB'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Josh[value] < 40 and (program_Josh[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        elif size_Josh[value] < 40 and (program_Josh[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Josh[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Josh[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Adan's folder.")
# Adan
for value in range(q):
    try:
        total += 1
        name = str(IA_num_Adan[value])
        type = str(type_Adan[value])
        full_name = r'G:\2021\21.00016\Reviews\AA\{}-AA'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Adan[value] < 40 and (program_Adan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        elif size_Adan[value] < 40 and (program_Adan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Adan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Adan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Joe's folder.")
# Joe
for value in range(k):
    try:
        total += 1
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2021\21.00016\Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Joe[value] < 40 and (program_Joe[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        elif size_Joe[value] < 40 and (program_Joe[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Joe[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Joe[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Now populating Nick's folder.")
# Nick
for value in range(i):
    try:
        total += 1
        name = str(IA_num_Nick[value])
        type = str(type_Nick[value])
        full_name = r'G:\2021\21.00016\Reviews\NC\{}-NC'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Nick[value] < 40 and (program_Nick[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        elif size_Nick[value] < 40 and (program_Nick[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Nick[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Nick[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ethan's folder.")
# Ethan
for value in range(j):
    try:
        total += 1
        name = str(IA_num_Ethan[value])
        type = str(type_Ethan[value])
        full_name = r'G:\2021\21.00016\Reviews\EU\{}-EU'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ethan[value] < 40 and (program_Ethan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        elif size_Ethan[value] < 40 and (program_Ethan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Ethan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Ethan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("And now Jose (if he has any).")
# Jose
for value in range(l):
    try:
        total += 1
        name = str(IA_num_Jose[value])
        type = str(type_Jose[value])
        full_name = r'G:\2021\21.00016\Reviews\JCN\{}-JCN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Jose[value] < 40 and (program_Jose[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        elif size_Jose[value] < 40 and (program_Jose[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Jose[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Jose[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ross's folder.")
# Ross
for value in range(m):
    try:
        total += 1
        name = str(IA_num_Ross[value])
        type = str(type_Ross[value])
        full_name = r'G:\2021\21.00016\Reviews\RK\{}-RK'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ross[value] < 40 and (program_Ross[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        elif size_Ross[value] < 40 and (program_Ross[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\ross.kirby\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)

            if size_Ross[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

            elif size_Ross[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

driver.close()
print("Folders are now all populated.")
print(f"{counter} folders successfully created out of a total of {total}!")
