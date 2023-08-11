# Made by Joe Nogosek
# 3/30/2022

import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import shutil

initials = ["JN", "RK", "NC", "EU", "JCN", "JB", "AA"]
pathToReviews = r'G:\2021\21.00016\Reviews'
type = ""
case_num = ""
type_SR = ""
under = ""
over = ""

for name in initials:
    print(f"Working on {name}.")
    pathToFolder = os.path.join(pathToReviews, name)
    curFolders = os.listdir(pathToFolder)
    for folder in curFolders:
        if folder[:2] == "IA":
            IA_num = folder[:8]
            pathToIA = os.path.join(pathToFolder, folder)
            full_name = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}'.format(name, IA_num, name)
            try:
                files = os.listdir(pathToIA)
                for file in files:
                    if file.endswith(".xlsm"):
                        under = str(file[18:23])
                        over = str(file[35:39])
                        type_SR = str(file[37:39])
                        pathToChecklist = os.path.join(pathToIA, file)
                        wb = load_workbook(pathToChecklist, keep_vba=True)
                        ws = wb['DER Study']
                        type = str(ws['I2'].value)
                        installer = str(ws['I3'].value)
                        case_num = str(ws['C1'].value)
                        if under == "under":
                            name_oneline_SF = ws['D13'].value
                            case_num_oneline_SF = ws['D14'].value
                            aggregate_SF = ws['D26'].value
                            inverter_rating_SF = ws['D28'].value
                            inverter_type_SF = ws['D29'].value
                            name_siteplan_SF = ws['D67'].value
                            address_SF = ws['D68'].value
                            installer_SF = ws['D72'].value
                            case_num_siteplan_SF = ws['D73'].value
                            feeder_num = ws['D137'].value
                            existing = ws['D34'].value
                            ess = ws['D35'].value
                        elif over == "over":
                            name_oneline_SF = ws['D12'].value
                            case_num_oneline_SF = ws['D13'].value
                            aggregate_SF = ws['D27'].value
                            inverter_rating_SF = ws['D29'].value
                            inverter_type_SF = ws['D30'].value
                            name_siteplan_SF = ws['D66'].value
                            address_SF = ws['D67'].value
                            coordinates_SF = ws['D81'].value
                            case_num_siteplan_SF = ws['D71'].value
                            feeder_num = ws['D126'].value
                        os.remove(pathToChecklist)
                    else:
                        continue
            except:
                pass
            if type_SR == "SR" and under == "under":
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
                new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(IA_num))
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = IA_num
                ws['I2'].value = type
                ws['I3'].value = installer
                ws['C1'].value = case_num
                ws['D13'].value = name_oneline_SF
                ws['D14'].value = case_num_oneline_SF
                ws['D26'].value = aggregate_SF
                ws['D28'].value = inverter_rating_SF
                ws['D29'].value = inverter_type_SF
                ws['D67'].value = name_siteplan_SF
                ws['D68'].value = address_SF
                ws['D72'].value = installer_SF
                ws['D73'].value = case_num_siteplan_SF
                ws['D137'].value = feeder_num
                ws['D34'].value = existing
                ws['D35'].value = ess
                wb.save(new_file)
            elif type_SR != "SR" and under == "under":
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
                new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(IA_num))
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = IA_num
                ws['I2'].value = type
                ws['I3'].value = installer
                ws['C1'].value = case_num
                ws['D13'].value = name_oneline_SF
                ws['D14'].value = case_num_oneline_SF
                ws['D26'].value = aggregate_SF
                ws['D28'].value = inverter_rating_SF
                ws['D29'].value = inverter_type_SF
                ws['D67'].value = name_siteplan_SF
                ws['D68'].value = address_SF
                ws['D72'].value = installer_SF
                ws['D73'].value = case_num_siteplan_SF
                ws['D137'].value = feeder_num
                ws['D34'].value = existing
                ws['D35'].value = ess
                wb.save(new_file)
            elif over == "over":
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
                new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(IA_num))
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = IA_num
                ws['I2'].value = type
                ws['I3'].value = installer
                ws['C1'].value = case_num
                ws['D12'].value = name_oneline_SF
                ws['D13'].value = case_num_oneline_SF
                ws['D27'].value = aggregate_SF
                ws['D29'].value = inverter_rating_SF
                ws['D30'].value = inverter_type_SF
                ws['D66'].value = name_siteplan_SF
                ws['D67'].value = address_SF
                ws['D71'].value = case_num_siteplan_SF
                ws['D81'].value = coordinates_SF
                ws['D126'].value = feeder_num
                wb.save(new_file)
            type = ""
            case_num = ""
            type_SR = ""
            under = ""
            over = ""
