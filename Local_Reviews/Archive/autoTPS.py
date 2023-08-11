# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 10/26/2022

import openpyxl
from openpyxl import load_workbook
import os, os.path
from os.path import join
import sys
import numpy as np
import shutil
import warnings
import glob
import zipfile
import formulas
import win32com.client as win32
import datetime

warnings.filterwarnings("ignore")

list_of_files = glob.glob(r"C:\Users\joe.nogosek\Downloads\*.zip") # * means all if need specific format then *.csv
latest_file = max(list_of_files, key=os.path.getctime)

with zipfile.ZipFile(latest_file, 'r') as zip_ref:
    zip_ref.extractall(r"C:\Users\joe.nogosek\Downloads")

pathToTPS = r"C:\Users\joe.nogosek\Downloads\TPS.xlsm"
pathToHelper = r"C:\Users\joe.nogosek\Downloads\Helper.xlsx"
pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\MN Completeness Review Tracker.xlsx'

wbHelper = openpyxl.Workbook()
wsHelper = wbHelper.active

wbTracker = load_workbook(pathToTracker)
wsTracker = wbTracker['2023']

caseNums = []
feeder = []
IA = []
numApps = 0
for row in range(wsTracker.max_row):
    if row == 0 or row == 1:
        pass
    else:
        if (wsTracker['L' + str(row)].value == 'CA' or wsTracker['L' + str(row)].value == 'Approved') and (wsTracker['R' + str(row)].value == None):
            caseNums.append(wsTracker['I' + str(row)].value)
            feeder.append(wsTracker['A' + str(row)].value)
            IA.append(wsTracker['B' + str(row)].value)
            numApps += 1

print(f"The number of applications for TPS: {numApps}")
for i in range(numApps):
    wsHelper['A' + str(i + 1)].value = caseNums[i]
    wsHelper['B' + str(i + 1)].value = feeder[i]
    wsHelper['C' + str(i + 1)].value = IA[i]

wbHelper.save(pathToHelper)

xl=win32.Dispatch("Excel.Application")
book = xl.Workbooks.Open(os.path.abspath(pathToTPS), ReadOnly=1)
file = 'NSPM_DER_Screening_Tool_v3.01_(Confidential_Internal_Only).xlsm'
xl.Run("TPS")
book.Close(SaveChanges=False)
xl.Quit()
del xl

wbHelper = load_workbook(pathToHelper, data_only = True)
wsHelper = wbHelper.active

date = datetime.date.strftime(datetime.date.today(), "%m/%d/%Y")

counter = 1
for row in range(wsTracker.max_row):
    if row == 0 or row == 1:
        pass
    else:
        if (wsTracker['L' + str(row)].value == 'CA' or wsTracker['L' + str(row)].value == 'Approved') and (wsTracker['R' + str(row)].value == None and wsTracker['M' + str(row)].value != None):
            wsTracker['R' + str(row)].value = wsHelper['F' + str(counter)].value
            if wsHelper['F' + str(counter)].value == "Fail":
                wsTracker['S' + str(row)].value = date
            else:
                wsTracker['S' + str(row)].value = 'x'
            wsTracker['T' + str(row)].value = 'JN'
            counter += 1

wbTracker.save(pathToTracker)

wbTracker.close()
os.remove(pathToHelper)
os.remove(pathToTPS)
os.remove(latest_file)

print("Finished successfully :)")
