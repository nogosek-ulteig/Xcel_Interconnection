#Made by Joe Nogosek, joe.nogosek@ulteig.com
#10/26/2021
#NOTE: Tracker must be closed on your local computer!

import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
import shutil
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import glob

twoFA = input("Enter six digit 2FA code: ")

driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe")
driver.maximize_window()
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element_by_xpath('//button[normalize-space()="Xcel Energy CORP credentials"]').click()

username = '239665'
user_box = driver.find_element_by_id('username')
user_box.send_keys(username)

password = 'airdoc1Ee'
pass_box = driver.find_element_by_id('password')
pass_box.send_keys(password)

sign_on_button = driver.find_element_by_css_selector("a[onclick^='postOk']").click()

passcode_box = driver.find_element_by_name('pf.pass')
passcode_box.send_keys('2846' + str(twoFA))

submit_button = driver.find_element_by_xpath("//button[contains(@onclick,'postOk')]").click()

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

wb = load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2022']

# Now get the newly added reviews ready for me to complete
IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
verify_Joe = []

# Joe
k=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        verify_Joe.append(ws[row][3].value)
        k=k+1

wb.close()
print("You can open the tracker now.")

# Count the total number of folders succesfully created
counter = 0

print("Populating Joe's folder.")
# Joe
for value in range(k):
    try:
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2021\21.00016\Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Joe[value] < 40 and (program_Joe[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        elif size_Joe[value] < 40 and (program_Joe[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-IA{}.xlsm'.format(name[-5:]))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-IA.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            for count in range(2,25):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(3)
                     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                     time.sleep(5)
                     driver.back()
                     time.sleep(1)
                     z+=1
                     if z == 2:
                         break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
        counter += 1
    except:
            continue

driver.close()
print("Folders are now all populated.")
print(f"{counter} folders successfully created!")
