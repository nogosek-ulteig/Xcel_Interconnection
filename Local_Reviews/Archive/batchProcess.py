# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 11/10/2022

import openpyxl
from openpyxl import load_workbook
import os, os.path
from os.path import join
import sys
import numpy as np
import shutil
import warnings
import glob
import win32com.client as win32
import datetime
import time

warnings.filterwarnings("ignore")

pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx'

wb = load_workbook(pathToTracker)
ws = wb['2023']

# Now get all the necessary information from the 2022 tab
IA_num = []
reviewer = []

numReviews = 0
for row in range(3,ws.max_row+1):
    if ws[row][11].value != None and ws[row][12].value == None:
        IA_num.append(ws[row][1].value)
        reviewer.append(ws[row][7].value)
        numReviews += 1

wb.close()
print("Good to open tracker.\n")

xl=win32.Dispatch("Excel.Application")
# tracker = xl.Workbooks.Open(os.path.abspath(pathToTracker))

pathToReadyQC = r'G:\2021\21.00016\Reviews\Ready_for_QC'
files = []
i=0
for i in range(numReviews):
    pause = input("Press enter when ready to process the next one.")
    print(IA_num[i])
    if reviewer[i] == "Joe Nogo":
        files.append(IA_num[i] + "-JN")
    elif reviewer[i] == "Ross K":
        files.append(IA_num[i] + "-RK")
    elif reviewer[i] == "Nick C":
        files.append(IA_num[i] + "-NC")
    pathToFile = os.path.join(pathToReadyQC,files[i])
    list_of_files = glob.glob(os.path.join(pathToFile,"*.xlsm")) # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    book = xl.Workbooks.Open(os.path.abspath(latest_file))
    try:
        xl.Run("autoSF")
    except:
        pass
    time.sleep(3)

del xl
print("This script is finished. Make sure to help all the open Chrome windows finish.")
