# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 3/22/2022

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import sys
import os

IA = input("Enter the IA number of the app to process: ")
initials = input("Enter the initials of the person who did the review: ")
SRorNo = input("Is this application solar rewards? (y/n): ")
twoFA = input("Enter six digit 2FA code: ")
pathToReadyForQC = r'G:\2021\21.00016\Reviews\Ready_for_QC'
if SRorNo == 'n':
    pathToChecklist = os.path.join(pathToReadyForQC, 'IA' + IA + '-' + initials, 'Xcel CR Checklist_under40kW_metering - IA' + IA + '.xlsm')
elif SRorNo == 'y':
    pathToChecklist = os.path.join(pathToReadyForQC, 'IA' + IA + '-' + initials, 'Xcel CR Checklist_under40kW_metering_SR - IA' + IA + '.xlsm')
# if SRorNo == 'N':
#     pathToChecklist = os.path.join(pathToReadyForQC, 'IA' + IA + '-' + initials, 'Xcel_CR_Checklist_under40kW_metering-IA' + IA + '.xlsm')
# elif SRorNo == 'Y':
#     pathToChecklist = os.path.join(pathToReadyForQC, 'IA' + IA + '-' + initials, 'Xcel_CR_Checklist_under40kW_metering_SR-IA' + IA + '.xlsm')
overUnder = "under"
wb = load_workbook(pathToChecklist, data_only = True)
ws = wb.active

driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe")
driver.maximize_window()
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element_by_xpath('//button[normalize-space()="Xcel Energy CORP credentials"]').click()

username = '239665'
user_box = driver.find_element_by_id('username')
user_box.send_keys(username)

password = 'airdoc1Ee'
pass_box = driver.find_element_by_id('password')
pass_box.send_keys(password)

sign_on_button = driver.find_element_by_css_selector("a[onclick^='postOk']").click()

passcode_box = driver.find_element_by_name('pf.pass')
passcode_box.send_keys('2846' + str(twoFA))

submit_button = driver.find_element_by_xpath("//button[contains(@onclick,'postOk')]").click()

# Put the IA that is being processed here and pull it from the folder name
driver.find_element_by_id('phSearchInput').send_keys(IA)
driver.find_element_by_id('phSearchButton').click()

# Click the top result from case number selector
driver.find_element_by_css_selector("a[target='_top']").click()
time.sleep(5)

# Change the name of Area and meter engineer (if applicable)
driver.find_element_by_css_selector("[title='Edit']").click()
# If not Verify complete or ESS, then do both Area and Meter otherwise just do Area
driver.find_element_by_css_selector("[title='Area Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
driver.find_element_by_css_selector("[title='Area Engineer Approver']").send_keys('Joseph Nogosek')
if overUnder == "under" and ws["I2"].value != "ESS":
    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys('Joseph Nogosek')
driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
time.sleep(5)

# If verify complete, skip. If meter details already filled out, skip (Could be the same check for verify complete??)
if ws["C3"].value != "Verify Complete" and ws["I2"].value != "ESS" and overUnder == "under":
    # Open meter details action
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    time.sleep(1)
    for count in range(2,25):
        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Meter Details':
             driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
             time.sleep(2)
             driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
             break
    time.sleep(3)
    # Get window handles
    p = driver.current_window_handle
    chwd = driver.window_handles
    for w in chwd:
        if(w!=p):
            driver.switch_to.window(w)
    time.sleep(0.9)

    meter_swap = ws["H127"].value
    production_meter = ws["H128"].value
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]/lightning-textarea[1]/div[1]/textarea[1]").send_keys(meter_swap)
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/lightning-textarea[1]/div[1]/textarea[1]").send_keys(production_meter)
    # 99% of time this is no, so it's no by default and if it needs to be changed, then it must be done manually after the fact
    driver.find_element_by_xpath("//span[contains(text(),'No')]").click()
    # Depending on produciton meter select proper from dropdown
    if production_meter == "NEX":
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        driver.find_element_by_css_selector("button[name='NEXT']").click()
    else:
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        driver.find_element_by_css_selector("button[name='NEXT']").click()
        time.sleep(2)
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[3]/button[1]").click()

    # Close meter details and select correct window handle going back to the correct IA page
    time.sleep(8)
    driver.close()
    driver.switch_to.window(p)
    driver.find_element_by_id('phSearchInput').send_keys(IA)
    driver.find_element_by_id('phSearchButton').click()
    driver.find_element_by_css_selector("a[target='_top']").click()
    time.sleep(5)

# Uncheck everything that is checked if verify complete
if ws["C3"].value == 'Verify Complete':
    # Click the 'Go to List' page for checks
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[17]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    # 3.21
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.2
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.3
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.4
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.2
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.3
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # # If larger than 40 kW
    # # 3.4
    #     if ws["F"] .value == 'x' or ws["
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[10]/td[1]/a[1]").click()
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.10
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.11
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.12
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.13
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.14
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[15]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[15]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.15
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.16
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.17
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.5
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[19]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[19]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.6
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[20]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[20]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.7
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[21]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[21]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.8
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[22]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[22]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.9
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[23]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[23]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # # If over 40 kW
    #     # 3.22
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.222.1.2
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.3
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.3
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.4
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.5
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.2
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.2.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    # 3.18
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[33]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[33]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[34]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[34]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.2
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[35]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[35]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.3
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[36]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[36]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

    # Go to next page
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[1]/a[1]").click()

    # 3.19
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.19.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.20
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.2
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.3
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.4
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.5
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.6
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[10]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[10]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.10
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.11
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.7
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.7.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.8
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.8.1
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.9
    isChecked = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[5]/img[1]").get_attribute("alt")
    if isChecked == 'Checked':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # Go back to interconnection app
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[1]/div[2]/a[1]").click()
    time.sleep(3)


# Fill out checks if application is getting rejected
if ws["C5"].value == 'Rejected':
    # Click the 'Go to List' page for checks
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[17]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    # 3.21
    if ws["F49"].value == 'x' or ws["F49"].value == 'X' or ws["F49"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.1
    if ws["F50"].value == 'x' or ws["F50"].value == 'X' or ws["F50"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.2
    if ws["F52"].value == 'x' or ws["F52"].value == 'X' or ws["F52"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.3
    if ws["F53"].value == 'x' or ws["F53"].value == 'X' or ws["F53"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.21.4
    if ws["F54"].value == 'x' or ws["F54"].value == 'X' or ws["F54"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.1
    if ws["F13"].value == 'x' or ws["F13"].value == 'X' or ws["F13"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.2
    if ws["F14"].value == 'x' or ws["F14"].value == 'X' or ws["F14"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.3
    if ws["F15"].value == 'x' or ws["F15"].value == 'X' or ws["F15"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # # If larger than 40 kW
    # # 3.4
    #     if ws["F"] .value == 'x' or ws["
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[10]/td[1]/a[1]").click()
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
    #     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.10
    if ws["F27"].value == 'x' or ws["F27"].value == 'X' or ws["F27"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.11
    if ws["F28"].value == 'x' or ws["F28"].value == 'X' or ws["F28"].value == 'CA' or ws["F29"].value == 'x' or ws["F29"].value == 'X' or ws["F29"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.12
    if ws["F30"].value == 'x' or ws["F30"].value == 'X' or ws["F30"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.13
    if ws["F31"].value == 'x' or ws["F31"].value == 'X' or ws["F31"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.14
    if ws["F34"].value == 'x' or ws["F34"].value == 'X' or ws["F34"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[15]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.15
    if ws["F35"].value == 'x' or ws["F35"].value == 'X' or ws["F35"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.16
    if ws["F36"].value == 'x' or ws["F36"].value == 'X' or ws["F36"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.17
    if ws["F37"].value == 'x' or ws["F37"].value == 'X' or ws["F37"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.5
    if ws["F20"].value == 'x' or ws["F20"].value == 'X' or ws["F20"].value == 'CA' or ws["F21"].value == 'x' or ws["F21"].value == 'X' or ws["F21"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[19]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.6
    if ws["F22"].value == 'x' or ws["F22"].value == 'X' or ws["F22"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[20]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.7
    if ws["F23"].value == 'x' or ws["F23"].value == 'X' or ws["F23"].value == 'CA' or ws["F24"].value == 'x' or ws["F24"].value == 'X' or ws["F24"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[21]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.8
    if ws["F25"].value == 'x' or ws["F25"].value == 'X' or ws["F25"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[22]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.9
    if ws["F26"].value == 'x' or ws["F26"].value == 'X' or ws["F26"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[23]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # # If over 40 kW
    #     # 3.22
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.222.1.2
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.3
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.3
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.4
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.1.5
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.2
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #     # 3.22.2.1
    #     if ws["F"] .value == 'x' or ws["
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    #         driver.find_element_by_xpath("").click()
    # 3.18
    if ws["F39"].value == 'x' or ws["F39"].value == 'X' or ws["F39"].value == 'CA' or ws["F42"].value == 'x' or ws["F42"].value == 'X' or ws["F42"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[33]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.1
    if ws["F40"].value == 'x' or ws["F40"].value == 'X' or ws["F40"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[34]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.2
    if ws["F41"].value == 'x' or ws["F41"].value == 'X' or ws["F41"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[35]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.18.3
    if ws["F43"].value == 'x' or ws["F43"].value == 'X' or ws["F43"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[36]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

    # Go to next page
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[1]/a[1]").click()

    # 3.19
    if ws["F44"].value == 'x' or ws["F44"].value == 'X' or ws["F44"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.19.1
    if ws["F46"].value == 'x' or ws["F46"].value == 'X' or ws["F46"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 3.20
    if ws["F47"].value == 'x' or ws["F47"].value == 'X' or ws["F47"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.1
    if ws["F67"].value == 'x' or ws["F67"].value == 'X' or ws["F67"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[5]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.2
    if ws["F68"].value == 'x' or ws["F68"].value == 'X' or ws["F68"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[6]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.3
    if ws["F72"].value == 'x' or ws["F72"].value == 'X' or ws["F72"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[7]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.4
    if ws["F73"].value == 'x' or ws["F73"].value == 'X' or ws["F73"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[8]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.5
    if ws["F74"].value == 'x' or ws["F74"].value == 'X' or ws["F74"].value == 'CA' or ws["F75"].value == 'x' or ws["F75"].value == 'X' or ws["F75"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[9]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.6
    if ws["F76"].value == 'x' or ws["F76"].value == 'X' or ws["F76"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[10]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.10
    if ws["F88"].value == 'x' or ws["F88"].value == 'X' or ws["F88"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[11]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.11
    if ws["F89"].value == 'x' or ws["F89"].value == 'X' or ws["F89"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[12]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.7
    if ws["F81"].value == 'x' or ws["F81"].value == 'X' or ws["F81"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[13]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.7.1
    if ws["F82"].value == 'x' or ws["F82"].value == 'X' or ws["F82"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[14]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.8
    if ws["F85"].value == 'x' or ws["F85"].value == 'X' or ws["F85"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[16]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.8.1
    if ws["F86"].value == 'x' or ws["F86"].value == 'X' or ws["F86"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[17]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # 4.9
    if ws["F87"].value == 'x' or ws["F87"].value == 'X' or ws["F87"].value == 'CA':
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[18]/td[1]/a[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    # Go back to interconnection app
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[1]/div[2]/a[1]").click()
    time.sleep(3)

# Reopen actions
flag = 0
driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
time.sleep(1)
try:
    for count in range(2,25):
        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Reopen Action (Review Missing Info)':
             driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
             time.sleep(2)
             driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
             flag += 1
             break
except:
    pass
if flag == 0:
    try:
        for count in range(2,25):
            if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Reopen Action (Initiate Application)':
                 driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                 time.sleep(2)
                 driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                 break
    except:
        pass
time.sleep(4)
# Move to the correct tab
p = driver.current_window_handle
chwd = driver.window_handles
for w in chwd:
    if(w!=p):
        driver.switch_to.window(w)
time.sleep(0.9)

# Check if actions other than what was caught in checks need to be reopened
flag_2 = 0
if ws["C5"].value == 'Rejected':
    # Check system details
    if ws["C7"].value == 'x' or ws["C7"].value == 'X':
        try:
            for count in range(1,20):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'System Details':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                     flag_2 += 1
                     break
        except:
            pass
    # Check Application Details
    if ws["C8"].value == 'x' or ws["C8"].value == 'X':
        try:
            for count in range(1,20):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Application Details':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                     flag_2 += 1
                     break
        except:
            pass
    # Check Exhibit B Details
    if ws["C9"].value == 'x' or ws["C9"].value == 'X':
        try:
            for count in range(1,20):
                if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Exhibit B Details':
                     driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                     flag_2 += 1
                     break
        except:
            pass
    # Now we have checked all that are necessary. If flag_2 has been raised, we must stay on 'Yes' and click 'Submit'
    if flag_2 > 0:
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/div[1]/button[1]").click()
        time.sleep(5)
        driver.close()
        driver.switch_to.window(p)

    # If flag_2 has not been raised, then we select 'No' and click 'Submit'
    else:
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
        time.sleep(5)
        driver.close()
        driver.switch_to.window(p)

# If not getting rejected, just say no and move on
else:
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
    time.sleep(5)
    driver.close()
    driver.switch_to.window(p)

# Go back to the main page so we are ready for the next task
driver.back()
driver.back()
time.sleep(2)

# Go to approve/reject page
driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[2]").click()
# Paste the comments in
driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/textarea[1]").send_keys(ws["D5"].value)
# # Click approve or reject
# if ws["C5"].value == 'Rejected':
#     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
# else:
#     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
#
# # Wait for page to load again and give it about one minute then refresh and push application through again
# time.sleep(60)
# driver.refresh()
# driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[1]/a[2]").click()
# if ws["C5"].value == 'Rejected':
#     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
# else:
#     driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
#
# time.sleep(60)
# # Close the page now that the application has been processed
# driver.close()
