# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 8/11/2022
# Close tracker and make sure G:\ directory is green

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import openpyxl
import glob
import win32com.client as win32
import calendar
import os
import warnings
import datetime
import numpy as np
import shutil
from copy import copy
from decimal import Decimal
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options

username = '239665'
password = 'airdoc2Ee'
pin = '2846'

warnings.filterwarnings("ignore", category=DeprecationWarning)

twoFA = input("Enter six digit 2FA code: ")

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe", options=options)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element_by_xpath('//button[normalize-space()="Xcel Energy CORP credentials"]').click()

user_box = driver.find_element_by_id('username')
user_box.send_keys(username)

pass_box = driver.find_element_by_id('password')
pass_box.send_keys(password)

sign_on_button = driver.find_element_by_css_selector("a[onclick^='postOk']").click()

passcode_box = driver.find_element_by_name('pf.pass')
passcode_box.send_keys(pin + str(twoFA))
submit_button = driver.find_element_by_xpath("//button[contains(@onclick,'postOk')]").click()

# Click the report tab
try:
    for i in range(1,10):
        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[2]/div[1]/div[1]/nav[1]/ul[1]/li[{str(i)}]/a[1]").get_attribute("innerText") == 'Reports':
            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[2]/div[1]/div[1]/nav[1]/ul[1]/li[{str(i)}]/a[1]").click()
            break
except:
    pass

sleep(5)
try:
    for i in range(1,10):
        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").get_attribute("innerText") == 'CR_2022_NSP_Ulteig':
            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").click()
            break
except:
    pass

sleep(3)
driver.find_element_by_css_selector("[title*='Export Details']").click()

driver.find_element_by_css_selector("[title*='Export']").click()
sleep(5)

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Define the path to the Completeness Review Tracker and your downloads folder
pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx'
pathToDownloadsFolder = r'C:\Users\joe.nogosek\Downloads'

# Define paths to report and convert to correct file format
if not os.path.exists(os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))):
    raise ValueError('No report file present in the Completeness Review Reports folder')
else:
    pathToReportxls = os.path.join(pathToDownloadsFolder, str(glob.glob(r'C:\Users\joe.nogosek\Downloads\report*')[0]))
    pathToReportxlsx = pathToReportxls + 'x'

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(pathToReportxls)
    wb.SaveAs(pathToReportxlsx, FileFormat = 51) # 51 corresponds to the .xlsx extension file format. 56 would be for .xls
    wb.Close()
    os.remove(pathToReportxls)

# Load in the new report file just created and current tracker and set some values which will be useful
wbReport = load_workbook(pathToReportxlsx)
wbTracker = load_workbook(pathToTracker)

wsReport = wbReport.active
row_max_report = wsReport.max_row - 6

IAnums_tracker = []
existingMeters_tracker = []
dueDates_tracker = []
types_tracker = []
developers_tracker = []
history_tracker = []
lineDistances_tracker = []
sizes_tracker = []
submittal_date_tracker = []
# Go through all the sheets in the tracker to get all the necessary information in lists
for sheet in wbTracker.sheetnames:
    wsTracker = wbTracker[sheet]
    # Get all the IA numbers from this sheet in the book
    for row in wsTracker['B']:
        if row.row == 1:
            pass
        else:
            IAnums_tracker.append(row.value)
            #print(f"IA Num: {row.value}")
    # Get all the existing meter types from this sheet in the book
    for row in wsTracker['D']:
        if row.row == 1:
            pass
        else:
            existingMeters_tracker.append(row.value)
            #print(f"Existing Meter: {row.value}")
    # Get all the due dates from this sheet in the book
    for row in wsTracker['N']:
        if row.row == 1:
            pass
        else:
            dueDates_tracker.append(row.value)
            #print(f"Due date: {row.value}")
    # Get all the review types (ESS or not) from this sheet in the book
    for row in wsTracker['E']:
        if row.row == 1:
            pass
        else:
            types_tracker.append(row.value)
            #print(f"Type: {row.value}")
    # Get all the developers from this sheet in the book
    for row in wsTracker['J']:
        if row.row == 1:
            pass
        else:
            developers_tracker.append(row.value)
            #print(f"Developer: {row.value}")
    # Get all the approval/rejection from this sheet in the book
    for row in wsTracker['L']:
        if row.row == 1:
            pass
        else:
            history_tracker.append(row.value)
            #print(f"History: {row.value}")
    # Get all the line distances from this sheet in the book
    for row in wsTracker['F']:
        if row.row == 1:
            pass
        else:
            lineDistances_tracker.append(row.value)
            #print(f"Line Distance: {row.value}")
    # Get all the system sizes from this sheet in the book
    for row in wsTracker['G']:
        if row.row == 1:
            pass
        else:
            sizes_tracker.append(row.value)
            #print(f"Size: {row.value}")
    # Get all the submittal dates from this sheet in the book
    for row in wsTracker['M']:
        if row.row == 1:
            pass
        else:
            submittal_date_tracker.append(row.value)
            #print(f"Submittal date: {row.value}")

# Get all the information from the report in lists
# Initialize necessary lists
IAnums_report = []
caseNums_report = []
developers_report = []
dueDates_report = []
statuses_report = []
sizes_report = []
types_report = []
premise_report = []
programType_report = []
feederNums_report = []

# Get the IA numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=5, max_col=5, values_only=True):
    IAnums_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Get the premise numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=6, max_col=6, values_only=True):
    premise_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Extract just the premise number
for premise in range(len(premise_report)):
    if type(premise_report[premise]) != type(None):
        premise_report[premise] = premise_report[premise][-11:]
        premise_report[premise] = premise_report[premise][:9]
        if premise_report[premise] == 'der Premi':
            premise_report[premise] = 'New Construction'
# Get the case numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=7, max_col=7, values_only=True):
    caseNums_report.append(row[0])
    #print(f"Case num: {row[0]}")
# Get all the developers from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=8, max_col=8, values_only=True):
    developers_report.append(row[0])
    #print(f"Developer: {row[0]}")
# Get the due dates from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=9, max_col=9, values_only=True):
    dueDates_report.append(row[0])
    #print(f"Due Date: {row[0]}")
# Get all the application statuses from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=12, max_col=12, values_only=True):
    statuses_report.append(row[0])
    #print(f"Status: {row[0]}")
# Get the application sizes from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=13, max_col=13, values_only=True):
    sizes_report.append(row[0])
    #print(f"Size: {row[0]}")
# Get all the application types (ESS or not) from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=14, max_col=14, values_only=True):
    types_report.append(row[0])
    #print(f"Type: {row[0]}")
# Get the program types
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=10, max_col=10, values_only=True):
    programType_report.append(row[0])
    #print(f"Type: {row[0]}")
# Get the feeder numbers
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=17, max_col=17, values_only=True):
    feederNums_report.append(row[0])
    #print(f"Type: {row[0]}")

# Create styles
redFill = PatternFill(start_color='FFFF0000', end_color='FF0000', fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF', end_color='000000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFF00', fill_type='solid')
greyFill = PatternFill(start_color='FFF2F2F2', end_color='F2F2F2', fill_type='solid')
greenFill = PatternFill(start_color = 'FFA9D08E', end_color = 'A9D08E', fill_type='solid')

row_max_tracker = wsTracker.max_row
wsTracker['O' + str(row_max_tracker)] = ''
next_row_tracker = row_max_tracker + 1

amountNewReviews = 0
# Count the total number of reviews to be added
for w in range(row_max_report-1):
    if statuses_report[w] == 'Initiate Application':
        if (IAnums_report[w] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])] == 'Rejected') and (dueDates_report[w] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])]):
            amountNewReviews += 1
        elif (IAnums_report[w] not in IAnums_tracker):
            amountNewReviews += 1

# Number of times we have made it through
k = 1
flipFlop = 0
for i in range(row_max_report-1):
    if statuses_report[i] == 'Initiate Application':
        if (IAnums_report[i] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Rejected') and (dueDates_report[i] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])]):
            if feederNums_report[i] == '-':
                wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
                wsTracker['A' + str(next_row_tracker)].fill = redFill
            else:
                wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
                wsTracker['A' + str(next_row_tracker)].fill = whiteFill
            wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)].fill = whiteFill
            wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
            wsTracker['C' + str(next_row_tracker)].fill = whiteFill
            wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
            wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
            wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
            if sizes_report[i] >= 20:
                 wsTracker['F' + str(next_row_tracker)].fill = redFill
            else:
                 wsTracker['F' + str(next_row_tracker)].fill = greyFill
            wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
            if sizes_report[i] < 40:
                 wsTracker['G' + str(next_row_tracker)].fill = yellowFill
            else:
                wsTracker['G' + str(next_row_tracker)].fill = whiteFill
            wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
            wsTracker['H' + str(next_row_tracker)].fill = whiteFill
            wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
            wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
            wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
            wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
            wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
            wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
            wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
            wsTracker['P' + str(next_row_tracker)]._style = copy(wsTracker['P' + str(row_max_tracker)]._style)
            wsTracker['Q' + str(next_row_tracker)]._style = copy(wsTracker['Q' + str(row_max_tracker)]._style)
            wsTracker['R' + str(next_row_tracker)]._style = copy(wsTracker['R' + str(row_max_tracker)]._style)
            wsTracker['S' + str(next_row_tracker)]._style = copy(wsTracker['S' + str(row_max_tracker)]._style)
            wsTracker['T' + str(next_row_tracker)]._style = copy(wsTracker['T' + str(row_max_tracker)]._style)
            dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
            if feederNums_report[i] == '-':
                wsTracker['A' + str(next_row_tracker)] = ''
            else:
                wsTracker['A' + str(next_row_tracker)] = feederNums_report[i]
            wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
            wsTracker['C' + str(next_row_tracker)] = premise_report[i]
            wsTracker['D' + str(next_row_tracker)] = 'Verify Complete'
            wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
            if types_report[i] == 'Yes':
                #print('Reassign ' + str(IAnums_report[i]) + ' to metering (ESS)')
                wsTracker['E' + str(next_row_tracker)] = 'ESS'
            if lineDistances_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Complete':
                wsTracker['F' + str(next_row_tracker)] = 'Complete'
            wsTracker['G' + str(next_row_tracker)] = sizes_report[i]
            if sizes_report[i] < 40:
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
            else:
                #print('Reassign ' + str(IAnums_report[i]) + ' to metering (Larger than 40 kW)')
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
            if developers_report[i] == 'All Energy Solar, Inc.' or developers_report[i] == 'Wolf River Electric' or developers_report[i] == 'Everlight Solar, LLC':
                wsTracker['H' + str(next_row_tracker)] = 'Joe Nogo'
            else:
                wsTracker['H' + str(next_row_tracker)] = 'Josh B'
            wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
            wsTracker['J' + str(next_row_tracker)] = developers_report[i]
            wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
            wsTracker['P' + str(next_row_tracker)] = programType_report[i]
            wsTracker['Q' + str(next_row_tracker)] = f'=IF(K{next_row_tracker}="RK","RK",IF(K{next_row_tracker}="x","","JN"))'
            wsTracker['S' + str(next_row_tracker)] = 'x'
            next_row_tracker = next_row_tracker + 1
            k += 1
        elif (IAnums_report[i] not in IAnums_tracker):
            if feederNums_report[i] == '-':
                wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
                wsTracker['A' + str(next_row_tracker)].fill = redFill
            else:
                wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
                wsTracker['A' + str(next_row_tracker)].fill = whiteFill
            wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
            wsTracker['B' + str(next_row_tracker)].fill = whiteFill
            wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
            wsTracker['C' + str(next_row_tracker)].fill = whiteFill
            wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
            wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
            wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
            if sizes_report[i] >= 20:
                 wsTracker['F' + str(next_row_tracker)].fill = redFill
            else:
                 wsTracker['F' + str(next_row_tracker)].fill = greyFill
            wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
            if sizes_report[i] < 40:
                 wsTracker['G' + str(next_row_tracker)].fill = yellowFill
            else:
                wsTracker['G' + str(next_row_tracker)].fill = whiteFill
            wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
            wsTracker['H' + str(next_row_tracker)].fill = whiteFill
            wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
            wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
            wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
            wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
            wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
            wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
            wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
            wsTracker['P' + str(next_row_tracker)]._style = copy(wsTracker['P' + str(row_max_tracker)]._style)
            wsTracker['Q' + str(next_row_tracker)]._style = copy(wsTracker['Q' + str(row_max_tracker)]._style)
            wsTracker['R' + str(next_row_tracker)]._style = copy(wsTracker['R' + str(row_max_tracker)]._style)
            wsTracker['S' + str(next_row_tracker)]._style = copy(wsTracker['S' + str(row_max_tracker)]._style)
            wsTracker['T' + str(next_row_tracker)]._style = copy(wsTracker['T' + str(row_max_tracker)]._style)
            dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
            if feederNums_report[i] == '-':
                wsTracker['A' + str(next_row_tracker)] = ''
            else:
                wsTracker['A' + str(next_row_tracker)] = feederNums_report[i]
            wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
            wsTracker['C' + str(next_row_tracker)] = premise_report[i]
            if types_report[i] == 'Yes' or sizes_report[i] >= 40:
                wsTracker['D' + str(next_row_tracker)] = 'N/A'
                wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
            else:
                wsTracker['D' + str(next_row_tracker)] = '2S'
                wsTracker['D' + str(next_row_tracker)].font = Font(bold=True)
            if types_report[i] == 'Yes':
                # Put the IA that is being processed here and pull it from the folder name
                driver.find_element_by_id('phSearchInput').send_keys(str(IAnums_report[i]))
                driver.find_element_by_id('phSearchButton').click()

                # Click the top result from case number selector
                driver.find_element_by_css_selector("a[target='_top']").click()
                sleep(5)

                # Change the name of Area and meter engineer (if applicable)
                driver.find_element_by_css_selector("[title='Edit']").click()
                sleep(3)
                driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                if flipFlop % 3 == 0:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Spencer Doriot")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Spencer Doriot (Larger than 40 kW)')
                elif flipFlop % 3 == 1:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Shaun Ly")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Shaun Ly (Larger than 40 kW)')
                else:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Chi Vang")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Chi Vang (Larger than 40 kW)')
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                sleep(5)
                flipFlop += 1
                wsTracker['E' + str(next_row_tracker)] = 'ESS'
            wsTracker['G' + str(next_row_tracker)] = sizes_report[i]
            if sizes_report[i] < 40:
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
            else:
                # Put the IA that is being processed here and pull it from the folder name
                driver.find_element_by_id('phSearchInput').send_keys(str(IAnums_report[i]))
                driver.find_element_by_id('phSearchButton').click()

                # Click the top result from case number selector
                driver.find_element_by_css_selector("a[target='_top']").click()
                sleep(5)

                # Change the name of Area and meter engineer (if applicable)
                driver.find_element_by_css_selector("[title='Edit']").click()
                driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                if flipFlop % 3 == 0:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Spencer Doriot")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Spencer Doriot (Larger than 40 kW)')
                elif flipFlop % 3 == 1:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Shaun Ly")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Shaun Ly (Larger than 40 kW)')
                else:
                    driver.find_element_by_css_selector("[title='Meter Engineer Approver']").send_keys("Chi Vang")
                    print('Reassigned ' + str(IAnums_report[i]) + ' to Chi Vang (Larger than 40 kW)')
                sleep(2)
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                sleep(5)
                flipFlop += 1
                wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
            if developers_report[i] == 'All Energy Solar, Inc.' or developers_report[i] == 'Wolf River Electric' or developers_report[i] == 'Everlight Solar, LLC':
                wsTracker['H' + str(next_row_tracker)] = 'Joe Nogo'
            else:
                wsTracker['H' + str(next_row_tracker)] = 'Josh B'
            wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
            wsTracker['J' + str(next_row_tracker)] = developers_report[i]
            wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
            wsTracker['P' + str(next_row_tracker)] = programType_report[i]
            wsTracker['Q' + str(next_row_tracker)] = f'=IF(K{next_row_tracker}="RK","RK",IF(K{next_row_tracker}="x","","JN"))'
            wsTracker['S' + str(next_row_tracker)] = 'x'
            next_row_tracker = next_row_tracker + 1
            k += 1

# Put the date and time of the update
time = datetime.datetime.today().strftime("%I:%M %p")
date = datetime.date.strftime(datetime.date.today(), "%m/%d/%Y")
wsTracker['O' + str(next_row_tracker - 1)] = 'Tracker updated as of ' + str(time) + ' MST on ' + str(date)
# Save the Excel file
wbTracker.save(pathToTracker)

wbTracker.close()
wbReport.close()

# Delete the report so the script is ready for next time
os.remove(pathToReportxlsx)


"""
Now it's time to setup the folders
"""


# Now get the newly added reviews ready for me to complete
IA_num_Nick = []
case_num_Nick = []
size_Nick = []
program_Nick = []
type_Nick = []
installer_Nick = []
due_date_Nick = []
reviewer_Nick = []
meter_Nick = []

IA_num_Ethan = []
case_num_Ethan = []
size_Ethan = []
program_Ethan = []
type_Ethan = []
installer_Ethan = []
due_date_Ethan = []
reviewer_Ethan = []
meter_Ethan = []

IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
installer_Joe = []
due_date_Joe = []
reviewer_Joe = []
meter_Joe = []

IA_num_Jose = []
case_num_Jose = []
size_Jose = []
program_Jose = []
type_Jose = []
installer_Jose = []
due_date_Jose = []
reviewer_Jose = []
meter_Jose = []

IA_num_Ross = []
case_num_Ross = []
size_Ross = []
program_Ross = []
type_Ross = []
installer_Ross = []
due_date_Ross = []
reviewer_Ross = []
meter_Ross = []

IA_num_Josh = []
case_num_Josh = []
size_Josh = []
program_Josh = []
type_Josh = []
installer_Josh = []
due_date_Josh = []
reviewer_Josh = []
meter_Josh = []

IA_num_Adan = []
case_num_Adan = []
size_Adan = []
program_Adan = []
type_Adan = []
installer_Adan = []
due_date_Adan = []
reviewer_Adan = []
meter_Adan = []

wb = load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER System Impact Studies - DER Reviews - NSP\Completeness Review Tracker.xlsx')
ws = wb['2022']

# Nick
i=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Nick C" and ws[row][10].value==None:
        IA_num_Nick.append(ws[row][1].value)
        case_num_Nick.append(ws[row][8].value)
        size_Nick.append(ws[row][6].value)
        program_Nick.append(ws[row][15].value)
        type_Nick.append(ws[row][4].value)
        installer_Nick.append(ws[row][9].value)
        due_date_Nick.append(ws[row][13].value)
        reviewer_Nick.append(ws[row][7].value)
        meter_Nick.append(ws[row][3].value)
        i=i+1

# Ethan
j=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ethan U" and ws[row][10].value==None:
        IA_num_Ethan.append(ws[row][1].value)
        case_num_Ethan.append(ws[row][8].value)
        size_Ethan.append(ws[row][6].value)
        program_Ethan.append(ws[row][15].value)
        type_Ethan.append(ws[row][4].value)
        installer_Ethan.append(ws[row][9].value)
        due_date_Ethan.append(ws[row][13].value)
        reviewer_Ethan.append(ws[row][7].value)
        meter_Ethan.append(ws[row][3].value)
        j=j+1

# Joe
k=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        installer_Joe.append(ws[row][9].value)
        due_date_Joe.append(ws[row][13].value)
        reviewer_Joe.append(ws[row][7].value)
        meter_Joe.append(ws[row][3].value)
        k=k+1

# Jose
l=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Jose CN" and ws[row][10].value==None:
        IA_num_Jose.append(ws[row][1].value)
        case_num_Jose.append(ws[row][8].value)
        size_Jose.append(ws[row][6].value)
        program_Jose.append(ws[row][15].value)
        type_Jose.append(ws[row][4].value)
        installer_Jose.append(ws[row][9].value)
        due_date_Jose.append(ws[row][13].value)
        reviewer_Jose.append(ws[row][7].value)
        meter_Jose.append(ws[row][3].value)
        l=l+1

# Ross
m=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Ross K" and ws[row][10].value==None:
        IA_num_Ross.append(ws[row][1].value)
        case_num_Ross.append(ws[row][8].value)
        size_Ross.append(ws[row][6].value)
        program_Ross.append(ws[row][15].value)
        type_Ross.append(ws[row][4].value)
        installer_Ross.append(ws[row][9].value)
        due_date_Ross.append(ws[row][13].value)
        reviewer_Ross.append(ws[row][7].value)
        meter_Ross.append(ws[row][3].value)
        m=m+1

# Josh
p=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Josh B" and ws[row][10].value==None:
        IA_num_Josh.append(ws[row][1].value)
        case_num_Josh.append(ws[row][8].value)
        size_Josh.append(ws[row][6].value)
        program_Josh.append(ws[row][15].value)
        type_Josh.append(ws[row][4].value)
        installer_Josh.append(ws[row][9].value)
        due_date_Josh.append(ws[row][13].value)
        reviewer_Josh.append(ws[row][7].value)
        meter_Josh.append(ws[row][3].value)
        p=p+1

# Adan
q=0
for row in range(3,ws.max_row+1):
    if ws[row][7].value=="Adan A" and ws[row][10].value==None:
        IA_num_Adan.append(ws[row][1].value)
        case_num_Adan.append(ws[row][8].value)
        size_Adan.append(ws[row][6].value)
        program_Adan.append(ws[row][15].value)
        type_Adan.append(ws[row][4].value)
        installer_Adan.append(ws[row][9].value)
        due_date_Adan.append(ws[row][13].value)
        reviewer_Adan.append(ws[row][7].value)
        meter_Adan.append(ws[row][3].value)
        q=q+1

wb.close()
print("You can open the tracker now.")

# Count the total number of folders succesfully created
counter = 0
total = 0

print("Populating Josh's folder.")
# Josh
for value in range(p):
    no_DoS_flag = False
    try:
        total += 1
        name = str(IA_num_Josh[value])
        type = str(type_Josh[value])
        size = float(size_Josh[value])
        full_name = r'G:\2021\21.00016\Reviews\JB\{}-JB'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Josh[value] < 40 and (program_Josh[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        elif size_Josh[value] < 40 and (program_Josh[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I3'].value = installer_Josh[value]
                ws['C1'].value = case_num_Josh[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Josh[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Josh[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Josh[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Josh[value]
                ws["C3"].value = meter_Josh[value]
                ws["C4"].value = due_date_Josh[value]
                wb.save(new_file)

            elif size_Josh[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Josh[value]
                ws["C3"].value = due_date_Josh[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Adan's folder.")
# Adan
for value in range(q):
    try:
        total += 1
        name = str(IA_num_Adan[value])
        type = str(type_Adan[value])
        full_name = r'G:\2021\21.00016\Reviews\AA\{}-AA'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Adan[value] < 40 and (program_Adan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        elif size_Adan[value] < 40 and (program_Adan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Adan[value]
                ws['C1'].value = case_num_Adan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Adan[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Adan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Adan[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Adan[value]
                ws["C3"].value = meter_Adan[value]
                ws["C4"].value = due_date_Adan[value]
                wb.save(new_file)

            elif size_Adan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Adan[value]
                ws["C3"].value = due_date_Adan[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Joe's folder.")
# Joe
for value in range(k):
    no_DoS_flag = False
    try:
        total += 1
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2021\21.00016\Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Joe[value] < 40 and (program_Joe[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        elif size_Joe[value] < 40 and (program_Joe[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I3'].value = installer_Joe[value]
                ws['C1'].value = case_num_Joe[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Joe[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Joe[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Joe[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Joe[value]
                ws["C3"].value = meter_Joe[value]
                ws["C4"].value = due_date_Joe[value]
                wb.save(new_file)

            elif size_Joe[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Joe[value]
                ws["C3"].value = due_date_Joe[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Now populating Nick's folder.")
# Nick
for value in range(i):
    try:
        total += 1
        name = str(IA_num_Nick[value])
        type = str(type_Nick[value])
        full_name = r'G:\2021\21.00016\Reviews\NC\{}-NC'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Nick[value] < 40 and (program_Nick[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        elif size_Nick[value] < 40 and (program_Nick[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Nick[value]
                ws['C1'].value = case_num_Nick[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Nick[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Nick[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Nick[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Nick[value]
                ws["C3"].value = meter_Nick[value]
                ws["C4"].value = due_date_Nick[value]
                wb.save(new_file)

            elif size_Nick[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Nick[value]
                ws["C3"].value = due_date_Nick[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ethan's folder.")
# Ethan
for value in range(j):
    try:
        total += 1
        name = str(IA_num_Ethan[value])
        type = str(type_Ethan[value])
        full_name = r'G:\2021\21.00016\Reviews\EU\{}-EU'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ethan[value] < 40 and (program_Ethan[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        elif size_Ethan[value] < 40 and (program_Ethan[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ethan[value]
                ws['C1'].value = case_num_Ethan[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Ethan[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Ethan[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Ethan[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Ethan[value]
                ws["C3"].value = meter_Ethan[value]
                ws["C4"].value = due_date_Ethan[value]
                wb.save(new_file)

            elif size_Ethan[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Ethan[value]
                ws["C3"].value = due_date_Ethan[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("And now Jose (if he has any).")
# Jose
for value in range(l):
    try:
        total += 1
        name = str(IA_num_Jose[value])
        type = str(type_Jose[value])
        full_name = r'G:\2021\21.00016\Reviews\JCN\{}-JCN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Jose[value] < 40 and (program_Jose[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        elif size_Jose[value] < 40 and (program_Jose[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Jose[value]
                ws['C1'].value = case_num_Jose[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Jose[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Jose[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Jose[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Jose[value]
                ws["C3"].value = meter_Jose[value]
                ws["C4"].value = due_date_Jose[value]
                wb.save(new_file)

            elif size_Jose[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Jose[value]
                ws["C3"].value = due_date_Jose[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ross's folder.")
# Ross
for value in range(m):
    try:
        total += 1
        name = str(IA_num_Ross[value])
        type = str(type_Ross[value])
        full_name = r'G:\2021\21.00016\Reviews\RK\{}-RK'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        if size_Ross[value] < 40 and (program_Ross[value] == 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering_SR-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering_SR-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        elif size_Ross[value] < 40 and (program_Ross[value] != 'Solar*Rewards'):
            old_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_CR_Checklist_under40kW_metering-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_CR_Checklist_under40kW_metering-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)
        else:
            old_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-.xlsm')
            new_file = os.path.join(full_name, 'Xcel_Completeness_Review_Checklist_over40kW-{}.xlsm'.format(name))
            if not os.path.exists(new_file):
                shutil.copy(r'G:\2021\21.00016\Reviews\Xcel_Completeness_Review_Checklist_over40kW-.xlsm',full_name)
                os.rename(old_file, new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                ws['I1'].value = name
                ws['I2'].value = type
                ws['I3'].value = installer_Ross[value]
                ws['C1'].value = case_num_Ross[value]
                wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element_by_id('phSearchInput').send_keys(name)
            driver.find_element_by_id('phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element_by_css_selector("a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            # z is used to check how many files we have downloaded
            z=0
            # If we are not in an ESS application, we only need two files (oneline and site plan)
            if type_Ross[value] != "ESS":
                # Count through all of the actions looking for the oneline and site plan actions
                for count in range(2,25):
                    # If we find the oneline or site plan, enter
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                        # We have found the oneline or site plan action, so let's click on the action number to enter the action
                        driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                        time.sleep(3)
                        # Now we are on the action page, let's download the first (and therefore newest) file uploaded to the portal
                        driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                        time.sleep(5)
                        # After this next line we are back at the actions page
                        driver.back()
                        time.sleep(1)
                        # We have succesfully downloaded one file, so add one to z
                        z+=1
                        # If we have hit two files succesfully downloaded, let's get out of this for loop
                        if z == 2:
                            break
            # We are in an ESS application, so we need three files (oneline, site plan, and declaration of storage [if one was uplaoded])
            else:
                # Let's count through all the actions
                for count in range(2,25):
                    # Let's look for the oneline, site plan, and spec sheets actions
                    if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                        # If we are on the oneline or site plan actions, enter
                        if driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                            # Click the action number for this to enter the action page for either the oneline or the site plan
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # Download the most recently uploaded file (and therefore the newest)
                            driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                            time.sleep(5)
                            # Go back to the action number page
                            driver.back()
                            time.sleep(1)
                            z+=1
                        # Okay now we are on the spec sheet action, so we need to do something special to find the declaration of storage
                        elif driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Specification Sheet(s)':
                            # Let's enter the action number for the spec sheets
                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                            time.sleep(3)
                            # We are going to give our best effort to find the DoS doc, but if we can't, we want to continue on anyway
                            try:
                                # There may be a lot of documents in the spec sheets action, so let's try going to the list
                                try:
                                    # Click go to list
                                    driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
                                    time.sleep(1)
                                    # Let's count through all of the files in this list
                                    for count2 in range(2,15):
                                        # Search through the titles (which have been forced lowercase) for 'storage declaration', 'declaration of storage', or 'ds'.
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the spec sheets action page
                                            driver.back()
                                            time.sleep(1)
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                # There was no go to list option, so let's proceed this way instead
                                except:
                                    for count2 in range(2,15):
                                        if ("storage declaration" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "ds" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower() or "declaration of storage" in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower()) and "-ds-" not in driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/th[1]/a[1]").get_attribute("innerText").lower():
                                            # Great we got a match, let's click download on that one
                                            driver.find_element_by_xpath(f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count2}]/td[1]/a[2]").click()
                                            time.sleep(5)
                                            # Let's go back to the action number action page
                                            driver.back()
                                            time.sleep(1)
                                            break
                                    z+=1
                                    pass
                            # We failed to find the DoS document, so let's throw a flag so we know we only need to move two documents instead of three
                            except:
                                no_DoS_flag = True
                                file_object = open(os.path.join(full_name,"No DoS doc found. Double check SF!.txt"),"w")
                                z+=1
                                driver.back()
                        # If we have downloaded (or tried to download) three files, let's get out of this for loop that is searching through the actions
                        if z == 3:
                            break
                    # While searching through the actions, if we hit an action that isn't one of the three we want, let's keep searching with the next action
                    else:
                        pass

            # After either of those breaks, we make it here
            # After all this action nonsense, we go back from the action page to the home page for this application
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
            sorted_files = sorted(list_of_files, key=os.path.getmtime)
            shutil.move(sorted_files[-1], full_name)
            shutil.move(sorted_files[-2], full_name)
            if z == 3 and no_DoS_flag == False:
                shutil.move(sorted_files[-3], full_name)

            if size_Ross[value] < 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D14"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D26"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D67"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D68"].value = "Premise: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[29]/table[1]/tbody[1]/tr[1]/td[2]/a[1]").get_attribute("innerText") + \
                "\nAddress: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") + \
                "\nCity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]").get_attribute("innerText")
                # Installer (site plan)
                ws["D72"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                "\n" + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                # Case number (site plan)
                ws["D73"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                # Feeder number
                ws["D137"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]").get_attribute("innerText")
                # Grab storage charge source if necessary
                if type_Ross[value] == "ESS":
                    try:
                        ws["D35"].value = "Storage charge source: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                    except:
                        pass
                wb.save(new_file)
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D28"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Ross[value]
                ws["C3"].value = meter_Ross[value]
                ws["C4"].value = due_date_Ross[value]
                wb.save(new_file)

            elif size_Ross[value] >= 40:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['DER Study']
                # Case number (oneline)
                ws["D13"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D12"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D27"].value = "AC capacity: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[21]/table[1]/tbody[1]/tr[1]/td[4]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[3]").get_attribute("innerText")
                # Existing DG
                if driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == " ":
                    ws["D34"].value = "Existing DG: Blank in SF"
                else:
                    ws["D34"].value = "Existing DG: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText")
                # Name of customer (site plan)
                ws["D66"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[4]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]/div[1]").get_attribute("innerText")
                wb.save(new_file)
                # Go into inverter information
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D30"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D29"].value = driver.find_element_by_xpath("/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                ws["C2"].value = reviewer_Ross[value]
                ws["C3"].value = due_date_Ross[value]
                wb.save(new_file)

        counter += 1
    except:
        print(f"Error with {name}")
        continue

driver.close()
print("Folders are now all populated.")
print(f"{counter} folders successfully created out of a total of {total}!")
