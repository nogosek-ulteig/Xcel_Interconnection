# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 11/10/2022

import openpyxl
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import os, os.path
from os.path import join
import sys
import numpy as np
import shutil
import warnings
import glob
import win32com.client as win32
import datetime
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert
import pyautogui
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

username = credentials.username
password = credentials.password
pathToTracker = credentials.path_to_CO_CR_tracker
pathToDownloadsFolder = credentials.path_to_downloads
path_to_driver = credentials.path_to_driver
name = credentials.name

warnings.filterwarnings("ignore")

pathToReadyQC = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC'
pathToForgotten = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\Forgotten_Reviews_pt2'

wb = load_workbook(pathToTracker)
ws = wb['2023']

# Now get all the necessary information from the 2022 tab
IA_num = []
reviewers = []
files = []

numReviews = 0
for row in range(2,ws.max_row+1):
    if str(ws[row][11].value) != 'None' and str(ws[row][12].value) == 'None':
        IA_num.append(ws[row][1].value)
        reviewers.append(ws[row][7].value)
        numReviews += 1

wb.close()
print("Good to open tracker.\n")

warnings.filterwarnings("ignore", category=DeprecationWarning)

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument("--disable-notifications")
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(path_to_driver, options=options)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')
time.sleep(1)

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

i=0
for i in range(numReviews):
    pyautogui.move(0,1)
    if reviewers[i] == "Joe Nogo":
        files.append(IA_num[i] + "-JN")
    elif reviewers[i] == "Ross K":
        files.append(IA_num[i] + "-RK")
    elif reviewers[i] == "Nick C":
        files.append(IA_num[i] + "-NC")
    elif reviewers[i] == "Jose CN":
        files.append(IA_num[i] + "-JCN")
    elif reviewers[i] == "Ethan U":
        files.append(IA_num[i] + "-EU")
    elif reviewers[i] == "Adan A":
        files.append(IA_num[i] + "-AA")
    elif reviewers[i] == "Josh B":
        files.append(IA_num[i] + "-JB")
    elif reviewers[i] == "Andrew N":
        files.append(IA_num[i] + "-AN")
    elif reviewers[i] == "Jason H":
        files.append(IA_num[i] + "-JH")
    elif reviewers[i] == "Abby M":
        files.append(IA_num[i] + "-AM")
    elif reviewers[i] == "Ed S":
        files.append(IA_num[i] + "-ES")
    elif reviewers[i] == "Josh G":
        files.append(IA_num[i] + "-JG")
    elif reviewers[i] == "Andre B":
        files.append(IA_num[i] + "-AB")
    elif reviewers[i] == "Caleb S":
        files.append(IA_num[i] + "-CS")
    elif reviewers[i] == "Anna R":
        files.append(IA_num[i] + "-AR")
    print(IA_num[i])
    try:
        pathToFile = os.path.join(pathToReadyQC,files[i])
        list_of_files = glob.glob(os.path.join(pathToFile,"*.xlsm")) # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
    except:
        pathToFile = os.path.join(pathToForgotten,files[i])
        list_of_files = glob.glob(os.path.join(pathToFile,"*.xlsm")) # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)

    try:
        wb = load_workbook(latest_file, data_only = True)
        ws = wb.active

        IA = ws["C1"].value
        pathToChecklist = latest_file
        overUnder = "Doesn't matter yet"
        status = ws["C5"].value
        reviewer = ws["J1"].value
        processor = "JN"

        # if str(ws["H92"].value) == 'None' or str(ws["H93"].value) == 'None':
        #     print("Populate meter information in checklist! It is empty currently.")
        #     continue

        # if str(ws["E24"].value) != 'None' or str(ws["F24"].value) != 'None':
        #     print("Meter collar. Manually process.")
        #     continue
        #
        # if str(ws["C4"].value) == 'Yes':
        #     print("ESS application. Manually process.")
        #     continue

        # Put the IA that is being processed here and pull it from the folder name
        driver.find_element("id", 'phSearchInput').send_keys(IA)
        driver.find_element("id", 'phSearchButton').click()

        # Click the top result from case number selector
        driver.find_element("css selector", "a[target='_top']").click()
        time.sleep(8)
        # If verify complete, skip. If meter details already filled out, skip (Could be the same check for verify complete??)
        # if str(ws["C3"].value).lower() != "verify complete" and (driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == "Josh Guck" or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == "Jose Coelho Neto" or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == "Joseph Nogosek" or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == "Ethan Unruh" or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == "Zakary Studebaker" or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[11]/table[1]/tbody[1]/tr[5]/td[2]/a[1]").get_attribute("innerText") == 'Cormac Heneghan'):
        if str(ws["C3"].value).lower() != "verify complete" and str(ws["C3"].value).lower() != "re-review" and str(ws["C3"].value).lower() != "n/a":
            # Open meter details action
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(1)
            try:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Meter Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(2)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                         break
                time.sleep(3)
            except:
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/div[1]/a[1]").click()
                time.sleep(1)
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Meter Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(2)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                         break
                time.sleep(3)
            # Get window handles
            p = driver.current_window_handle
            chwd = driver.window_handles
            for w in chwd:
                if(w!=p):
                    driver.switch_to.window(w)
            time.sleep(0.9)

            if ws["H102"].value == None and ws["H103"].value == None:
                meter_swap = ws["H92"].value
                production_meter = ws["H93"].value
            else:
                meter_swap = ws["H102"].value
                production_meter = ws["H103"].value
            time.sleep(1)
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[2]/flowruntime-aura-field/div/div/lightning-textarea/div/textarea").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[2]/flowruntime-aura-field/div/div/lightning-textarea/div/textarea").send_keys(meter_swap)
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-aura-field/div/div/lightning-textarea/div/textarea").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-aura-field/div/div/lightning-textarea/div/textarea").send_keys(production_meter)
            # Choose whether it's an existing installation or not
            if "existing" in str(ws["C3"].value).lower():
                driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-aura-field/div/div/lightning-radio-group/fieldset/div/div/span[1]/label").click()
            else:
                driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-aura-field/div/div/lightning-radio-group/fieldset/div/div/span[2]/label/span").click()
            # Depending on production meter select proper from dropdown
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[5]/flowruntime-aura-field/div/div/div[1]/div/select").send_keys(Keys.ENTER, Keys.UP, Keys.UP, Keys.UP, Keys.UP, Keys.ENTER)
            if production_meter == "N/A":
                driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[5]/flowruntime-aura-field/div/div/div[1]/div/select").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
                button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[6]/flowruntime-aura-field/div/div/div/button")
                driver.execute_script("arguments[0].click();", button)
            else:
                driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[5]/flowruntime-aura-field/div/div/div[1]/div/select").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
                button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[6]/flowruntime-aura-field/div/div/div/button")
                driver.execute_script("arguments[0].click();", button)
                time.sleep(2)
                button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[5]/flowruntime-aura-field/div/div/div[3]/button")
                driver.execute_script("arguments[0].click();", button)

            # Close meter details and select correct window handle going back to the correct IA page
            time.sleep(8)
            driver.close()
            driver.switch_to.window(p)
            driver.find_element("id", 'phSearchInput').send_keys(IA)
            driver.find_element("id", 'phSearchButton').click()
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)
        # Reopen actions
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
        time.sleep(3)
        try:
            for count in range(2,25):
                if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Reopen Action (Engr)':
                     driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(2)
                     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                     break
        except:
            pass

        time.sleep(4)
        # Move to the correct tab
        p = driver.current_window_handle
        chwd = driver.window_handles
        for w in chwd:
            if(w!=p):
                driver.switch_to.window(w)
        time.sleep(2)

        oneline = 0
        site_plan = 0
        # See if we need to open oneline and/or site plan
        if ws["K8"].value == "One-Line":
            oneline += 1
        if ws["K9"].value == "Site Plan":
            site_plan += 1

        # Check if actions other than what was caught in checks need to be reopened
        flag_2 = 0
        if ws["C5"].value == 'Rejected':
            # Check system details
            if ws["C7"].value == 'x' or ws["C7"].value == 'X':
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'System Details':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Application Details':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass
            # Check Meter Collar
            if (ws["C8"].value == 'x' or ws["C8"].value == 'X') and ws["B8"].value == 'Meter Collar Consent Form':
                print("Meter collar consent form checked.")
                # try:
                #     for count in range(1,20):
                #         if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Application Details':
                #              driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                #              flag_2 += 1
                #              break
                # except:
                #     pass
            # Check Battery Details
            if ws["C9"].value == 'x' or ws["C9"].value == 'X':
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Battery Details':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass
            # Check Declaration
            if ws["F39"].value == 'x' or ws["F39"].value == 'X':
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Declaration':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass
            # Check Oneline
            if oneline > 0:
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'One-Line Diagram':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass

            # Check Site Plan
            if site_plan > 0:
                try:
                    for count in range(1,20):
                        if driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Site Plan':
                             driver.find_element("xpath", f"/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-lwc-field/div/flowruntime-multi-checkbox-lwc/fieldset/div/span[{count}]/label/span[1]").click()
                             flag_2 += 1
                             break
                except:
                    pass
            # Now we have checked all that are necessary. If flag_2 has been raised, we must stay on 'Yes' and click 'Submit'
            if flag_2 > 0:
                button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[6]/flowruntime-aura-field/div/div/div/button")
                driver.execute_script("arguments[0].click();", button)
                time.sleep(5)
                driver.close()
                driver.switch_to.window(p)

            # If flag_2 has not been raised, then we select 'No' and click 'Submit'
            else:
                driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-lwc-field/div/flowruntime-picklist-input-lwc/div/lightning-select/div/div/select").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
                button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[6]/flowruntime-aura-field/div/div/div/button")
                driver.execute_script("arguments[0].click();", button)
                time.sleep(5)
                driver.close()
                driver.switch_to.window(p)

        # If not getting rejected, just say no and move on
        else:
            driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-lwc-field/div/flowruntime-picklist-input-lwc/div/lightning-select/div/div/select").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[6]/flowruntime-aura-field/div/div/div/button")
            driver.execute_script("arguments[0].click();", button)
            time.sleep(5)
            driver.close()
            driver.switch_to.window(p)

        # Go back to the main page so we are ready for the next task
        driver.back()
        driver.back()
        time.sleep(5)

        p = driver.current_window_handle

        # Populate Exhibit D action (if needed)
        if ws["C4"].value == "Yes" and (str(ws["C3"].value).lower() != "verify complete" and str(ws["C3"].value).lower() != "re-review"):
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[4]/div[1]/table/tbody/tr/td[2]/input[2]").click()
            time.sleep(1)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[23]/table/tbody/tr[1]/td[4]/input").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[23]/table/tbody/tr[1]/td[4]/input").send_keys(ws["F8"].value)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[23]/table/tbody/tr[4]/td[4]/input").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[23]/table/tbody/tr[4]/td[4]/input").send_keys(ws["F7"].value)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[3]/table/tbody/tr/td[2]/input[1]").click()
            time.sleep(5)

            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(1)
            try:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Populate Exhibit D':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(2)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                         break
            except:
                pass

            time.sleep(4)
            # Move to the correct tab
            p = driver.current_window_handle
            chwd = driver.window_handles
            for w in chwd:
                if(w!=p):
                    driver.switch_to.window(w)
            time.sleep(2)

            driver.close()
            driver.switch_to.window(p)

            # Go back to the main page so we are ready for the next task
            driver.back()
            driver.back()
            time.sleep(5)

            p = driver.current_window_handle

        # Go to approve/reject page
        for count in range(3,5):
            try:
                if (driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Kshitiz Karki' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Nimesh Shrestha' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Madeleine Balchan' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Kristin Gaspar') and driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Gihyeon Hong' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Abby Martin' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Grace Wenham' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Melkam Alemu' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Cormac Heneghan':
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    time.sleep(1)
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    time.sleep(1)
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[3]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[9]/table[1]/tbody[1]/tr[3]/td[2]/div[1]/span[1]/input[1]").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[9]/table[1]/tbody[1]/tr[3]/td[2]/div[1]/span[1]/input[1]").send_keys("Abby Martin")
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    time.sleep(5)
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    alert = Alert(driver)
                    alert.accept()
                    time.sleep(5)
                    break
            except:
                pass
        time.sleep(2)

        # for count in range(3,5):
        #     if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ross Kirby' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ethan Unruh' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Jose Coelho Neto' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == name or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Gihyeon Hong' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Grace Wenham' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Melkam Alemu' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Zakary Studebaker' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Cormac Heneghan':
        #         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[1]/a[2]").click()
        #         break

        for count in range(3,6):
            if driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Abby Martin' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Andrew Norman' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Josh Berg' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == 'Jose Coelho Neto' or driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[3]/a").get_attribute("innerText") == name:
                driver.find_element("xpath", f"/html/body/div/div[3]/table/tbody/tr/td[2]/div[14]/div[1]/div/div[2]/table/tbody/tr[{count}]/td[1]/a[2]").click()
                break
        # Paste the comments in
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/textarea[1]").send_keys(ws["D5"].value)

        # Click approve or reject
        if ws["C5"].value == 'Rejected':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
        else:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

        if status == 'Approved':
            status = 'Appr'
        elif status == 'Rejected':
            status = 'Rej'
        else:
            status = 'CA'

        print("First approval/rejection successfully completed.")
        pathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}'.format(IA, reviewer)
        newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}'.format(IA, reviewer, status)
        v2Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v2'.format(IA, reviewer, status)
        v3Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v3'.format(IA, reviewer, status)
        v4Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v4'.format(IA, reviewer, status)
        v5Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v5'.format(IA, reviewer, status)
        v6Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v6'.format(IA, reviewer, status)
        v7Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v7'.format(IA, reviewer, status)
        v8Path = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v8'.format(IA, reviewer, status)
        os.rename(pathToFolder, newPathToFolder)

        pathToApproved = r'G:\2023\23.22984\CO_Reviews\Complete-Approved'
        pathToRejected = r'G:\2023\23.22984\CO_Reviews\Complete-Rejected'
        rejectedV2 = os.path.join(pathToRejected, '{0}-{1}-{2}'.format(IA, reviewer, status))
        rejectedV3 = os.path.join(pathToRejected, '{0}-{1}-{2}v2'.format(IA, reviewer, status))
        rejectedV4 = os.path.join(pathToRejected, '{0}-{1}-{2}v3'.format(IA, reviewer, status))
        rejectedV5 = os.path.join(pathToRejected, '{0}-{1}-{2}v4'.format(IA, reviewer, status))
        rejectedV6 = os.path.join(pathToRejected, '{0}-{1}-{2}v5'.format(IA, reviewer, status))
        rejectedV7 = os.path.join(pathToRejected, '{0}-{1}-{2}v6'.format(IA, reviewer, status))
        rejectedV8 = os.path.join(pathToRejected, '{0}-{1}-{2}v7'.format(IA, reviewer, status))
        if status == 'Appr' or status == 'CA':
            shutil.move(newPathToFolder, pathToApproved)
        else:
            if not os.path.exists(rejectedV2) and not os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and not os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
                os.rename(newPathToFolder, v2Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v2'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
                os.rename(newPathToFolder, v3Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v3'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
                os.rename(newPathToFolder, v4Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v4'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and os.path.exists(rejectedV4) and os.path.exists(rejectedV5) and not os.path.exists(rejectedV6):
                os.rename(newPathToFolder, v5Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v5'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and os.path.exists(rejectedV4) and os.path.exists(rejectedV5) and os.path.exists(rejectedV6) and not os.path.exists(rejectedV7):
                os.rename(newPathToFolder, v6Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v6'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and os.path.exists(rejectedV4) and os.path.exists(rejectedV5) and os.path.exists(rejectedV6) and os.path.exists(rejectedV7) and not os.path.exists(rejectedV8):
                os.rename(newPathToFolder, v7Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v7'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)
            else:
                os.rename(newPathToFolder, v8Path)
                newPathToFolder = r'G:\2023\23.22984\CO_Reviews\Ready_for_QC\{0}-{1}-{2}v8'.format(IA, reviewer, status)
                shutil.move(newPathToFolder, pathToRejected)

        time.sleep(10)
        successFlag = 0
        try:
            try:
                driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/div/a[2]").click()
                time.sleep(1)
                for count in range(2,7):
                    if driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Provide Missing Information (Engineering)' or driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Provide Missing Information' or driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Deemed Complete Date' or driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Level 1 Screens (Simplified)':
                        if driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[4]/img[1]").get_attribute("alt") == 'Checked':
                            successFlag = 1
                            break

                # Go back to the main page so we are ready for the next task
                driver.back()

                if successFlag == 1:
                    pass
                else:
                    print("!!!!!! Error with " + IA)
            except:
                for count in range(2,6):
                    if driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Provide Missing Information (Engineering)' or driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Provide Missing Information' or driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Deemed Complete Date' or driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Level 1 Screens (Simplified)':
                        if driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[4]/img").get_attribute("alt") == 'Checked':
                            successFlag = 1
                            break

                if successFlag == 1:
                    pass
                else:
                    print("!!!!!! Error with " + IA)

        except Exception as e:
            print("!!!!!! Error with " + IA)

    except Exception as e:
        print(e)
        print("!!!!!! Error with " + IA)
        continue

print("This script is finished. Make sure to help all the open Chrome windows finish.")

input("Press enter to exit.")
