# Made by Joe Nogosek
# 3/30/2022

import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import shutil
import warnings
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

warnings.filterwarnings('ignore')

pathToTracker = credentials.path_to_CO_CR_tracker
#pathToTracker = r'C:\Users\joe.nogosek\Downloads\CR & LV1 Tracker - PSCO - Copy2.xlsx'
pathToReviews = r'G:\2023\23.22984\CO_Reviews'

wb = load_workbook(pathToTracker)
ws = wb['2023']

og_rev = []
new_rev = []
IA = []

total=0
for row in range(2,ws.max_row+1):
    if ws[row][17].value != None:
        og_rev.append(ws[row][7].value)
        new_rev.append(ws[row][17].value)
        IA.append(ws[row][1].value)
        #ws[row][7].value = ws[row][17].value
        #ws[row][17].value = None
        total=total+1

#wb.save(pathToTracker)
wb.close()

successes = 0
fails = 0
for i in range(total):
    if og_rev[i] == "Josh B":
        og_initials = "JB"
    elif og_rev[i] == "Ethan U":
        og_initials = "EU"
    elif og_rev[i] == "Jose CN":
        og_initials = "JCN"
    elif og_rev[i] == "Andrew N":
        og_initials = "AN"
    elif og_rev[i] == "Adan A":
        og_initials = "AA"
    elif og_rev[i] == "Abby M":
        og_initials = "AM"
    elif og_rev[i] == "Anna R":
        og_initials = "AR"
    elif og_rev[i] == "Joe Nogo":
        og_initials = "JN"
    elif og_rev[i] is None:
        og_initials = "No folder"
    if new_rev[i] == "Josh B":
        new_initials = "JB"
    elif new_rev[i] == "Ethan U":
        new_initials = "EU"
    elif new_rev[i] == "Jose CN":
        new_initials = "JCN"
    elif new_rev[i] == "Andrew N":
        new_initials = "AN"
    elif new_rev[i] == "Adan A":
        new_initials = "AA"
    elif new_rev[i] == "Abby M":
        new_initials = "AM"
    elif new_rev[i] == "Anna R":
        new_initials = "AR"
    elif new_rev[i] == "Joe Nogo":
        new_initials = "JN"
    elif new_rev[i] == "Delete":
        new_initials = "Delete"
    og_folder_path = os.path.join(pathToReviews,og_initials,IA[i]+"-"+og_initials)
    new_folder_path = os.path.join(pathToReviews,new_initials,IA[i]+"-"+new_initials)
    try:
        if og_initials == "No folder":
            print(f'No folder created for {IA[i]}! Therefore no folder moved over.')
            successes += 1
        elif new_initials == "Delete":
            shutil.rmtree(og_folder_path)
            print(f'Deleted {IA[i]}!')
            successes += 1
        else:
            shutil.move(og_folder_path,new_folder_path)
            print(f"Moved {IA[i]}")
            successes += 1
    except:
        print(f"!!!!!!!!!!!!!!!! Issue moving {IA[i]} !!!!!!!!!!!!!!!!!!!!")
        fails += 1

print(f"{successes} folders moved over (if they were created) and {fails} not moved over.")
input("Press enter to exit.")
