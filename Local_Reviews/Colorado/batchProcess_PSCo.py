# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 11/10/2022

import openpyxl
from openpyxl import load_workbook
import os, os.path
from os.path import join
import sys
import numpy as np
import shutil
import warnings
import glob
import win32com.client as win32
import datetime
import time

warnings.filterwarnings("ignore")

# pathToTracker = r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER Work - DER Reviews - PSCo\CR & LV1 Tracker - PSCO.xlsx'
pathToTracker = r'C:\Users\joe.nogosek\Downloads\CR & LV1 Tracker - PSCO - Copy.xlsx'

wb = load_workbook(pathToTracker)
ws = wb['2022 CR']

# Now get all the necessary information from the 2022 tab
IA_num = []
reviewer = []

numReviews = 0
for row in range(2,ws.max_row+1):
    if str(ws[row][11].value) != '' and str(ws[row][12].value) == 'None':
        IA_num.append(ws[row][1].value)
        reviewer.append(ws[row][7].value)
        numReviews += 1

wb.close()
print("Good to open tracker.\n")

xl=win32.Dispatch("Excel.Application")
# tracker = xl.Workbooks.Open(os.path.abspath(pathToTracker))

pathToReadyQC = r'G:\2021\21.10704\Reviews\Ready_for_QC'
files = []
i=0
for i in range(numReviews):
    pause = input("Press enter when ready to process the next one.")
    print(IA_num[i])
    if reviewer[i] == "Joe Nogo":
        files.append(IA_num[i] + "-JN")
    elif reviewer[i] == "Ross K":
        files.append(IA_num[i] + "-RK")
    elif reviewer[i] == "Nick C":
        files.append(IA_num[i] + "-NC")
    elif reviewer[i] == "Jose CN":
        files.append(IA_num[i] + "-JCN")
    elif reviewer[i] == "Ethan U":
        files.append(IA_num[i] + "-EU")
    elif reviewer[i] == "Adan A":
        files.append(IA_num[i] + "-AA")
    elif reviewer[i] == "Josh B":
        files.append(IA_num[i] + "-JB")
    elif reviewer[i] == "Andrew N":
        files.append(IA_num[i] + "-AN")
    elif reviewer[i] == "Jason H":
        files.append(IA_num[i] + "-JH")
    elif reviewer[i] == "Abby M":
        files.append(IA_num[i] + "-AM")
    elif reviewer[i] == "Ed S":
        files.append(IA_num[i] + "-ES")
    elif reviewer[i] == "Josh G":
        files.append(IA_num[i] + "-JG")
    elif reviewer[i] == "Andre B":
        files.append(IA_num[i] + "-AB")
    pathToFile = os.path.join(pathToReadyQC,files[i])
    list_of_files = glob.glob(os.path.join(pathToFile,"*.xlsm")) # * means all if need specific format then *.csv
    latest_file = max(list_of_files, key=os.path.getctime)
    book = xl.Workbooks.Open(os.path.abspath(latest_file))
    try:
        xl.Run("autoSF_CO")
    except:
        pass

del xl
print("This script is finished. Make sure to help all the open Chrome windows finish.")

input("Press enter to exit.")
