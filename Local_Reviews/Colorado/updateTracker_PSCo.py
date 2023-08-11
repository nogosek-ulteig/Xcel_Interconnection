# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 11/30/2021
# Close tracker and make sure G:\ directory is green

from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import glob
import win32com.client as win32
import calendar
import warnings
import datetime
import numpy as np
import shutil
from copy import copy
from decimal import Decimal
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

username = credentials.username
password = credentials.password
pathToTracker = credentials.path_to_CO_CR_tracker
pathToDownloadsFolder = credentials.path_to_downloads
path_to_driver = credentials.path_to_driver
name = credentials.name

sendOption = input("Would you like to send the email? (y/n?): ")

JB_arr = ["Porticade construction LLC","Skyline Solar","ADT Solar","Golden Solar Electric, LLC","Small Town Solar Inc","LGCY Installation Services, LLC","The Solar Revolution","Buglet Solar","Tephra Solar","Peak Solar Designs","Lifetime Energy LLP","Cascade Solar and Electric","1 Solar","Big Dog Renewable Energy","Active Energies Solar, LLC","WattMore","Glyde Solar LLC","Glyde Solar","Progreen Solar","Apollo Energy","Custom Solar","Infinity Energy Inc.","Palmetto Solar","SoCo Solar & Power","SandboAndrew N Solar","Klick Solar","Energy Advantage roof and Solar","Titanium Solar","GAF Energy","Ambia","No Problem Electric","Sunsense Solar","Complete Energy Solutions, LLC","Colorado Energy Systems","BRS Feild Ops LLC","Go Green Electric, Inc.","High Noon Solar & Energy Products","Ideal Home Energy","Unico Solar Investors","LGCY INSTALLATION SERVICES","BriteStreet Solar","Douglass Colony Group","Fluent Solar","Solcius, LLC","Solar Wave","Complete Solar","Blue Sky Solar and Roofing","CAM Solar","Reconnect Clean Energy, LLC","Green Electrical Solutions","Solarise Solar","Advantage Solar","GRID Alternatives","Innovative Energy","SunPower Direct","Roof Check Inc","Sopris Solar LTD","BriteStreet","BRS Field Ops, LLC.","Vanguard Solar Services","CJR Partners LLC dba Sopris Solar","Electrical & Lighting Services Company, LLC","Solar Side Up LLC","Impact Solar","0","REenergizeCO Inc","It's Lit Solar Electric","Southard Solar Energy"]
AN_arr = ["ION Solar","Freedom Forever","Smart Wave Solar LLC","1Solar","SunTalk Solar","Freedom Solar","POWUR, PBC","RTG Investments","Conundrum Technologies","Jacob Johnson","Solar Supply of Colorado, LLC","Roof Check","1st Light Energy Inc.","Solar Design Studio","Alt-E Wind & Solar","Jason Becker","SOPRIS SOLAR","C.A.M. Electric","SoL Energy, LLC","Rural Solar Restoration, LLC","Sandhill Solar","Rural Solar Restoration","Active Energies Solar","Aspen Solar Inc","SoleAndrew N Power, LLC","Colorado Energy Office","EnergyLink LLC","Leadville Solar","Christopher McLeod","Weddle and Sons, Inc","Independent Power","GenPro Energy Solutions, LLC","CAM Electric Inc","Verde Solutions","Elevated Independent Energy","james lewton","Green Home Systems L.L.C.","TELT Ventures dba One Solar","Roper Roofing and Solar","Total Solar Solutions","EAndrew Npert Electric, LLC","Ventara, LLC.","Beautifi Solar","Pivot Energy","LGCY INSTALLATION SERVICES LLC","Liv Solar","Freedom Forever CO, LLC","Project Solar","IPOWER Alliance, LLC","  ","Daniela Andrade","Greenstar Power","Covenant Solar Tech","Empower Energy Solutions Inc","Freedom Forever Solar LLC","Custom Solar LLC","TELT Ventures LLC dba 1Solar","Matt Palen","Vivint Solar Developer, LLC","EAndrew Ncel Construction Group","Mile High Solar Guy","Rise Power LLC","Advantage Solar LLC DBA: Solarise Solar","Grand Valley Solar","Soco solar power","Avolta Power Inc","JP Electric","Green Cardinal Energy"]
AA_arr = ["Sunrun Installation Services","Klick Solar LLC","Photon Brothers","Generation Solar","Tesla Energy Operations, Inc.","Premier Renewables","Solcius LLC","Infinity Energy Inc","Freedom Solar Power","BRS Field Ops, LLC","Solar Power Pros Inc.","NeAndrew Nt Energy Solar","SUNRUN INSTALLATION SERVICES INC","ARE Solar","Pro Bid Energy","Glyde Solar, LLC","SUNRUN","Tesla Energy Operations"]
AR_arr = ["Porticade Construction DBA Action Solar","Namaste Solar","Solcius","Blue Raven Solar","Titan Solar Power","Namaste Solar Electric, Inc.","Atlasta Solar Center, LLC","Independent Power Systems, Inc.","TruSun Energy","Small Town Solar","Top Electric","Elevated Independent Energy LLC","Mountain Power Solutions","ALE Energy","Sunsense Inc","Altitude Solar","Covenant Solar Tech LLC","Southard Solar","Empower Energies","IoniAndrew N Smart Solutions","Ridgeline Electric & Solar","Sun Valley Electric","Empower Energy Solutions Inc.","Retired","Green Home Systems","Constructive Alternatives","James Bryner","Sunnyside Solar","AST Solae","CEG Solutions","Avolta Power","Dean Reitz Construction","Dark Forest Solar","MEGA HOMES LLC","Peak to Peak Roofing & EAndrew Nteriors, LLC","Isaac Jeffs","BecauseSolar","South Plains Solar, LLC","Lone Star Solar Services LLC","Michelle Orwick"]

warnings.filterwarnings("ignore", category=DeprecationWarning)

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

driver = webdriver.Chrome(executable_path=path_to_driver, options=options)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(10)

try:
    driver.find_element("id", "idBtn_Back").click()
    time.sleep(3)
except:
    pass

time.sleep(3)
# Click the report tab
try:
    for i in range(1,10):
        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[2]/div[1]/div[1]/nav[1]/ul[1]/li[{str(i)}]/a[1]").get_attribute("innerText") == 'Reports':
            driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[2]/div[1]/div[1]/nav[1]/ul[1]/li[{str(i)}]/a[1]").click()
            break
except:
    pass

time.sleep(5)
try:
    try:
        for i in range(1,20):
            if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").get_attribute("innerText") == 'Joe_PSCo_CR_Report':
                driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").click()
                break
    except:
        pass

    time.sleep(3)
    driver.find_element("css selector", "[title*='Export Details']").click()
except:
    time.sleep(180)
    try:
        for i in range(1,20):
            if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").get_attribute("innerText") == 'Joe_PSCo_CR_Report':
                driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[1]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/form[1]/div[3]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[{str(i)}]/table[1]/tbody[1]/tr[1]/td[3]/div[1]/div[2]/a[1]/span[1]").click()
                break
    except:
        pass

    time.sleep(3)
    driver.find_element("css selector", "[title*='Export Details']").click()

driver.find_element("css selector", "[title*='Export']").click()
time.sleep(5)

# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

# Define paths to report and convert to correct file format
if not os.path.exists(os.path.join(pathToDownloadsFolder, str(glob.glob(os.path.join(pathToDownloadsFolder,'report*'))[0]))):
    raise ValueError('No report file present in the Completeness Review Reports folder')
else:
    pathToReportxls = os.path.join(pathToDownloadsFolder, str(glob.glob(os.path.join(pathToDownloadsFolder,'report*'))[0]))
    pathToReportxlsx = pathToReportxls + 'x'

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(pathToReportxls)
    wb.SaveAs(pathToReportxlsx, FileFormat = 51) # 51 corresponds to the .xlsx extension file format. 56 would be for .xls
    wb.Close()
    os.remove(pathToReportxls)

# For email creation purposes
num=0
list_to_email = []

# Load in the new report file just created and current tracker and set some values which will be useful
wbReport = load_workbook(pathToReportxlsx)
wbTracker = load_workbook(pathToTracker)

wsReport = wbReport.active
row_max_report = wsReport.max_row - 6

IAnums_tracker = []
existingMeters_tracker = []
premises_tracker = []
dueDates_tracker = []
types_tracker = []
developers_tracker = []
history_tracker = []
lineDistances_tracker = []
AC_sizes_tracker = []
DC_sizes_tracker = []
submittal_date_tracker = []
QC_tracker = []
feederNums_tracker = []
ground_ref_tracker = []
# Go through all the sheets in the tracker to get all the necessary information in lists
for sheet in wbTracker.sheetnames:
    wsTracker = wbTracker[sheet]
    # Get all the IA numbers from this sheet in the book
    for row in wsTracker['B']:
        if row.row == 1:
            pass
        else:
            IAnums_tracker.append(row.value)
            #print(f"IA Num: {row.value}")
    # Get all the premises from this sheet in the book
    for row in wsTracker['C']:
        if row.row == 1:
            pass
        else:
            premises_tracker.append(row.value)
            #print(f"Existing Meter: {row.value}")
    # Get all the existing meter types from this sheet in the book
    for row in wsTracker['D']:
        if row.row == 1:
            pass
        else:
            existingMeters_tracker.append(row.value)
            #print(f"Existing Meter: {row.value}")
    # Get all the due dates from this sheet in the book
    for row in wsTracker['N']:
        if row.row == 1:
            pass
        else:
            dueDates_tracker.append(row.value)
            #print(f"Due date: {row.value}")
    # Get all the review types (ESS or not) from this sheet in the book
    for row in wsTracker['E']:
        if row.row == 1:
            pass
        else:
            types_tracker.append(row.value)
            #print(f"Type: {row.value}")
    # Get all the developers from this sheet in the book
    for row in wsTracker['J']:
        if row.row == 1:
            pass
        else:
            developers_tracker.append(row.value)
            #print(f"Developer: {row.value}")
    # Get all the approval/rejection from this sheet in the book
    for row in wsTracker['L']:
        if row.row == 1:
            pass
        else:
            history_tracker.append(row.value)
            #print(f"History: {row.value}")
    # Get all the line distances from this sheet in the book
    for row in wsTracker['F']:
        if row.row == 1:
            pass
        else:
            lineDistances_tracker.append(row.value)
            #print(f"Line Distance: {row.value}")
    # Get all the system sizes from this sheet in the book
    for row in wsTracker['G']:
        if row.row == 1:
            pass
        else:
            AC_sizes_tracker.append(row.value)
            #print(f"Size: {row.value}")
    # Get all the submittal dates from this sheet in the book
    for row in wsTracker['M']:
        if row.row == 1:
            pass
        else:
            submittal_date_tracker.append(row.value)
            #print(f"Submittal date: {row.value}")
    # Get all the QC info from this sheet in the book
    for row in wsTracker['K']:
        if row.row == 1:
            pass
        else:
            QC_tracker.append(row.value)
    for row in wsTracker['A']:
        if row.row == 1:
            pass
        else:
            feederNums_tracker.append(row.value)
    for row in wsTracker['Q']:
        if row.row == 1:
            pass
        else:
            DC_sizes_tracker.append(row.value)
    for row in wsTracker['P']:
        if row.row == 1:
            pass
        else:
            ground_ref_tracker.append(row.value)

# Get all the information from the report in lists
# Initialize necessary lists
IAnums_report = []
caseNums_report = []
developers_report = []
dueDates_report = []
statuses_report = []
AC_sizes_report = []
DC_sizes_report = []
types_report = []
premise_report = []
programType_report = []
feederNums_report = []

# Get the IA numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=2, max_col=2, values_only=True):
    IAnums_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Get the premise numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=3, max_col=3, values_only=True):
    premise_report.append(row[0])
    #print(f"IA Num: {row[0]}")
# Extract just the premise number
for premise in range(len(premise_report)):
    if type(premise_report[premise]) != type(None):
        premise_report[premise] = premise_report[premise][-11:]
        premise_report[premise] = premise_report[premise][:9]
        if premise_report[premise] == 'der Premi':
            premise_report[premise] = 'New Construction'
# Get the case numbers from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=6, max_col=6, values_only=True):
    caseNums_report.append(row[0])
    #print(f"Case num: {row[0]}")
# Get all the developers from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=7, max_col=7, values_only=True):
    developers_report.append(row[0])
    #print(f"Developer: {row[0]}")
# Get the due dates from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=1, max_col=1, values_only=True):
    dueDates_report.append(row[0])
    #print(f"Due Date: {row[0]}")
# Get all the DC sizes from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=9, max_col=9, values_only=True):
    DC_sizes_report.append(row[0])
    #print(f"Status: {row[0]}")
# Get the application sizes from the report into a list
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=5, max_col=5, values_only=True):
    AC_sizes_report.append(row[0])
    #print(f"Size: {row[0]}")
# Get all the application types (ESS or not) from this sheet in the book
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=4, max_col=4, values_only=True):
    types_report.append(row[0])
    #print(f"Type: {row[0]}")
# Get the program types
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=13, max_col=13, values_only=True):
    programType_report.append(row[0])
    #print(f"Type: {row[0]}")
# Get the feeder numbers
for row in wsReport.iter_rows(min_row=2, max_row=row_max_report, min_col=19, max_col=19, values_only=True):
    feederNums_report.append(row[0])
    #print(f"Type: {row[0]}")

# Create styles
redFill = PatternFill(start_color='FFFF0000', end_color='FF0000', fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF', end_color='000000', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFF00', fill_type='solid')
greyFill = PatternFill(start_color='FFF2F2F2', end_color='F2F2F2', fill_type='solid')
greenFill = PatternFill(start_color = 'FFA9D08E', end_color = 'A9D08E', fill_type='solid')
salmonFill = PatternFill(start_color = 'FFF4B084', end_color = 'F4B084', fill_type='solid')

row_max_tracker = wsTracker.max_row
wsTracker['O' + str(row_max_tracker)] = ''
next_row_tracker = row_max_tracker + 1

amountNewReviews = 0
# Count the total number of reviews to be added
for w in range(row_max_report-1):
    if (IAnums_report[w] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])] == 'Rejected') and (dueDates_report[w] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[w])]) and (submittal_date_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] != None):
        amountNewReviews += 1
    elif (IAnums_report[w] not in IAnums_tracker):
        amountNewReviews += 1

# Number of times we have made it through
k = 1
# flipFlop = 0
for i in range(row_max_report-1):
    if (IAnums_report[i] in IAnums_tracker) and (history_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Rejected') and (dueDates_report[i] != dueDates_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])]) and (submittal_date_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] != None):
        if AC_sizes_report[i] is None:
            AC_sizes_report[i] = 0
        if DC_sizes_report[i] is None:
            DC_sizes_report[i] = 0
        wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
        if str(feederNums_report[i]) == 'None':
            wsTracker['A' + str(next_row_tracker)].fill = salmonFill
        else:
            wsTracker['A' + str(next_row_tracker)].fill = whiteFill
        wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
        wsTracker['B' + str(next_row_tracker)].fill = whiteFill
        wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
        wsTracker['C' + str(next_row_tracker)].fill = whiteFill
        wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
        wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
        wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] >= 25:
             wsTracker['F' + str(next_row_tracker)].fill = redFill
        else:
             wsTracker['F' + str(next_row_tracker)].fill = greyFill
        wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] < 100:
             wsTracker['G' + str(next_row_tracker)].fill = yellowFill
        else:
            wsTracker['G' + str(next_row_tracker)].fill = whiteFill
        wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
        wsTracker['H' + str(next_row_tracker)].fill = whiteFill
        wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
        wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
        wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
        wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
        wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
        wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
        wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
        wsTracker['P' + str(next_row_tracker)]._style = copy(wsTracker['P' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] >= 100:
            wsTracker['P' + str(next_row_tracker)].fill = greenFill
        else:
            wsTracker['P' + str(next_row_tracker)].fill = greyFill
        wsTracker['Q' + str(next_row_tracker)]._style = copy(wsTracker['Q' + str(row_max_tracker)]._style)
        wsTracker['R' + str(next_row_tracker)]._style = copy(wsTracker['R' + str(row_max_tracker)]._style)
        wsTracker['S' + str(next_row_tracker)]._style = copy(wsTracker['S' + str(row_max_tracker)]._style)
        wsTracker['T' + str(next_row_tracker)]._style = copy(wsTracker['T' + str(row_max_tracker)]._style)
        wsTracker['U' + str(next_row_tracker)]._style = copy(wsTracker['U' + str(row_max_tracker)]._style)
        dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
        if feederNums_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] != '':
            wsTracker['A' + str(next_row_tracker)] = feederNums_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])]
        wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
        wsTracker['C' + str(next_row_tracker)] = premise_report[i]
        wsTracker['D' + str(next_row_tracker)] = 'Verify Complete'
        wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
        if types_report[i] == 'Yes' and str(programType_report[i]) == "Renewable Battery Connect":
            #print('Reassign ' + str(IAnums_report[i]) + ' to metering (ESS)')
            wsTracker['E' + str(next_row_tracker)] = 'RBC'
        elif types_report[i] == 'Yes':
            wsTracker['E' + str(next_row_tracker)] = 'ESS'
        if lineDistances_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] == 'Complete':
            wsTracker['F' + str(next_row_tracker)] = 'Complete'
        wsTracker['G' + str(next_row_tracker)] = AC_sizes_report[i]
        if AC_sizes_report[i] < 100:
            wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
        else:
            #print('Reassign ' + str(IAnums_report[i]) + ' to metering (Larger than 40 kW)')
            wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
        if developers_report[i] in JB_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Josh B'
        elif developers_report[i] in AN_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Andrew N'
        elif developers_report[i] in AA_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Adan A'
        elif developers_report[i] in AR_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Anna R'
        else:
            wsTracker['H' + str(next_row_tracker)] = 'Andrew N'
        wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
        wsTracker['J' + str(next_row_tracker)] = developers_report[i]
        wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
        if ground_ref_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])] != '':
            wsTracker['P' + str(next_row_tracker)] = ground_ref_tracker[len(IAnums_tracker) - 1 - IAnums_tracker[::-1].index(IAnums_report[i])]
        wsTracker['Q' + str(next_row_tracker)] = DC_sizes_report[i]
        wsTracker['T' + str(next_row_tracker)] = 'x'
        next_row_tracker = next_row_tracker + 1
        k += 1
    elif (IAnums_report[i] not in IAnums_tracker):
        if AC_sizes_report[i] is None:
            AC_sizes_report[i] = 0
        if DC_sizes_report[i] is None:
            DC_sizes_report[i] = 0
        wsTracker['A' + str(next_row_tracker)]._style = copy(wsTracker['A' + str(row_max_tracker)]._style)
        if str(feederNums_report[i]) == 'None':
            wsTracker['A' + str(next_row_tracker)].fill = salmonFill
        else:
            wsTracker['A' + str(next_row_tracker)].fill = whiteFill
        wsTracker['B' + str(next_row_tracker)]._style = copy(wsTracker['B' + str(row_max_tracker)]._style)
        wsTracker['B' + str(next_row_tracker)].fill = whiteFill
        wsTracker['C' + str(next_row_tracker)]._style = copy(wsTracker['C' + str(row_max_tracker)]._style)
        wsTracker['C' + str(next_row_tracker)].fill = whiteFill
        wsTracker['D' + str(next_row_tracker)]._style = copy(wsTracker['D' + str(row_max_tracker)]._style)
        wsTracker['E' + str(next_row_tracker)]._style = copy(wsTracker['E' + str(row_max_tracker)]._style)
        wsTracker['F' + str(next_row_tracker)]._style = copy(wsTracker['F' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] >= 25:
         wsTracker['F' + str(next_row_tracker)].fill = redFill
        else:
             wsTracker['F' + str(next_row_tracker)].fill = greyFill
        wsTracker['G' + str(next_row_tracker)]._style = copy(wsTracker['G' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] < 100:
            wsTracker['G' + str(next_row_tracker)].fill = yellowFill
        else:
            wsTracker['G' + str(next_row_tracker)].fill = whiteFill
        wsTracker['H' + str(next_row_tracker)]._style = copy(wsTracker['H' + str(row_max_tracker)]._style)
        wsTracker['H' + str(next_row_tracker)].fill = whiteFill
        wsTracker['I' + str(next_row_tracker)]._style = copy(wsTracker['I' + str(row_max_tracker)]._style)
        wsTracker['J' + str(next_row_tracker)]._style = copy(wsTracker['J' + str(row_max_tracker)]._style)
        wsTracker['K' + str(next_row_tracker)]._style = copy(wsTracker['K' + str(row_max_tracker)]._style)
        wsTracker['L' + str(next_row_tracker)]._style = copy(wsTracker['L' + str(row_max_tracker)]._style)
        wsTracker['M' + str(next_row_tracker)]._style = copy(wsTracker['M' + str(row_max_tracker)]._style)
        wsTracker['N' + str(next_row_tracker)]._style = copy(wsTracker['N' + str(row_max_tracker)]._style)
        wsTracker['O' + str(next_row_tracker)]._style = copy(wsTracker['O' + str(row_max_tracker)]._style)
        if AC_sizes_report[i] >= 100:
            wsTracker['P' + str(next_row_tracker)].fill = greenFill
        else:
            wsTracker['P' + str(next_row_tracker)].fill = greyFill
        wsTracker['Q' + str(next_row_tracker)]._style = copy(wsTracker['Q' + str(row_max_tracker)]._style)
        wsTracker['R' + str(next_row_tracker)]._style = copy(wsTracker['R' + str(row_max_tracker)]._style)
        wsTracker['S' + str(next_row_tracker)]._style = copy(wsTracker['S' + str(row_max_tracker)]._style)
        wsTracker['T' + str(next_row_tracker)]._style = copy(wsTracker['T' + str(row_max_tracker)]._style)
        wsTracker['U' + str(next_row_tracker)]._style = copy(wsTracker['U' + str(row_max_tracker)]._style)
        dueDate = datetime.datetime.strptime(str(dueDates_report[i]), '%Y-%m-%d %H:%M:%S')
        if feederNums_report[i] == '-':
            wsTracker['A' + str(next_row_tracker)].value = ''
        else:
            wsTracker['A' + str(next_row_tracker)].value = feederNums_report[i]
        wsTracker['B' + str(next_row_tracker)] = IAnums_report[i]
        wsTracker['C' + str(next_row_tracker)] = premise_report[i]
        if premise_report[i] == 'New Construction':
            list_to_email.append(IAnums_report[i])
            num += 1
        if DC_sizes_report[i] > 20:
            wsTracker['D' + str(next_row_tracker)] = 'N/A'
            wsTracker['D' + str(next_row_tracker)].font = Font(bold=False)
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(str(IAnums_report[i]))
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Change the name of Area and meter engineer (if applicable)
            driver.find_element("css selector", "[title='Edit']").click()
            time.sleep(3)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[9]/table/tbody/tr[5]/td[2]/div/span/input").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[2]/div[9]/table/tbody/tr[5]/td[2]/div/span/input").send_keys("David Wynkoop")
            print('Reassigned ' + str(IAnums_report[i]) + ' to David Wynkoop')
            driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/form/div/div[1]/table/tbody/tr/td[2]/input[1]").click()
            time.sleep(5)
        else:
            wsTracker['D' + str(next_row_tracker)] = '2S'
            wsTracker['D' + str(next_row_tracker)].font = Font(bold=True)
        if types_report[i] == 'Yes' and str(programType_report[i]) == "Renewable Battery Connect":
            #print('Reassign ' + str(IAnums_report[i]) + ' to metering (ESS)')
            wsTracker['E' + str(next_row_tracker)] = 'RBC'
        elif types_report[i] == 'Yes':
            wsTracker['E' + str(next_row_tracker)] = 'ESS'
        wsTracker['G' + str(next_row_tracker)] = AC_sizes_report[i]
        if AC_sizes_report[i] < 100:
            wsTracker['G' + str(next_row_tracker)].font = Font(bold=True)
        else:
            wsTracker['G' + str(next_row_tracker)].font = Font(bold=False)
        if developers_report[i] in JB_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Josh B'
        elif developers_report[i] in AN_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Andrew N'
        elif developers_report[i] in AA_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Adan A'
        elif developers_report[i] in AR_arr:
            wsTracker['H' + str(next_row_tracker)] = 'Anna R'
        else:
            wsTracker['H' + str(next_row_tracker)] = 'Andrew N'
        wsTracker['I' + str(next_row_tracker)] = (caseNums_report[i])[1:]
        wsTracker['J' + str(next_row_tracker)] = developers_report[i]
        wsTracker['N' + str(next_row_tracker)] = dueDates_report[i]
        wsTracker['Q' + str(next_row_tracker)] = DC_sizes_report[i]
        wsTracker['T' + str(next_row_tracker)] = 'x'
        next_row_tracker = next_row_tracker + 1
        k += 1

# Close the SF webpage
driver.close()

time = datetime.datetime.today().strftime("%I:%M %p")
date = datetime.date.strftime(datetime.date.today(), "%m/%d/%Y")
wsTracker['O' + str(next_row_tracker - 1)] = 'Tracker updated as of ' + str(time) + ' MST on ' + str(date)
# Save the Excel file
wbTracker.save(pathToTracker)

list_to_email_string = ""
for i in range(len(list_to_email)):
    list_to_email_string += list_to_email[i] + "\n"

if num >= 1:
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = "DER_IntegrationCO@xcelenergy.com"
    mail.CC = "andrew.norman@ulteig.com; josh.berg@ulteig.com"
    mail.BCC = "joe.nogosek@ulteig.com"
    mail.Subject = "Interconnection Application(s) Missing Premise Info"
    mail.Body = "Hi,\n\nCan you please populate the premise field and the address for the following application(s)?\n\n" + str(list_to_email_string) + "\nThanks!\n" + str(name)
    if sendOption.lower() == "y":
        mail.send
        print("Email sent.")
    else:
        pass
        print("No email sent.")
else:
    print("No need to send email.")

wbTracker.close()
wbReport.close()

# Delete the report so the script is ready for next time
os.remove(pathToReportxlsx)

input("Press enter to exit.")
