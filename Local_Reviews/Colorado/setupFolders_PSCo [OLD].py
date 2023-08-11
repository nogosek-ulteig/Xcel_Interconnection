# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 10/26/2021
# NOTE: Tracker must be closed on your local computer!

import openpyxl
from openpyxl import load_workbook
import os
import numpy as np
import shutil
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import glob
from copy import copy
import win32com.client as win32

warnings.filterwarnings('ignore')

email = 'joseph.h.nogosek@xcelenergy.com'
username = '239665'
password = 'airdoc6Ee'
pin = '2846'

# Now get the newly added reviews ready for me to complete
IA_num_Nick = []
case_num_Nick = []
size_Nick = []
program_Nick = []
type_Nick = []
installer_Nick = []
due_Date_Nick = []
sizeDC_Nick = []
meter_Nick = []

IA_num_Ethan = []
case_num_Ethan = []
size_Ethan = []
program_Ethan = []
type_Ethan = []
installer_Ethan = []
due_Date_Ethan = []
sizeDC_Ethan = []
meter_Ethan = []

IA_num_Joe = []
case_num_Joe = []
size_Joe = []
program_Joe = []
type_Joe = []
installer_Joe = []
due_Date_Joe = []
sizeDC_Joe = []
meter_Joe = []

IA_num_Jose = []
case_num_Jose = []
size_Jose = []
program_Jose = []
type_Jose = []
installer_Jose = []
due_Date_Jose = []
sizeDC_Jose = []
meter_Jose = []

IA_num_Ross = []
case_num_Ross = []
size_Ross = []
program_Ross = []
type_Ross = []
installer_Ross = []
due_Date_Ross = []
sizeDC_Ross = []
meter_Ross = []

IA_num_Josh = []
case_num_Josh = []
size_Josh = []
program_Josh = []
type_Josh = []
installer_Josh = []
due_Date_Josh = []
sizeDC_Josh = []
meter_Josh = []

IA_num_Adan = []
case_num_Adan = []
size_Adan = []
program_Adan = []
type_Adan = []
installer_Adan = []
due_Date_Adan = []
sizeDC_Adan = []
meter_Adan = []

IA_num_Ed = []
case_num_Ed = []
size_Ed = []
program_Ed = []
type_Ed = []
installer_Ed = []
due_Date_Ed = []
sizeDC_Ed = []
meter_Ed = []

IA_num_Abby = []
case_num_Abby = []
size_Abby = []
program_Abby = []
type_Abby = []
installer_Abby = []
due_Date_Abby = []
sizeDC_Abby = []
meter_Abby = []

IA_num_Andre = []
case_num_Andre = []
size_Andre = []
program_Andre = []
type_Andre = []
installer_Andre = []
due_Date_Andre = []
sizeDC_Andre = []
meter_Andre = []

IA_num_Jason = []
case_num_Jason = []
size_Jason = []
program_Jason = []
type_Jason = []
installer_Jason = []
due_Date_Jason = []
sizeDC_Jason = []
meter_Jason = []

IA_num_Andrew = []
case_num_Andrew = []
size_Andrew = []
program_Andrew = []
type_Andrew = []
installer_Andrew = []
due_Date_Andrew = []
sizeDC_Andrew = []
meter_Andrew = []

IA_num_JoshG = []
case_num_JoshG = []
size_JoshG = []
program_JoshG = []
type_JoshG = []
installer_JoshG = []
due_Date_JoshG = []
sizeDC_JoshG = []
meter_JoshG = []

IA_num_Anna = []
case_num_Anna = []
size_Anna = []
program_Anna = []
type_Anna = []
installer_Anna = []
due_Date_Anna = []
sizeDC_Anna = []
meter_Anna = []

# Open up Tracker and go to the newewst tab
# C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER Work - DER Reviews - PSCo\CR & LV1 Tracker - PSCO.xlsx
wb = load_workbook(r'C:\Users\joe.nogosek\Ulteig Engineers, Inc\Xcel DER Work - DER Reviews - PSCo\CO Completeness Review Tracker.xlsx')
# wb = load_workbook(r'C:\Users\joe.nogosek\Downloads\CO Completeness Review Tracker - Copy3.xlsx')
ws = wb['2023']

# Nick
i=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Nick C" and ws[row][10].value==None:
        IA_num_Nick.append(ws[row][1].value)
        case_num_Nick.append(ws[row][8].value)
        size_Nick.append(ws[row][6].value)
        program_Nick.append(ws[row][15].value)
        type_Nick.append(ws[row][4].value)
        installer_Nick.append(ws[row][9].value)
        due_Date_Nick.append(ws[row][13].value)
        sizeDC_Nick.append(ws[row][16].value)
        meter_Nick.append(ws[row][3].value)
        i=i+1

# Ethan
j=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Ethan U" and ws[row][10].value==None:
        IA_num_Ethan.append(ws[row][1].value)
        case_num_Ethan.append(ws[row][8].value)
        size_Ethan.append(ws[row][6].value)
        program_Ethan.append(ws[row][15].value)
        type_Ethan.append(ws[row][4].value)
        installer_Ethan.append(ws[row][9].value)
        due_Date_Ethan.append(ws[row][13].value)
        sizeDC_Ethan.append(ws[row][16].value)
        meter_Ethan.append(ws[row][3].value)
        j=j+1

# Joe
k=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Joe Nogo" and ws[row][10].value==None:
        IA_num_Joe.append(ws[row][1].value)
        case_num_Joe.append(ws[row][8].value)
        size_Joe.append(ws[row][6].value)
        program_Joe.append(ws[row][15].value)
        type_Joe.append(ws[row][4].value)
        installer_Joe.append(ws[row][9].value)
        due_Date_Joe.append(ws[row][13].value)
        sizeDC_Joe.append(ws[row][16].value)
        meter_Joe.append(ws[row][3].value)
        k=k+1

# Jose
l=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Jose CN" and ws[row][10].value==None:
        IA_num_Jose.append(ws[row][1].value)
        case_num_Jose.append(ws[row][8].value)
        size_Jose.append(ws[row][6].value)
        program_Jose.append(ws[row][15].value)
        type_Jose.append(ws[row][4].value)
        installer_Jose.append(ws[row][9].value)
        due_Date_Jose.append(ws[row][13].value)
        sizeDC_Jose.append(ws[row][16].value)
        meter_Jose.append(ws[row][3].value)
        l=l+1

# Ross
m=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Ross K" and ws[row][10].value==None:
        IA_num_Ross.append(ws[row][1].value)
        case_num_Ross.append(ws[row][8].value)
        size_Ross.append(ws[row][6].value)
        program_Ross.append(ws[row][15].value)
        type_Ross.append(ws[row][4].value)
        installer_Ross.append(ws[row][9].value)
        due_Date_Ross.append(ws[row][13].value)
        sizeDC_Ross.append(ws[row][16].value)
        meter_Ross.append(ws[row][3].value)
        m=m+1

# Josh
p=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Josh B" and ws[row][10].value==None:
        IA_num_Josh.append(ws[row][1].value)
        case_num_Josh.append(ws[row][8].value)
        size_Josh.append(ws[row][6].value)
        program_Josh.append(ws[row][15].value)
        type_Josh.append(ws[row][4].value)
        installer_Josh.append(ws[row][9].value)
        due_Date_Josh.append(ws[row][13].value)
        sizeDC_Josh.append(ws[row][16].value)
        meter_Josh.append(ws[row][3].value)
        p=p+1

# Adan
q=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Adan A" and ws[row][10].value==None:
        IA_num_Adan.append(ws[row][1].value)
        case_num_Adan.append(ws[row][8].value)
        size_Adan.append(ws[row][6].value)
        program_Adan.append(ws[row][15].value)
        type_Adan.append(ws[row][4].value)
        installer_Adan.append(ws[row][9].value)
        due_Date_Adan.append(ws[row][13].value)
        sizeDC_Adan.append(ws[row][16].value)
        meter_Adan.append(ws[row][3].value)
        q=q+1

# Ed
r=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Ed S" and ws[row][10].value==None:
        IA_num_Ed.append(ws[row][1].value)
        case_num_Ed.append(ws[row][8].value)
        size_Ed.append(ws[row][6].value)
        program_Ed.append(ws[row][15].value)
        type_Ed.append(ws[row][4].value)
        installer_Ed.append(ws[row][9].value)
        due_Date_Ed.append(ws[row][13].value)
        sizeDC_Ed.append(ws[row][16].value)
        meter_Ed.append(ws[row][3].value)
        r=r+1

# Abby
s=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Abby M" and ws[row][10].value==None:
        IA_num_Abby.append(ws[row][1].value)
        case_num_Abby.append(ws[row][8].value)
        size_Abby.append(ws[row][6].value)
        program_Abby.append(ws[row][15].value)
        type_Abby.append(ws[row][4].value)
        installer_Abby.append(ws[row][9].value)
        due_Date_Abby.append(ws[row][13].value)
        sizeDC_Abby.append(ws[row][16].value)
        meter_Abby.append(ws[row][3].value)
        s=s+1

# Andre
t=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Andre B" and ws[row][10].value==None:
        IA_num_Andre.append(ws[row][1].value)
        case_num_Andre.append(ws[row][8].value)
        size_Andre.append(ws[row][6].value)
        program_Andre.append(ws[row][15].value)
        type_Andre.append(ws[row][4].value)
        installer_Andre.append(ws[row][9].value)
        due_Date_Andre.append(ws[row][13].value)
        sizeDC_Andre.append(ws[row][16].value)
        meter_Andre.append(ws[row][3].value)
        t=t+1

# Jason
u=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Jason H" and ws[row][10].value==None:
        IA_num_Jason.append(ws[row][1].value)
        case_num_Jason.append(ws[row][8].value)
        size_Jason.append(ws[row][6].value)
        program_Jason.append(ws[row][15].value)
        type_Jason.append(ws[row][4].value)
        installer_Jason.append(ws[row][9].value)
        due_Date_Jason.append(ws[row][13].value)
        sizeDC_Jason.append(ws[row][16].value)
        meter_Jason.append(ws[row][3].value)
        u=u+1

# Andrew
v=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Andrew N" and ws[row][10].value==None:
        IA_num_Andrew.append(ws[row][1].value)
        case_num_Andrew.append(ws[row][8].value)
        size_Andrew.append(ws[row][6].value)
        program_Andrew.append(ws[row][15].value)
        type_Andrew.append(ws[row][4].value)
        installer_Andrew.append(ws[row][9].value)
        due_Date_Andrew.append(ws[row][13].value)
        sizeDC_Andrew.append(ws[row][16].value)
        meter_Andrew.append(ws[row][3].value)
        v=v+1

# JoshG
w=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Josh G" and ws[row][10].value==None:
        IA_num_JoshG.append(ws[row][1].value)
        case_num_JoshG.append(ws[row][8].value)
        size_JoshG.append(ws[row][6].value)
        program_JoshG.append(ws[row][15].value)
        type_JoshG.append(ws[row][4].value)
        installer_JoshG.append(ws[row][9].value)
        due_Date_JoshG.append(ws[row][13].value)
        sizeDC_JoshG.append(ws[row][16].value)
        meter_JoshG.append(ws[row][3].value)
        w=w+1

# Anna
x=0
for row in range(2,ws.max_row+1):
    if ws[row][7].value=="Anna R" and ws[row][10].value==None:
        IA_num_Anna.append(ws[row][1].value)
        case_num_Anna.append(ws[row][8].value)
        size_Anna.append(ws[row][6].value)
        program_Anna.append(ws[row][15].value)
        type_Anna.append(ws[row][4].value)
        installer_Anna.append(ws[row][9].value)
        due_Date_Anna.append(ws[row][13].value)
        sizeDC_Anna.append(ws[row][16].value)
        meter_Anna.append(ws[row][3].value)
        x=x+1

wb.close()
print("You can open the tracker now.")

#twoFA = input("Enter six digit 2FA code: ")

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Downloads\Python\chromedriver.exe", options = options)

driver.maximize_window()
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(10)
driver.find_element("id", 'i0116').send_keys(email)
driver.find_element("id", 'idSIButton9').click()
time.sleep(3)
driver.find_element("id", 'i0118').send_keys(password)
driver.find_element("id", 'idSIButton9').click()

#passcode_box = driver.find_element_by_name('pf.pass')
#passcode_box.send_keys(pin + str(twoFA))
time.sleep(15)
#submit_button = driver.find_element("xpath", "//button[contains(@onclick,'postOk')]").click()
try:
    driver.find_element("id", 'idSIButton9').click()
    time.sleep(4)
except:
    pass
# Suppress all warning messages (This is bad practice, but I think I know what I am doing lol)
warnings.filterwarnings('ignore')

counter = 0
total = 0

print("Populating Josh's folder.")
# Josh
for value in range(p):
    try:
        total += 1
        name = str(IA_num_Josh[value])
        type = str(type_Josh[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\JB\{}-JB'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Josh[value]
            ws['I1'].value = case_num_Josh[value]
            ws['I4'].value = due_Date_Josh[value]
            if type_Josh[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Josh[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            if type_Josh[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)
            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

            if size_Josh[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Adan's folder.")
# Adan
for value in range(q):
    try:
        total += 1
        name = str(IA_num_Adan[value])
        type = str(type_Adan[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\AA\{}-AA'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Adan[value]
            ws['I1'].value = case_num_Adan[value]
            ws['I4'].value = due_Date_Adan[value]
            if type_Adan[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Adan[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Adan[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Adan[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Adan[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Joe's folder.")
# Joe
for value in range(k):
    try:
        total += 1
        name = str(IA_num_Joe[value])
        type = str(type_Joe[value])
        size = float(size_Joe[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\JN\{}-JN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Joe[value]
            ws['I1'].value = case_num_Joe[value]
            ws['I4'].value = due_Date_Joe[value]
            if type_Joe[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Joe[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Joe[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Joe[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Joe[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Now populating Nick's folder.")
# Nick
for value in range(i):
    try:
        total += 1
        name = str(IA_num_Nick[value])
        type = str(type_Nick[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\NC\{}-NC'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Nick[value]
            ws['I1'].value = case_num_Nick[value]
            ws['I4'].value = due_Date_Nick[value]
            if type_Nick[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Nick[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Nick[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Nick[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Nick[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ethan's folder.")
# Ethan
for value in range(j):
    try:
        total += 1
        name = str(IA_num_Ethan[value])
        type = str(type_Ethan[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\EU\{}-EU'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Ethan[value]
            ws['I1'].value = case_num_Ethan[value]
            ws['I4'].value = due_Date_Ethan[value]
            if type_Ethan[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Ethan[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Ethan[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Ethan[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Ethan[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("And now Jose (if he has any).")
# Jose
for value in range(l):
    try:
        total += 1
        name = str(IA_num_Jose[value])
        type = str(type_Jose[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\JCN\{}-JCN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Jose[value]
            ws['I1'].value = case_num_Jose[value]
            ws['I4'].value = due_Date_Jose[value]
            if type_Jose[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Jose[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Jose[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Jose[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Jose[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue
print("Now populating Ross's folder.")
# Ross
for value in range(m):
    try:
        total += 1
        name = str(IA_num_Ross[value])
        type = str(type_Ross[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\RK\{}-RK'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Ross[value]
            ws['I1'].value = case_num_Ross[value]
            ws['I4'].value = due_Date_Ross[value]
            if type_Ross[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Ross[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Ross[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Ross[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Ross[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Ed's folder.")
# Ed
for value in range(r):
    try:
        total += 1
        name = str(IA_num_Ed[value])
        type = str(type_Ed[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\ES\{}-ES'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Ed[value]
            ws['I1'].value = case_num_Ed[value]
            ws['I4'].value = due_Date_Ed[value]
            if type_Ed[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Ed[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Ed[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Ed[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

                wb.save(new_file)

            elif size_Ed[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Abby's folder.")
# Abby
for value in range(s):
    try:
        total += 1
        name = str(IA_num_Abby[value])
        type = str(type_Abby[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\AM\{}-AM'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Abby[value]
            ws['I1'].value = case_num_Abby[value]
            ws['I4'].value = due_Date_Abby[value]
            if type_Abby[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Abby[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Abby[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Abby[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Abby[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Andre's folder.")
# Andre
for value in range(t):
    try:
        total += 1
        name = str(IA_num_Andre[value])
        type = str(type_Andre[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\AB\{}-AB'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Andre[value]
            ws['I1'].value = case_num_Andre[value]
            ws['I4'].value = due_Date_Andre[value]
            if type_Andre[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Andre[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Andre[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Andre[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Andre[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Jason's folder.")
# Jason
for value in range(u):
    try:
        total += 1
        name = str(IA_num_Jason[value])
        type = str(type_Jason[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\JH\{}-JH'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Jason[value]
            ws['I1'].value = case_num_Jason[value]
            ws['I4'].value = due_Date_Jason[value]
            if type_Jason[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Jason[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Jason[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Jason[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Jason[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Andrew's folder.")
# Andrew
for value in range(v):
    try:
        total += 1
        name = str(IA_num_Andrew[value])
        type = str(type_Andrew[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\AN\{}-AN'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Andrew[value]
            ws['I1'].value = case_num_Andrew[value]
            ws['I4'].value = due_Date_Andrew[value]
            if type_Andrew[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Andrew[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Andrew[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Andrew[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Andrew[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating JoshG's folder.")
# JoshG
for value in range(w):
    try:
        total += 1
        name = str(IA_num_JoshG[value])
        type = str(type_JoshG[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\JG\{}-JG'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_JoshG[value]
            ws['I1'].value = case_num_JoshG[value]
            ws['I4'].value = due_Date_JoshG[value]
            if type_JoshG[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_JoshG[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_JoshG[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_JoshG[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_JoshG[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

print("Populating Anna's folder.")
# Anna
for value in range(x):
    try:
        total += 1
        name = str(IA_num_Anna[value])
        type = str(type_Anna[value])
        full_name = r'G:\2023\23.22984\CO_Reviews\AR\{}-AR'.format(name)
        if not os.path.exists(full_name):
            os.makedirs(full_name)
        old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
        new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(name))
        if not os.path.exists(new_file):
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C1'].value = name
            ws['I2'].value = type
            ws['I3'].value = installer_Anna[value]
            ws['I1'].value = case_num_Anna[value]
            ws['I4'].value = due_Date_Anna[value]
            if type_Anna[value] == "ESS":
                ws['C4'].value = "Yes"
            else:
                ws['C4'].value = "No"
            wb.save(new_file)

        if len(os.listdir(full_name)) > 1:
            pass
        else:
            # Grab the oneline and site plan files and put them in the folder
            # Put the IA that is being processed here and pull it from the folder name
            driver.find_element("id", 'phSearchInput').send_keys(name)
            driver.find_element("id", 'phSearchButton').click()

            # Click the top result from case number selector
            driver.find_element("css selector", "a[target='_top']").click()
            time.sleep(5)

            # Go to actions
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
            time.sleep(3)

            z=0
            if type_Anna[value] == "ESS":
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Declaration':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 3:
                             break
            else:
                for count in range(2,25):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'One-Line Diagram' or driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Site Plan':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                         time.sleep(3)
                         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[9]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[1]/a[2]").click()
                         time.sleep(5)
                         driver.back()
                         time.sleep(1)
                         z+=1
                         if z == 2:
                             break
            driver.back()
            time.sleep(1)

            # Now put the file in the correct spot
            if type_Anna[value] == "ESS":
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)
                shutil.move(sorted_files[-3], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])
                head_3, tail_3 = os.path.split(sorted_files[-3])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0] or newest_file == tail_3.rsplit(".",1)[0]:
                    pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            else:
                list_of_files = glob.glob(r'C:\Users\joe.nogosek\Downloads\*.pdf') # * means all if need specific format then *.csv
                sorted_files = sorted(list_of_files, key=os.path.getmtime)
                shutil.move(sorted_files[-1], full_name)
                shutil.move(sorted_files[-2], full_name)

                head_1, tail_1 = os.path.split(sorted_files[-1])
                head_2, tail_2 = os.path.split(sorted_files[-2])

                newest_file = driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[13]/div[1]/div/div[2]/table/tbody/tr[2]/th/a").get_attribute("innerText")
                if newest_file == tail_1.rsplit(".",1)[0] or newest_file == tail_2.rsplit(".",1)[0]:
                        pass
                else:
                    txt_file = open("!!! Make sure these are the most recent files!", "w+")
                    txt_file.close()

            if size_Anna[value] < 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                try:
                    # Case number (oneline)
                    ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Name of customer (oneline)
                    ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # AC capacity (oneline)
                    ws["D27"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]").get_attribute("innerText")  + "\nDC Capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #DC Size
                    ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                    #####ws["D67"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                    # Address (site plan)
                    ws["D14"].value = "Premise: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[2]/td[4]/a[1]").get_attribute("innerText") + \
                    "\nAddress: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]").get_attribute("innerText") + \
                    "\nCity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText") + \
                    "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[10]/td[2]").get_attribute("innerText")
                    # Installer (site plan)
                    ws["D72"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[4]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[3]").get_attribute("innerText") + \
                    "\n" + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[3]").get_attribute("innerText")
                except:
                    pass

                try:
                    # ESS Charge source
                    ws["D37"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") + \
                    "\nESS Export: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[25]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    #Existing DG
                    ws["D35"].value =  "Existing DG: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                    # Case number (site plan)
                    #####ws["D169"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[7]/table[1]/tbody[1]/tr[2]/td[2]/a[1]").get_attribute("innerText")
                    # Feeder number
                    ##ws["D69"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    # PV Module and quantity
                    ws["D31"].value = "PV Module: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[8]").get_attribute("innerText")  + "\n# of modules: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[5]/td[2]").get_attribute("innerText")
                    #City to check if in mountain division
                    ws["D79"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[9]/td[2]").get_attribute("innerText")
                    # Inverter name (oneline)
                    ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[8]").get_attribute("innerText") + "  \n# of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                except:
                    pass
                try:
                    # Go into inverter information
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                    time.sleep(2)
                    # Click into inverter
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                    time.sleep(2)
                    # Individual inverter rating (oneline)
                    ws["D29"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText")
                    driver.back()
                    time.sleep(1)
                    driver.back()
                    time.sleep(1)
                except:
                    pass

                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

            elif size_Anna[value] >= 1000:
                wb = load_workbook(new_file, keep_vba=True)
                ws = wb['Completeness Review']
                # Case number (oneline)
                ws["D13"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Name of customer (oneline)
                ws["D14"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # AC capacity (oneline)
                ws["D70"].value = "AC capacity: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + "    # of inverters: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText")
                #DC Size
                ws["I6"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[17]/table[1]/tbody[1]/tr[2]/td[2]").get_attribute("innerText")
                ws["D66"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/td[4]").get_attribute("innerText")
                # Address (site plan)
                ws["D67"].value = "City: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[7]/td[2]/div[1]").get_attribute("innerText") + \
                "\nZip code: " + driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[8]/td[2]/div[1]").get_attribute("innerText")
                # Case number (site plan)
                ws["D71"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[19]/table[1]/tbody[1]/tr[2]/td[2]/div[1]/a[1]").get_attribute("innerText")
                # Feeder number
                feed_num = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[15]/table[1]/tbody[1]/tr[5]/td[2]/div[1]").get_attribute("innerText")
                if len(feed_num) > 5:
                    ws["D126"].value = feed_num
                else:
                    ws["D126"].value = "Feeder number missing!"
                # Coordinates
                ws["D81"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[6]/td[2]/div[1]").get_attribute("innerText")
                # Go into inverter information
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[12]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[2]/th[1]/a[1]").click()
                time.sleep(2)
                # Inverter name (oneline)
                ws["D73"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").get_attribute("innerText")
                # Click into inverter
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[3]/table[1]/tbody[1]/tr[3]/td[4]/div[1]/a[1]").click()
                time.sleep(2)
                # Individual inverter rating (oneline)
                ws["D30"].value = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[5]/table[1]/tbody[1]/tr[1]/td[2]/div[1]").get_attribute("innerText") + " W"
                driver.back()
                time.sleep(1)
                driver.back()
                time.sleep(1)
                wb.save(new_file)

                xl=win32.Dispatch("Excel.Application")
                book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
                xl.Run("existingMeter")
                xl.Run("hideRows")
                book.Close(SaveChanges=True)
                del xl

        counter += 1
    except:
        print(f"Error with {name}")
        continue

driver.close()
print("Folders are now all populated.")
print(f"{counter} folders successfully created out of a total of {total}!")

input("Press enter to exit.")
