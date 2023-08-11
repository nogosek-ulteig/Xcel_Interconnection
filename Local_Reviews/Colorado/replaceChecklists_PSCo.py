# Made by Joe Nogosek
# 3/30/2022

import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import shutil
import warnings
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

warnings.filterwarnings("ignore")

# initials = ["JN", "RK", "NC", "EU", "JCN", "JB", "AA", "AB", "AN", "JH", "AM", "JG", "ES"]
initials = ["JB", "AA", "AN", "AM", "AR"]
pathToReviews = r'G:\2023\23.22984\CO_Reviews'
type = ""
case_num = ""
IA = ""
under = ""
over = ""
reviewer = ""

for name in initials:
    print(f"Working on {name}.")
    pathToFolder = os.path.join(pathToReviews, name)
    curFolders = os.listdir(pathToFolder)
    for folder in curFolders:
        if folder[:2] == "IA":
            IA_num = folder.split('-')[0]
            print(IA_num)
            pathToIA = os.path.join(pathToFolder, folder)
            full_name = r'G:\2023\23.22984\CO_Reviews\{0}\{1}-{2}'.format(name, IA_num, name)
            try:
                files = os.listdir(pathToIA)
                for file in files:
                    if file.endswith(".xlsm"):
                        under = str(file[25:30])
                        over = str(file[18:22])
                        pathToChecklist = os.path.join(pathToIA, file)
                        wb = load_workbook(pathToChecklist, keep_vba=True)
                        ws = wb['Completeness Review']
                        type = str(ws['C4'].value)
                        installer = str(ws['I3'].value)
                        case_num = str(ws['I1'].value)
                        IA = str(ws['C1'].value)
                        name_oneline_SF = ws['D13'].value
                        address_SF = ws['D14'].value
                        aggregate_SF = ws['D27'].value
                        inverter_rating_SF = ws['D29'].value
                        inverter_type_SF = ws['D30'].value
                        modules_SF = ws['D31'].value
                        existingDG_SF = ws['D35'].value
                        ESSexport_SF = ws['D37'].value
                        division_SF = ws['D89'].value
                        installer_82 = ws['D82'].value
                        os.remove(pathToChecklist)
                    else:
                        continue
            except Exception as e:
                print(f"Error {e} with {IA_num}")
                continue
            shutil.copy(r'G:\2023\23.22984\CO_Reviews\PSCo_CR_Checklist-.xlsm',full_name)
            old_file = os.path.join(full_name, 'PSCo_CR_Checklist-.xlsm')
            new_file = os.path.join(full_name, 'PSCo_CR_Checklist-{}.xlsm'.format(IA_num))
            os.rename(old_file, new_file)
            wb = load_workbook(new_file, keep_vba=True)
            ws = wb['Completeness Review']
            ws['C4'].value = type
            ws['I3'].value = installer
            ws['I1'].value = case_num
            ws['C1'].value = IA
            ws['D13'].value = name_oneline_SF
            ws['D14'].value = address_SF
            ws['D27'].value = aggregate_SF
            ws['D29'].value = inverter_rating_SF
            ws['D30'].value = inverter_type_SF
            ws['D31'].value = modules_SF
            ws['D35'].value = existingDG_SF
            ws['D37'].value = ESSexport_SF
            ws['D89'].value = division_SF
            ws['D82'].value = installer_82
            wb.save(new_file)
            xl=win32.Dispatch("Excel.Application")
            book = xl.Workbooks.Open(os.path.abspath(new_file), ReadOnly=0)
            xl.Run("existingMeter")
            xl.Run("hideRows")
            book.Close(SaveChanges=True)
            del xl
            type = ""
            case_num = ""
            type_SR = ""
            under = ""
            over = ""

input("Press enter to exit.")
