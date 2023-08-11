# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 3/9/2023

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import shutil
import sys
import datetime
import warnings
from selenium.webdriver.chrome.options import Options
import pathlib
import pyautogui
import glob
from win32com import client
import zipfile
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

username = credentials.username
password = credentials.password
pathToTracker = credentials.path_to_MN_preapps_tracker
pathToDownloadsFolder = credentials.path_to_downloads
path_to_driver = credentials.path_to_driver
name = credentials.name

pathToReadytoProcess = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Preapps\Ready_to_Process"
pathToCompleted = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Preapps\Completed"

try:
    to_unzip = glob.glob(os.path.join(pathToDownloadsFolder,"202*.zip"))
    latest_file = max(to_unzip, key=os.path.getctime)

    with zipfile.ZipFile(to_unzip[0], 'r') as zip_ref:
        zip_ref.extractall(r"C:\Users\joe.nogosek\Documents\Projects\NSP_Preapps\Ready_to_Process")
        print("Folders successfully moved over!")
        list_of_files = glob.glob(os.path.join(pathToDownloadsFolder,"*")) # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        zip_ref.close()
        os.remove(os.path.join(pathToDownloads,latest_file))
except:
    print("No files to unzip.")
    pass

wbTracker = load_workbook(pathToTracker)
wsTracker = wbTracker['2023']

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings('ignore')

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

# Process in Salesforce
driver = webdriver.Chrome(executable_path=path_to_driver, options=options)
driver.maximize_window()
driver.implicitly_wait(5)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')
time.sleep(1)

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

# Now get all the necessary information from the 2022 tab
listOfIA = []
listofCases = []

numIA = 0
for row in range(2,wsTracker.max_row+1):
    if str(wsTracker[row][9].value) == 'x' and str(wsTracker[row][5].value) == 'None':
        listOfIA.append(str(wsTracker[row][2].value))
        listofCases.append(str(wsTracker[row][0].value))
        numIA += 1

i = 0
for IA in listOfIA:
    print("Working on " + IA)
    wb = load_workbook(os.path.join(pathToReadytoProcess,IA,"Report.xlsm"), data_only = True)
    ws = wb["Pre-App Data Report"]

    excel = client.Dispatch("Excel.Application")
    sheets = excel.Workbooks.Open(os.path.join(pathToReadytoProcess,IA,"Report.xlsm"))
    work_sheets = sheets.Worksheets["Pre-App Data Report"]
    work_sheets.ExportAsFixedFormat(0,os.path.join(pathToReadytoProcess,IA,"Report.pdf"))
    sheets.Close(True)
    excel.Quit()

    subName = str(ws["B50"].value)
    subInitials = str(ws["B59"].value)[:3]
    feeder = ws["B59"].value
    gardenName = ws["B13"].value

    wb.close()
    os.rename(os.path.join(pathToReadytoProcess,IA,"Report.xlsm"),os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Report.xlsm"))
    newFilePath = os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Report.xlsm")
    os.rename(os.path.join(pathToReadytoProcess,IA,"Report.pdf"),os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Report.pdf"))
    newFilePDF = os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Report.pdf")
    os.rename(os.path.join(pathToReadytoProcess,IA,"Request.pdf"),os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Request.pdf"))
    os.rename(os.path.join(pathToReadytoProcess,IA,"Territory Map.png"),os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Territory Map.png"))
    os.rename(os.path.join(pathToReadytoProcess,IA,"Supplemental.docx"),os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Supplemental.docx"))
    pathToSite = os.path.join(pathToReadytoProcess,IA,str(glob.glob(os.path.join(pathToReadytoProcess,IA,'Site Map*'))[0]))
    fileExt = pathToSite[-3:]
    os.rename(pathToSite,os.path.join(pathToReadytoProcess,IA,feeder+" "+listofCases[i]+" "+gardenName+" Site Map."+fileExt))

    wb = load_workbook(os.path.join(pathToReadytoProcess,IA,str(feeder)+" "+str(listofCases[i])+" "+str(gardenName)+" Preapp Report.xlsm"), data_only = True)
    ws = wb["Pre-App Data Report"]

    # Put the IA that is being processed here and pull it from the folder name
    driver.find_element("id", 'phSearchInput').send_keys(IA)
    driver.find_element("id", 'phSearchButton').click()

    # Click the top result from case number selector
    driver.find_element("css selector", "a[target='_top']").click()
    time.sleep(8)

    # Go to actions and open initial review results action
    for count in range(2,5):
        if driver.find_element("xpath", f"/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[10]/div[1]/div/div[2]/table/tbody/tr[" + str(count)+ "]/td[2]").get_attribute("innerText") == 'Pre-Application Report Generate':
             driver.find_element("xpath", f"/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[10]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/th/a").click()
             time.sleep(2)
             driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[4]/div[2]/div[2]/table/tbody/tr[10]/td[2]/div/a").click()
             break
    time.sleep(3)

    # Get window handles
    p = driver.current_window_handle
    chwd = driver.window_handles
    for w in chwd:
        if(w!=p):
            driver.switch_to.window(w)
    time.sleep(3)

    # Click 'Upload Files'
    driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-aura-field/div/div/lightning-file-upload/lightning-input/div/div/lightning-primitive-file-droppable-zone/slot/label/span[1]").click()
    time.sleep(5)
    try:
        pyautogui.click(520, 468)
        pyautogui.hotkey('ctrl','a')
        pyautogui.press('delete')
        pyautogui.typewrite(newFilePDF)
        pyautogui.click(790,500)
        time.sleep(0.5)
    except:
        time.sleep(3)
        pyautogui.click(520, 468)
        pyautogui.hotkey('ctrl','a')
        pyautogui.press('delete')
        pyautogui.typewrite(newFilePDF)
        pyautogui.click(790,500)
        time.sleep(0.5)
    time.sleep(10)
    driver.find_element("xpath", "/html/body/div[1]/article/div/div/div[2]/div/div[3]/div/span[2]/button").click()
    time.sleep(3)
    button = driver.find_element("name", "NEXT")
    driver.execute_script("arguments[0].click();", button)
    time.sleep(3)

    # Go back to our main IA page
    driver.close()
    driver.switch_to.window(p)
    driver.find_element("id", 'phSearchInput').send_keys(IA)
    driver.find_element("id", 'phSearchButton').click()
    driver.find_element("css selector", "a[target='_top']").click()
    time.sleep(5)

    driver.switch_to.frame("0664O000000hsL8")
    time.sleep(1)
    button = driver.find_element("xpath", "/html/body/div[1]/div[1]/div/div[2]/button")
    driver.execute_script("arguments[0].click();", button)

    driver.switch_to.default_content()

    time.sleep(25)

    print("Finished processing " + IA + "!")

    # Enter feeder and substation name into tracker
    for row in range(2,wsTracker.max_row+1):
        if str(wsTracker[row][10].value) == 'None' and str(wsTracker[row][11].value == 'None') :
            empty_row = row
            break

    wsTracker["K" + str(empty_row)].value = feeder
    wsTracker["L" + str(empty_row)].value = subName
    wsTracker["F" + str(empty_row)].value = datetime.date.strftime(datetime.date.today(), "%m/%d/%Y")

    wbTracker.save(pathToTracker)

    # Now let's move the screen and its pdf to the right spot
    if not os.path.exists(os.path.join(pathToCompleted,subInitials + " - " + subName,feeder)):
        os.makedirs(os.path.join(pathToCompleted,subInitials + " - " + subName,feeder))
    shutil.move(os.path.join(pathToReadytoProcess,IA),os.path.join(pathToCompleted,subInitials + " - " + subName,feeder))
    os.rename(os.path.join(pathToCompleted,subInitials + " - " + subName,feeder,IA),os.path.join(pathToCompleted,subInitials + " - " + subName,feeder,listofCases[i] + " " + gardenName))

    print("Moved successfully!")

    time.sleep(3)
    driver.refresh()
    time.sleep(2)

    successFlag = 0
    try:
        if driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[2]/td[4]/img").get_attribute("alt") == "Not Checked":
            successFlag = 1
            pass

    except Exception as e:
        print("Error :(")
        print(e)
        pause = input("Press a key and then enter when ready to process the next one: ")

    i+=1

wbTracker.close()
print("Good to open tracker.\n")

print("This script is finished. Make sure to help all the open Chrome windows finish.")

input("Press enter to exit.")
