# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 1/12/2023

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import shutil
import sys
import warnings
from selenium.webdriver.chrome.options import Options
import pathlib
import pyautogui
from win32com import client
import os
import sys

user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

username = credentials.username
password = credentials.password
pathToTracker = credentials.path_to_MN_initials_tracker
pathToDownloadsFolder = credentials.path_to_downloads
path_to_driver = credentials.path_to_driver
name = credentials.name

pathToJN = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\JN"
pathToCompleted = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\Completed"
pathToReadytoProcess = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\Ready_to_Process"

wbTracker = load_workbook(pathToTracker)
wsTracker = wbTracker['2023']

# Now get all the necessary information from the 2022 tab
listOfCases = []

numCases = 0
for row in range(2,wsTracker.max_row+1):
    if str(wsTracker[row][7].value) == 'Joe Nogo' and str(wsTracker[row][8].value) != 'None' and str(wsTracker[row][9].value) == 'None':
        listOfCases.append("Initial Review_Case#0" + str(wsTracker[row][5].value))
        numCases += 1

wbTracker.close()
print("Good to open tracker.\n")

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings('ignore')

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

# Process in Salesforce
driver = webdriver.Chrome(executable_path=path_to_driver, options=options)
driver.implicitly_wait(5)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')
time.sleep(1)

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

total = 0
outlook = client.Dispatch("Outlook.Application")
for caseNum in listOfCases:
    pyautogui.move(0,1)
    for i in range(0,3):
        pyautogui.press("shift")
    just_case = caseNum[-8:]
    print("Working on " + just_case)
    try:
        wb = load_workbook(os.path.join(pathToReadytoProcess,caseNum)+".xlsm", data_only = True)
        ws = wb["NSP Cstmr Report Initial Review"]

        excel = client.Dispatch("Excel.Application")
        sheets = excel.Workbooks.Open(os.path.join(pathToReadytoProcess,caseNum)+".xlsm")
        work_sheets = sheets.Worksheets["NSP Cstmr Report Initial Review"]
        work_sheets.ExportAsFixedFormat(0,os.path.join(pathToReadytoProcess,caseNum)+".pdf")
        excel.Quit()

        subInitials = str(ws["D11"].value)[:3]
        feeder = str(ws["D11"].value)
        address = ws["C8"].value
        address = address.replace("  ", " ")
        address = str(address.split(",", 1)[0])
        #size = float(str(ws["C29"].value))
        substation = str(ws["I11"].value)
        summary = str(ws["A18"].value)
        ground_ref = str(ws["A20"].value)
        desc = ws["A18"].value
        size = ws["C9"].value

        key = 0
        ground_key = 0
        if summary == "This project does not qualify for either the Simplified or Fast Track Process and will need to proceed to the Study Process.":
            key = 1
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the substation transformer, the details of which are provided below. The proposed project cannot be approved without incurring significant cost, which for this project would require at least a substation transformer upgrade to increase capacity. An enhanced version of the System impact Study will be required to determine whether, how, and at what indicative estimated costs the project could proceed. With the significant costs to allow the interconnection of the proposed project, it may not be financially viable for the interconnection customer to proceed. The interconnection customer has the option to attend a customer options meeting. This meeting must be accepted or declined prior to the initiation of the Supplemental Review or Study Process.":
            key = 2
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the distribution feeder, the details of which are provided below. The proposed project cannot be approved without incurring significant cost, which for this project would require at least a new distribution feeder to increase capacity. An enhanced version of the System impact Study will be required to determine whether, how, and at what indicative estimated costs the project could proceed. With the significant costs to allow the interconnection of the proposed project, it may not be financially viable for the interconnection customer to proceed. The interconnection customer has the option to attend a customer options meeting. This meeting must be accepted or declined prior to the initiation of the Supplemental Review or Study Process.":
            key = 3
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the substation transformer, the details of which are provided below. Further review or study will be required to determine if the project can be interconnected consistent with safety, reliability, and power quality standards. This review or study may identify required upgrades of significant costs to increase the substation transformer capacity. The interconnection customer has the option to attend a customer options meeting. This meeting must be accepted or declined prior to the initiation of the Supplemental Review or Study Process.":
            key = 4
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the distribution feeder, the details of which are provided below. Further review or study will be required to determine if the project can be interconnected consistent with safety, reliability, and power quality standards. This review or study may identify required upgrades of significant costs to increase the substation transformer capacity. The interconnection customer has the option to attend a customer options meeting. This meeting must be accepted or declined prior to the initiation of the Supplemental Review or Study Process.":
            key = 5
        elif summary == "This project has passed the Initial Review Screens and does not require construction of facilities by Xcel Energy. An executable Uniform Statewide Contract or MN DIA will be provided. ":
            key = 6
        elif summary == "This project has passed the Initial Review Screens. It will require construction of facilities by Xcel Energy. Details can be found below. A facilities study will be required to determine a construction cost estimate. A facilities study agreement will be provided. ":
            key = 7
        elif summary == "This project has failed the Initial Review Screens, but it was determined that the project can still interconnect safely and reliably and does not require construction of facilities by Xcel Energy. An executable Uniform Statewide Contract or MN DIA will be provided. ":
            key = 8
        elif summary == "This project has failed the Initial Review Screens, the details of which are provided below. A supplemental review will be required to determine if the DER may be interconnected consistent with safety, reliability, and power quality standards. The interconnection customer has the option to attend a customer options meeting. This meeting must be accepted or declined prior to the initiation of the Supplemental Review. ":
            key = 9
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the substation transformer, the details are provided below. The proposed project cannot be approved without incurring significant cost, which would require at least a substation transformer upgrade to increase capacity. An enhanced version of the System impact Study will be required to determine whether, how, and at what indicative estimated costs the project could proceed. With the significant costs, it may not be financially viable for the interconnection customer to proceed. The Interconnection Customer will be provided with the opportunity to attend a customer options meeting.":
            key = 10
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the distribution feeder, the details are provided below. The proposed project cannot be approved without incurring significant cost, which would require at least a new distribution feeder to increase capacity. An enhanced version of the System impact Study will be required to determine whether, how, and at what indicative estimated costs the project could proceed. With the significant costs, it may not be financially viable for the interconnection customer to proceed. The Interconnection Customer will be provided with the opportunity to attend a customer options meeting.":
            key = 11
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the substation transformer, the details of which are provided below. The project will need to enter the Study Process to determine if the project can be interconnected consistent with safety, reliability, and power quality standards. This study may identify required upgrades of significant costs to increase the substation transformer capacity. The Interconnection Customer will be provided with the opportunity to attend a customer options meeting.":
            key = 12
        elif summary == "This project has exceeded the Technical Planning Standard (TPS) on the distribution feeder, the details of which are provided below. The project will need to enter the Study Process to determine if the project can be interconnected consistent with safety, reliability, and power quality standards. This study may identify required upgrades of significant costs to increase the feeder capacity. The Interconnection Customer will be provided with the opportunity to attend a customer options meeting.":
            key = 13
        elif summary == "This project has failed the Initial Review Screens, but has passed the Supplemental Review. Details can be found below. It does not require construction of facilities by Xcel Energy. An executable Uniform Statewide Contract or MN DIA will be provided. ":
            key = 14
        elif summary == "This project has failed both the Initial Review and the Supplemental Review Screens, the details of which can be found below. However, with some facilities upgrades, the interconnection will be able to proceed. The details of these upgrades can be found on page 3 and 4 of this report. A scoping meeting option is available to the customer, as well.  A facilities study will be required to determine a construction cost estimate. A facilities study agreement will be provided.":
            key = 15
        elif summary == "This project has failed both the Initial Review and the Supplemental Review Screens, the details of which can be found below. To determine if the project can be interconnected consistent with safety, reliability, and power quality standards, the project will need to enter the Study Process. The Interconnection Customer will be provided with the opportunity to attend a customer options meeting.":
            key = 16
        elif summary == "This project has failed both the Initial Review and the Supplemental Review Screens, the details of which can be found below. However, with some facilities upgrades, the interconnection will be able to proceed. The details of these upgrades can be found on page 3 and 4 of this report. A scoping meeting option is available to the customer, as well.  A facilities study will be required to determine a construction cost estimate. A facilities study agreement will be provided.":
            key = 17
        else:
            print("!!!!!!!!!!!!!!!! Summary is blank !!!!!!!!!!!!!!!!!!!!!!!!!!")
            continue
        if ground_ref == "This project is not required to install ground referencing equipment as it is less than 100 kW.":
            ground_key = 1
        elif ground_ref == "This project is required to install ground referencing equipment. The submitted specifications for the ground bank do not meet the required specifications. The specific requirements are detailed at the end of this report.":
            ground_key = 2
        elif ground_ref == "Based on the project size and system configuration, the ground referencing equipment specifications appear to be adequate for installation with this interconnection. Should the size or configuration of this project change at any point in time, this determination will no longer be valid. It is the customer's responsibility to ensure that the ground referencing equipment specifications are reviewed and in compliance with Xcel Energy's Ground Reference Requirements. ":
            ground_key = 3
        else:
            print("!!!!!!!!!!!!!!!! Ground referencing is blank !!!!!!!!!!!!!!!!!!!!!!!!!!")
            continue

        # Put the IA that is being processed here and pull it from the folder name
        driver.find_element("id", 'phSearchInput').send_keys(just_case)
        driver.find_element("id", 'phSearchButton').click()

        # Click the top result from case number selector
        driver.find_element("css selector", "a[target='_top']").click()
        time.sleep(8)

        # Go to actions and open initial review results action
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
        time.sleep(1)
        try:
            for count in range(2,30):
                #if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Simple Track)' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Fast Track)':
                if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Simple Track)':
                     driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(2)
                     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                     break
        except:
            for count in range(2,30):
                #if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Simple Track)' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Fast Track)':
                if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Fast Track)':
                     driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                     time.sleep(2)
                     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                     break

        time.sleep(3)

        # Get window handles
        p = driver.current_window_handle
        chwd = driver.window_handles
        for w in chwd:
            if(w!=p):
                driver.switch_to.window(w)
        time.sleep(3)

        # Click 'Upload Files'
        driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[3]/flowruntime-aura-field/div/div/lightning-file-upload/lightning-input/div/div/lightning-primitive-file-droppable-zone/slot/label/span[1]").click()
        time.sleep(5)
        try:
            pyautogui.click(520, 468)
            pyautogui.typewrite(os.path.join(pathToReadytoProcess,caseNum+".pdf"))
            pyautogui.click(790,500)
            time.sleep(0.5)
        except:
            time.sleep(3)
            pyautogui.click(520, 468)
            pyautogui.typewrite(os.path.join(pathToReadytoProcess,caseNum+".pdf"))
            pyautogui.click(790,500)
            time.sleep(0.5)

        time.sleep(10)
        button = driver.find_element("xpath", "/html/body/div[1]/article/div/div/div[2]/div/div[3]/div/span[2]/button/span")
        driver.execute_script("arguments[0].click();", button)
        time.sleep(3)
        button = driver.find_element("xpath", "/html/body/div[1]/article/flowruntime-flow/flowruntime-lwc-body/div/flowruntime-list-container/div/flowruntime-base-section/div/flowruntime-screen-field[4]/flowruntime-aura-field/div/div/div/button")
        driver.execute_script("arguments[0].click();", button)
        time.sleep(3)

        # Go back to our main IA page
        driver.close()
        driver.switch_to.window(p)
        driver.find_element("id", 'phSearchInput').send_keys(just_case)
        driver.find_element("id", 'phSearchButton').click()
        driver.find_element("css selector", "a[target='_top']").click()
        time.sleep(5)

        # Go into Edit
        driver.find_element("name", "edit").click()
        time.sleep(1)
        # Put in screen result, detailed result, summary, and allowed actions (if applicable)
        if key == 1:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 2:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 3:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 4:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 5:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 6 and ground_key != 2:
            # Pass
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            # Safe and Reliable Interconnection
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Not needed
            # driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif key == 6 and ground_key == 2:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Ground referencing fails
            # Paste in description and ground referencing notes
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            # Now we have to send an email to PMO
            mail = outlook.CreateItem(0)
            mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
            mail.CC = "solarprogrammn@xcelenergy.com"
            mail.BCC = "joe.nogosek@ulteig.com"
            mail.Subject = "Material Modification Necessary for " + just_case
            mail.Body = "Hi,\n\nThe initial review for " + just_case + " passed all initial review screens. However, they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
            mail.send
            print("Email sent to PMO for " + just_case + "!")
            # Supplemental Review, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 7:
            # Pass
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            # Safe and Reliable Interconnection
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " passed all initial review screens. However, they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Facility Study, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 8:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Safe and Reliable Interconnection
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Not needed
            # driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 9:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, Customer Options Metting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 10:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # Supplemental Review, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 11:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 12:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 13:
            # Fail
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Further Study Needed
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
            # Check if ground referencing notes are required to be pasted in
            if ground_key == 3:
                # Ground referencing passes
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
            elif ground_key == 2:
                # Ground referencing fails
                # Paste in description and ground referencing notes
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys("\n\n")
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(ground_ref)
                # Now we have to send an email to PMO
                mail = outlook.CreateItem(0)
                mail.To = "nicholas.j.coquyt@xcelenergy.com; violeta.vidakovic@xcelenergy.com"
                mail.CC = "solarprogrammn@xcelenergy.com"
                mail.BCC = "joe.nogosek@ulteig.com"
                mail.Subject = "Material Modification Necessary for " + just_case
                mail.Body = "Hi,\n\nThe initial review for " + just_case + " has failed one or more initial review screens and they have insufficient ground referencing. Can you please have " + just_case + " moved into material modification status so the developer can update the ground referencing?\n\nThanks!\nJoe Nogosek"
                mail.send
                print("Email sent to PMO for " + just_case + "!")
            elif ground_key == 1:
                # Ground referencing notes not necessary
                # Paste in description
                driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
            # System Impact Study, Customer Options Meeting, Withdraw
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        if key == 14:
            print("!!!!!! Summary not applicable to initial review !!!!!!!!!")
            continue
        if key == 15:
            print("!!!!!! Summary not applicable to initial review !!!!!!!!!")
            continue
        if key == 16:
            print("!!!!!! Summary not applicable to initial review !!!!!!!!!")
            continue
        if key == 17:
            print("!!!!!! Summary not applicable to initial review !!!!!!!!!")
            continue

        # Put in screen result
        # if ws["I5"].value == "Pass":
        #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        # elif ws["I5"].value == "Fail":
        #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        # else:
        #     print("Issue with 'Passes Screen'. Please inspect manually.")
        #     quit()

        # # Enter detailed result
        # if ws["I5"].value == "Pass":
        #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        # # elif ws["I145"].value == "Yes":
        # #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        # else:
        #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        #
        # # Paste in description
        # driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)
        #
        # if ws["I5"].value == "Fail":
        #     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)

        # Save
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
        time.sleep(3)

        # Create a new screen result
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[16]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
        time.sleep(2)
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
        time.sleep(3)
        # In Xcel territory
        if ws["I52"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I52"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I52"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with DER is in Xcel territory!")
            quit()
        # Reverse power flow
        if ws["I59"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I59"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I59"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with reverse power flow!")
            quit()
        # Network interconnection size
        if ws["I68"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I68"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I68"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with network interconnection size!")
            quit()
        # 10% of fault contribution
        if ws["I77"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I77"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I77"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with 10% of fault contribution!")
            quit()
        # 87.5% of interrupt rating
        if ws["I87"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I87"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I87"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with 87.5% of interrupt rating!")
            quit()
        # Service type compatible
        if ws["I93"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I93"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I93"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with service type compatible!")
            quit()
        # DER size on shared secondary
        if ws["I100"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I100"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I100"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with DER size on shared secondary!")
            quit()
        # 120V DER on 240V service
        if ws["I108"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I108"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I108"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with 120V DER on 240V service!")
            quit()
        # 1PH DER on 3PH Service TR
        if ws["I117"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I117"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I117"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with 1PH DER on 3PH Service TR!")
            quit()
        # DER size behind volt reg
        if ws["I123"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I123"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I123"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with DER size behind volt reg!")
            quit()
        # Service TR loading
        if ws["I130"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I130"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I130"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with service TR loading!")
            quit()
        # Downstream of ATO
        if ws["I144"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I144"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I144"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with downstream of ATO!")
            quit()
        # VSR
        if ws["I143"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I143"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I143"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with VSR!")
            quit()
        # Other facilities required
        if ws["I145"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I145"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I145"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with other facilities required!")
            quit()
        # Grounding requirement 1
        if ws["I161"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I161"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I161"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with grounding requirement 1!")
            quit()
        # Grounding requirement 2
        if ws["I165"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I165"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I165"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with grounding requirement 2!")
            quit()
        # Grounding requirement 3
        if ws["I169"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I169"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I169"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with grounding requirement 3!")
            quit()
        # Grounding requirement 4
        if ws["I173"].value == 'Yes':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        elif ws["I173"].value == 'No':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        elif ws["I173"].value == 'N/A':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        else:
            print("Error with grounding requirement 4!")
            quit()
        # Click save
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

        # Go back to main IA page
        driver.find_element("id", 'phSearchInput').send_keys(just_case)
        driver.find_element("id", 'phSearchButton').click()

        # Click the top result from case number selector
        driver.find_element("css selector", "a[target='_top']").click()
        time.sleep(3)

        driver.maximize_window()

        time.sleep(1)
        driver.switch_to.frame("0664O000000hsL8")
        time.sleep(1)
        driver.find_element("xpath", "/html/body/div[1]/div[1]/div/div[2]/button").click()
        time.sleep(2)
        if key == 6:
            if ground_key == 1 or ground_key == 3:
                for count in range(1,3):
                    if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Initial Engineering Screens Complete. Provide Interconnection Agreement.':
                        driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
                        driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
            elif ground_key == 2:
                for count in range(1,3):
                    if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Further Study or Construction upgrade is required. The applicant will need to decide on next steps.':
                        driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
                        driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
        elif key == 8:
            for count in range(1,3):
                if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Initial Engineering Screens Complete. Provide Interconnection Agreement.':
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
        elif key == 1 or key == 2 or key == 3 or key == 4 or key == 5 or key == 7 or key == 9 or key == 10 or key == 11 or key == 12 or key == 13 or key == 14 or key == 15 or key == 16 or key == 17:
            for count in range(1,3):
                if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Further Study or Construction upgrade is required. The applicant will need to decide on next steps.':
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
        else:
            print("Issue submitting application!")
            quit()

        driver.switch_to.default_content()

        wb.close()

        time.sleep(15)

        print("Finished processing " + just_case + "!")

        # Now let's move the screen and its pdf to the right spot
        if not os.path.exists(os.path.join(pathToCompleted,subInitials+" - "+substation,feeder)):
            os.makedirs(os.path.join(pathToCompleted,subInitials+" - "+substation,feeder))
        os.makedirs(os.path.join(pathToCompleted,subInitials+" - "+substation,feeder,"Case"+just_case+" - "+address))
        shutil.move(os.path.join(pathToReadytoProcess,caseNum+".xlsm"),os.path.join(pathToCompleted,subInitials+" - "+substation,feeder,"Case"+just_case+" - "+address,caseNum+".xlsm"))
        shutil.move(os.path.join(pathToReadytoProcess,caseNum+".pdf"),os.path.join(pathToCompleted,subInitials+" - "+substation,feeder,"Case"+just_case+" - "+address,caseNum+".pdf"))

        print("Moved successfully!")

        time.sleep(3)
        driver.refresh()
        time.sleep(2)

        successFlag = 0
        try:
            try:
                driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/div/a[2]").click()
                time.sleep(1)
                for count in range(2,7):
                    if driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Sign the Interconnection Agreement' or driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Decision' or driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Generate IA and Construction Estimate':
                        if driver.find_element("xpath", "/html/body/div/div[3]/table/tbody/tr/td[2]/div[2]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[4]/img").get_attribute("alt") == 'Checked':
                            successFlag = 1
                            break

                # Go back to the main page so we are ready for the next task
                driver.back()

                if successFlag == 1:
                    pass
                else:
                    print("Error")
                    continue

            except:
                for count in range(2,6):
                    if driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Sign the Interconnection Agreement' or driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Decision' or driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Generate IA and Construction Estimate':
                        if driver.find_element("xpath", "/html/body/div[1]/div[3]/table/tbody/tr/td[2]/div[11]/div[1]/div/div[2]/table/tbody/tr[" + str(count) + "]/td[4]/img").get_attribute("alt") == 'Checked':
                            successFlag = 1
                            break

                if successFlag == 1:
                    pass
                else:
                    print("Error")
                    continue

        except Exception as e:
            print(e)
            continue

        total += 1

    except Exception as e:
        print(e)
        continue

print(f"Total processed: {total}!")
print("This script is finished. Make sure to help all the open Chrome windows finish.")

input("Press enter to exit.")
