# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 1/12/2023

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import shutil
import sys
import warnings
from selenium.webdriver.chrome.options import Options
import pathlib
import pyautogui

# Enter your information here
username = 'joseph.h.nogosek@xcelenergy.com'
password = 'airdoc5Ee'
name = 'Joseph Nogosek'

pathToJN = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\JN"
pathToCompleted = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\Completed"

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings('ignore')

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

caseNum = sys.argv[1]
caseNum = "0" + caseNum
pathToScreen = sys.argv[2]
subInitials = sys.argv[3]
feeder = sys.argv[4]
subName = sys.argv[5]

print(caseNum)

wb = load_workbook(pathToScreen, data_only = True)
ws = wb["NSP Cstmr Report Initial Review"]

print("Status: " + ws["I5"].value)

# Process in Salesforce
driver = webdriver.Chrome(executable_path=r"C:\Users\joe.nogosek\Documents\Python\chromedriver.exe", options=options)
driver.implicitly_wait(5)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')
time.sleep(1)

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

driver.minimize_window()
# Put the IA that is being processed here and pull it from the folder name
driver.find_element("id", 'phSearchInput').send_keys(caseNum)
driver.find_element("id", 'phSearchButton').click()

# Click the top result from case number selector
driver.find_element("css selector", "a[target='_top']").click()
time.sleep(8)

# Go to actions and open initial review results action
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
time.sleep(1)
for count in range(2,25):
    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Initial Review Results (Simple Track)':
         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
         time.sleep(2)
         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
         break
time.sleep(3)

# Get window handles
p = driver.current_window_handle
chwd = driver.window_handles
for w in chwd:
    if(w!=p):
        driver.switch_to.window(w)
time.sleep(3)

# Click 'Upload Files'
driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/lightning-file-upload[1]/lightning-input[1]/div[1]/div[1]/lightning-primitive-file-droppable-zone[1]/slot[1]/label[1]/span[1]").click()
time.sleep(1)
pyautogui.typewrite(os.path.join(pathToJN,caseNum+".pdf"))
pyautogui.click(790,500)
time.sleep(0.5)
driver.minimize_window()
time.sleep(10)
driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[3]/div[1]/span[2]/button[1]/span[1]").click()
time.sleep(3)
driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/div[1]/button[1]").click()
time.sleep(3)

# Go back to our main IA page
driver.close()
driver.switch_to.window(p)
driver.minimize_window()
driver.find_element("id", 'phSearchInput').send_keys(caseNum)
driver.find_element("id", 'phSearchButton').click()
driver.find_element("css selector", "a[target='_top']").click()
time.sleep(5)

# Go into Edit
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[3]").click()
time.sleep(1)
# Put in screen result
if ws["I5"].value == "Pass":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I5"].value == "Fail":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Issue with 'Passes Screen'. Please inspect manually.")
    quit()

# Enter detailed result
if ws["I5"].value == "Pass":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
# elif ws["I145"].value == "Yes":
#     driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)

# Paste in description
desc = ws["A18"].value
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[3]/td[2]/textarea[1]").send_keys(desc)

if ws["I5"].value == "Fail":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[33]/table[1]/tbody[1]/tr[1]/td[4]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)

# Save
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
time.sleep(3)

# Create a new screen result
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[16]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
time.sleep(2)
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
# In Xcel territory
if ws["I52"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I52"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I52"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with DER is in Xcel territory!")
    quit()
# Reverse power flow
if ws["I59"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I59"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I59"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[3]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with reverse power flow!")
    quit()
# Network interconnection size
if ws["I68"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I68"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I68"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[4]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with network interconnection size!")
    quit()
# 10% of fault contribution
if ws["I77"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I77"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I77"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[5]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with 10% of fault contribution!")
    quit()
# 87.5% of interrupt rating
if ws["I87"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I87"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I87"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[6]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with 87.5% of interrupt rating!")
    quit()
# Service type compatible
if ws["I93"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I93"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I93"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[7]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with service type compatible!")
    quit()
# DER size on shared secondary
if ws["I100"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I100"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I100"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[8]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with DER size on shared secondary!")
    quit()
# 120V DER on 240V service
if ws["I108"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I108"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I108"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[9]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with 120V DER on 240V service!")
    quit()
# 1PH DER on 3PH Service TR
if ws["I117"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I117"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I117"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with 1PH DER on 3PH Service TR!")
    quit()
# DER size behind volt reg
if ws["I123"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I123"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I123"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[11]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with DER size behind volt reg!")
    quit()
# Service TR loading
if ws["I130"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I130"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I130"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[12]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with service TR loading!")
    quit()
# Downstream of ATO
if ws["I144"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I144"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I144"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[13]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with downstream of ATO!")
    quit()
# VSR
if ws["I143"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I143"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I143"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[14]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with VSR!")
    quit()
# Other facilities required
if ws["I145"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I145"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I145"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[15]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with other facilities required!")
    quit()
# Grounding requirement 1
if ws["I161"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I161"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I161"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[16]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with grounding requirement 1!")
    quit()
# Grounding requirement 2
if ws["I165"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I165"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I165"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[17]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with grounding requirement 2!")
    quit()
# Grounding requirement 3
if ws["I169"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I169"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I169"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[18]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with grounding requirement 3!")
    quit()
# Grounding requirement 4
if ws["I173"].value == 'Yes':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
elif ws["I173"].value == 'No':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.ENTER)
elif ws["I173"].value == 'N/A':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[19]/td[2]/span[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
else:
    print("Error with grounding requirement 4!")
    quit()
# Click save
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

# Go back to main IA page
driver.find_element("id", 'phSearchInput').send_keys(caseNum)
driver.find_element("id", 'phSearchButton').click()

# Click the top result from case number selector
driver.find_element("css selector", "a[target='_top']").click()
time.sleep(3)

driver.maximize_window()

time.sleep(1)
driver.switch_to.frame("0664O000000hsL8")
time.sleep(1)
driver.find_element("xpath", "/html/body/div[1]/div[1]/div/div[2]/button").click()
time.sleep(2)
if ws["I5"].value == "Pass":
    for count in range(1,3):
        if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Initial Engineering Screens Complete. Provide Interconnection Agreement.':
            driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
            driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
elif ws["I5"].value == "Fail":
    for count in range(1,3):
        if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'Further Study or Construction upgrade is required. The applicant will need to decide on next steps.':
            driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
            driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
else:
    print("Issue submitting application!")
    quit()

driver.minimize_window()

time.sleep(15)

print("Finished processing successfully!")

# Now let's move the screen and its pdf to the right spot
if not os.path.exists(os.path.join(pathToCompleted,subInitials,feeder)):
    os.makedirs(os.path.join(pathToCompleted,subInitials,feeder))
shutil.move(os.path.join(pathToJN,caseNum+".xlsm"),os.path.join(pathToCompleted,subInitials,feeder,caseNum+".xlsm"))
shutil.move(os.path.join(pathToJN,caseNum+".pdf"),os.path.join(pathToCompleted,subInitials,feeder,caseNum+".pdf"))
