# Made by Joe Nogosek, joe.nogosek@ulteig.com
# 8/9/2023
# NOTE: Tracker must be closed on your local computer!

# Import all our necessary libraries
import openpyxl
from openpyxl import load_workbook
import os
import shutil
import warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import glob
import zipfile
from selenium.webdriver.chrome.options import Options
from datetime import datetime
import win32com.client as win32
import os
import sys
import csv
import xlwings as xw

# Grab the credentials and file paths from the credentials library
user_env = os.getlogin()
credentials_path = os.path.join('C:\\Users', user_env, 'Documents', 'Local_Reviews')
sys.path.append(credentials_path)

import credentials

username = credentials.username
password = credentials.password
pathToTracker = credentials.path_to_MN_initials_tracker
pathToDownloadsFolder = credentials.path_to_downloads
path_to_driver = credentials.path_to_driver
name = credentials.name

warnings.filterwarnings('ignore')

# Move the new tool over if necessary
try:
    list_of_files = glob.glob(os.path.join(pathToDownloadsFolder,"202*.zip"))
    latest_file = max(list_of_files, key=os.path.getctime)

    with zipfile.ZipFile(latest_file, 'r') as zip_ref:
        zip_ref.extractall(pathToDownloadsFolder)
except:
    print("No files to unzip.")
    pass

pathToTool = r"C:\Users\joe.nogosek\Downloads\Joe Initial Review Tool.xlsm"
pathToMaster = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\Joe Initial Review Tool.xlsm"

try:
    shutil.move(pathToTool, pathToMaster)
    print("Moved the new tool over!")
except:
    print("Didn't move new tool over.")
    pass

# Reset our filepaths with the new tool location
pathToJN = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\JN"
pathToTool = r"C:\Users\joe.nogosek\Documents\Projects\NSP_Initial_Reviews\Joe Initial Review Tool.xlsm"

# Open the tracker
wbTracker = load_workbook(pathToTracker)
wsTracker = wbTracker['2023']

# Grab information about reviews to be performed from the tracker and store it in lists
num_reviews = 0
caseNums = []
premises = []
proj_sizes = []
rows_arr = []
for row in range(2,wsTracker.max_row+1):
    if wsTracker[row][7].value=="Joe Nogo" and wsTracker[row][8].value==None:
        caseNums.append(wsTracker[row][5].value)
        premises.append(wsTracker[row][4].value)
        proj_sizes.append(float(wsTracker[row][2].value))
        rows_arr.append(row)
        num_reviews=num_reviews+1

# Open an instance of Excel
# try:
#     app = xw.Book()
# except:
#     app = xw.App()

num_complete = 0

try:
    for instance in range(num_reviews):
        curFile = os.path.join(pathToJN, 'Initial Review_Case#0{}.xlsm'.format(caseNums[instance]))
        if not os.path.exists(os.path.join(pathToJN,curFile)):
            shutil.copy(pathToTool,pathToJN)
            os.rename(os.path.join(pathToJN,"Joe Initial Review Tool.xlsm"),curFile)
            wbTool = xw.Book(curFile)
            wsTool = xw.sheets[2]

            wsTool.range("C11").value = caseNums[instance]

            try:
                