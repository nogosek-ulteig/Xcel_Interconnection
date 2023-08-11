# Made by Joe Nogosek; joe.nogosek@ulteig.com
# 3/22/2022

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import shutil
import sys
import warnings
from selenium.webdriver.chrome.options import Options

# Enter your information here
username = 'joseph.h.nogosek@xcelenergy.com'
password = 'airdoc4Ee'
name = 'Joseph Nogosek'

warnings.filterwarnings("ignore", category=DeprecationWarning)

options = webdriver.ChromeOptions()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_experimental_option("detach", True)

IA = sys.argv[1]
pathToChecklist = sys.argv[2]
overUnder = sys.argv[3]
status = sys.argv[4]
reviewer = sys.argv[5]

print(IA)

wb = load_workbook(pathToChecklist, data_only = True)
ws = wb.active

driver = webdriver.Chrome(executable_path=r"G:\2021\21.00016\Reviews\chromedriver.exe", options=options)
driver.get('https://xcelenergy.my.salesforce.com/?ec=302&startURL=%2Fhome%2Fhome.jsp')
driver.implicitly_wait(30)
time.sleep(1)

xcel_login = driver.find_element("xpath", '//button[normalize-space()="Log in with Xcel Energy CORP credentials"]').click()
time.sleep(5)

user_box = driver.find_element("id", 'i0116')
user_box.send_keys(username)
driver.find_element("id", "idSIButton9").click()
time.sleep(3)

pass_box = driver.find_element("id", 'i0118')
pass_box.send_keys(password)
driver.find_element("id", "idSIButton9").click()
time.sleep(15)

try:
    driver.find_element("id", "idSIButton9").click()
    time.sleep(3)
except:
    pass

driver.minimize_window()
# Put the IA that is being processed here and pull it from the folder name
driver.find_element("id", 'phSearchInput').send_keys(IA)
driver.find_element("id", 'phSearchButton').click()

# Click the top result from case number selector
driver.find_element("css selector", "a[target='_top']").click()
time.sleep(8)

# Change the name of Area and meter engineer (if applicable)
driver.find_element("css selector", "[title='Edit']").click()
# If not Verify complete or ESS, then do both Area and Meter otherwise just do Area
driver.find_element("css selector", "[title='Area Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
driver.find_element("css selector", "[title='Area Engineer Approver']").send_keys(name)
if overUnder == "under" and ws["I2"].value != "ESS" and ws["C3"].value != 'N/A':
    driver.find_element("css selector", "[title='Meter Engineer Approver']").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
    driver.find_element("css selector", "[title='Meter Engineer Approver']").send_keys(name)
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
time.sleep(5)

# If verify complete, skip. If meter details already filled out, skip (Could be the same check for verify complete??)
if ws["C3"].value != "Verify Complete" and ws["I2"].value != "ESS" and overUnder == "under" and ws["C3"].value != 'N/A':
    # Open meter details action
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    time.sleep(1)
    for count in range(2,25):
        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Meter Details':
             driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
             time.sleep(2)
             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
             break
    time.sleep(3)
    # Get window handles
    p = driver.current_window_handle
    chwd = driver.window_handles
    for w in chwd:
        if(w!=p):
            driver.switch_to.window(w)
            driver.minimize_window()
    time.sleep(0.9)

    meter_swap = ws["H127"].value
    production_meter = ws["H128"].value
    time.sleep(1)
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/div[1]/lightning-textarea[1]/div[1]/textarea[1]").send_keys(meter_swap)
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/lightning-textarea[1]/div[1]/textarea[1]").send_keys(production_meter)
    # Choose whether it's an existing installation or not
    if "BD Existing" in ws["C3"].value:
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/lightning-radio-group[1]/fieldset[1]/div[1]/div[1]/span[1]/label[1]/span[1]").click()
    else:
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/lightning-radio-group[1]/fieldset[1]/div[1]/div[1]/span[2]/label[1]/span[1]").click()
    # Depending on production meter select proper from dropdown
    if production_meter == "NEX":
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.DOWN, Keys.DOWN, Keys.ENTER)
        driver.find_element("css selector", "button[name='NEXT']").click()
    else:
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        driver.find_element("css selector", "button[name='NEXT']").click()
        time.sleep(2)
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[5]/div[1]/div[3]/button[1]").click()

    # Close meter details and select correct window handle going back to the correct IA page
    time.sleep(8)
    driver.close()
    driver.switch_to.window(p)
    driver.minimize_window()
    driver.find_element("id", 'phSearchInput').send_keys(IA)
    driver.find_element("id", 'phSearchButton').click()
    driver.find_element("css selector", "a[target='_top']").click()
    time.sleep(5)

# Uncheck everything that is checked if verify complete
if ws["C3"].value == 'Verify Complete' or ws["D5"].value == 'Verify Complete':
    # Click the 'Go to List' page for checks
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[17]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    for counter in range(2,100):
        isChecked = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[5]/img[1]").get_attribute("alt")
        if isChecked == 'Checked':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

    # Go to the next page
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[2]/a[1]").click()
    time.sleep(3)
    for counter in range(2,100):
        isChecked = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[5]/img[1]").get_attribute("alt")
        if isChecked == 'Checked':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[2]/a[2]").click()
    time.sleep(3)
    for counter in range(2,13):
        isChecked = driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[5]/img[1]").get_attribute("alt")
        if isChecked == 'Checked':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

    # Go back to interconnection app
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[1]/div[2]/a[1]").click()
    time.sleep(3)

# Fill out checks if application is getting rejected
if ws["C5"].value == 'Rejected' or ws["C4"].value == 'Rejected':
    # Click the 'Go to List' page for checks
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[17]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
    time.sleep(1)
    if overUnder == "under":
        # 3.21
        if ws["F49"].value == 'x' or ws["F49"].value == 'X' or ws["F49"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.1
        if ws["F50"].value == 'x' or ws["F50"].value == 'X' or ws["F50"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.2
        if ws["F52"].value == 'x' or ws["F52"].value == 'X' or ws["F52"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.3
        if ws["F53"].value == 'x' or ws["F53"].value == 'X' or ws["F53"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.4
        if ws["F54"].value == 'x' or ws["F54"].value == 'X' or ws["F54"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.1
        if ws["F13"].value == 'x' or ws["F13"].value == 'X' or ws["F13"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.2
        if ws["F14"].value == 'x' or ws["F14"].value == 'X' or ws["F14"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.3
        if ws["F15"].value == 'x' or ws["F15"].value == 'X' or ws["F15"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.10
        if ws["F27"].value == 'x' or ws["F27"].value == 'X' or ws["F27"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.10":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.11
        if ws["F28"].value == 'x' or ws["F28"].value == 'X' or ws["F28"].value == 'CA' or ws["F29"].value == 'x' or ws["F29"].value == 'X' or ws["F29"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.11":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.12
        if ws["F30"].value == 'x' or ws["F30"].value == 'X' or ws["F30"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.12":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.13
        if ws["F31"].value == 'x' or ws["F31"].value == 'X' or ws["F31"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.13":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.14
        if ws["F34"].value == 'x' or ws["F34"].value == 'X' or ws["F34"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.14":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.15
        if ws["F35"].value == 'x' or ws["F35"].value == 'X' or ws["F35"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.15":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.16
        if ws["F36"].value == 'x' or ws["F36"].value == 'X' or ws["F36"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.16":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.17
        if ws["F37"].value == 'x' or ws["F37"].value == 'X' or ws["F37"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.17":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.5
        if ws["F20"].value == 'x' or ws["F20"].value == 'X' or ws["F20"].value == 'CA' or ws["F21"].value == 'x' or ws["F21"].value == 'X' or ws["F21"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.5":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.6
        if ws["F22"].value == 'x' or ws["F22"].value == 'X' or ws["F22"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.6":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.7
        if ws["F23"].value == 'x' or ws["F23"].value == 'X' or ws["F23"].value == 'CA' or ws["F24"].value == 'x' or ws["F24"].value == 'X' or ws["F24"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.7":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break

        # Go to next page
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[1]/a[1]").click()

        # 3.8
        if ws["F25"].value == 'x' or ws["F25"].value == 'X' or ws["F25"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.8":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.9
        if ws["F26"].value == 'x' or ws["F26"].value == 'X' or ws["F26"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.9":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.18
        if ws["F39"].value == 'x' or ws["F39"].value == 'X' or ws["F39"].value == 'CA' or ws["F42"].value == 'x' or ws["F42"].value == 'X' or ws["F42"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.18":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.18.1
        if ws["F40"].value == 'x' or ws["F40"].value == 'X' or ws["F40"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.18.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.18.2
        if ws["F41"].value == 'x' or ws["F41"].value == 'X' or ws["F41"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.18.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.18.3
        if ws["F43"].value == 'x' or ws["F43"].value == 'X' or ws["F43"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.18.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.19
        if ws["F44"].value == 'x' or ws["F44"].value == 'X' or ws["F44"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.19":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.19.1
        if ws["F46"].value == 'x' or ws["F46"].value == 'X' or ws["F46"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.19.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.20
        if ws["F47"].value == 'x' or ws["F47"].value == 'X' or ws["F47"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.20":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.1
        if ws["F67"].value == 'x' or ws["F67"].value == 'X' or ws["F67"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.2
        if ws["F68"].value == 'x' or ws["F68"].value == 'X' or ws["F68"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break

        # Go to the next page
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[2]/a[2]").click()

        # 4.3
        if ws["F72"].value == 'x' or ws["F72"].value == 'X' or ws["F72"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.4
        if ws["F73"].value == 'x' or ws["F73"].value == 'X' or ws["F73"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.5
        if ws["F74"].value == 'x' or ws["F74"].value == 'X' or ws["F74"].value == 'CA' or ws["F75"].value == 'x' or ws["F75"].value == 'X' or ws["F75"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.5":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.6
        if ws["F76"].value == 'x' or ws["F76"].value == 'X' or ws["F76"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.6":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.10
        if ws["F88"].value == 'x' or ws["F88"].value == 'X' or ws["F88"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.10":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.11
        if ws["F89"].value == 'x' or ws["F89"].value == 'X' or ws["F89"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.11":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.7
        if ws["F81"].value == 'x' or ws["F81"].value == 'X' or ws["F81"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.7":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.7.1
        if ws["F82"].value == 'x' or ws["F82"].value == 'X' or ws["F82"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.7.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.8
        if ws["F85"].value == 'x' or ws["F85"].value == 'X' or ws["F85"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.8":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.8.1
        if ws["F86"].value == 'x' or ws["F86"].value == 'X' or ws["F86"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.8.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.9
        if ws["F87"].value == 'x' or ws["F87"].value == 'X' or ws["F87"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.9":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
    elif overUnder == "over":
        # 3.21
        if ws["F48"].value == 'x' or ws["F48"].value == 'X' or ws["F48"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.1
        if ws["F49"].value == 'x' or ws["F49"].value == 'X' or ws["F49"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.2
        if ws["F51"].value == 'x' or ws["F51"].value == 'X' or ws["F51"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.3
        if ws["F52"].value == 'x' or ws["F52"].value == 'X' or ws["F52"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.21.4
        if ws["F53"].value == 'x' or ws["F53"].value == 'X' or ws["F53"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.21.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.1
        if ws["F12"].value == 'x' or ws["F12"].value == 'X' or ws["F12"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.2
        if ws["F13"].value == 'x' or ws["F13"].value == 'X' or ws["F13"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.3
        if ws["F14"].value == 'x' or ws["F14"].value == 'X' or ws["F14"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.4
        if ws["F16"].value == 'x' or ws["F16"].value == 'X' or ws["F16"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.10
        if ws["F28"].value == 'x' or ws["F28"].value == 'X' or ws["F28"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.10":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.11
        if ws["F29"].value == 'x' or ws["F29"].value == 'X' or ws["F29"].value == 'CA' or ws["F30"].value == 'x' or ws["F30"].value == 'X' or ws["F30"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.11":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.12
        if ws["F31"].value == 'x' or ws["F31"].value == 'X' or ws["F31"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.12":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.13
        if ws["F32"].value == 'x' or ws["F32"].value == 'X' or ws["F32"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.13":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.14
        if ws["F33"].value == 'x' or ws["F33"].value == 'X' or ws["F33"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.14":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.15
        if ws["F34"].value == 'x' or ws["F34"].value == 'X' or ws["F34"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.15":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.16
        if ws["F35"].value == 'x' or ws["F35"].value == 'X' or ws["F35"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.16":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.17
        if ws["F36"].value == 'x' or ws["F36"].value == 'X' or ws["F36"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.17":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.5
        if ws["F19"].value == 'x' or ws["F19"].value == 'X' or ws["F19"].value == 'CA' or ws["F20"].value == 'x' or ws["F20"].value == 'X' or ws["F20"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.5":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.6
        if ws["F22"].value == 'x' or ws["F22"].value == 'X' or ws["F22"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.6":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.7
        if ws["F23"].value == 'x' or ws["F23"].value == 'X' or ws["F23"].value == 'CA' or ws["F24"].value == 'x' or ws["F24"].value == 'X' or ws["F24"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.7":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break

        # Go to next page
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[2]/div[1]/div[1]/a[1]").click()

        # 3.8
        if ws["F25"].value == 'x' or ws["F25"].value == 'X' or ws["F25"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.8":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.9
        if ws["F27"].value == 'x' or ws["F27"].value == 'X' or ws["F27"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.9":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22
        if ws["F55"].value == 'x' or ws["F55"].value == 'X' or ws["F55"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1
        if ws["F56"].value == 'x' or ws["F56"].value == 'X' or ws["F56"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1.1
        if ws["F57"].value == 'x' or ws["F57"].value == 'X' or ws["F57"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1.2
        if ws["F58"].value == 'x' or ws["F58"].value == 'X' or ws["F58"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1.3
        if ws["F59"].value == 'x' or ws["F59"].value == 'X' or ws["F59"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1.3":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1.4
        if ws["F60"].value == 'x' or ws["F60"].value == 'X' or ws["F60"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.1.5
        if ws["F61"].value == 'x' or ws["F61"].value == 'X' or ws["F61"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.1.5":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.2
        if ws["F62"].value == 'x' or ws["F62"].value == 'X' or ws["F62"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 3.22.2.1
        if ws["F63"].value == 'x' or ws["F63"].value == 'X' or ws["F63"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "3.22.2.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.1
        if ws["F66"].value == 'x' or ws["F66"].value == 'X' or ws["F66"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.2
        if ws["F67"].value == 'x' or ws["F67"].value == 'X' or ws["F67"].value == 'CA':
            for counter in range(2,100):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break

        # Go to the next page
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[5]/div[1]/div[2]/a[2]").click()

        # 4.4
        if ws["F71"].value == 'x' or ws["F71"].value == 'X' or ws["F71"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.4":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.5
        if ws["F72"].value == 'x' or ws["F72"].value == 'X' or ws["F72"].value == 'CA' or ws["F73"].value == 'x' or ws["F73"].value == 'X' or ws["F73"].value == 'CA' or ws["F74"].value == 'x' or ws["F74"].value == 'X' or ws["F74"].value == 'CA' or ws["F75"].value == 'x' or ws["F75"].value == 'X' or ws["F75"].value == 'CA' or ws["F76"].value == 'x' or ws["F76"].value == 'X' or ws["F76"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.5":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.6
        if ws["F77"].value == 'x' or ws["F77"].value == 'X' or ws["F77"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.6":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.10
        if ws["F87"].value == 'x' or ws["F87"].value == 'X' or ws["F87"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.10":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.11
        if ws["F88"].value == 'x' or ws["F88"].value == 'X' or ws["F88"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.11":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.7
        if ws["F79"].value == 'x' or ws["F79"].value == 'X' or ws["F79"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.7":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.7.1
        if ws["F80"].value == 'x' or ws["F80"].value == 'X' or ws["F80"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.7.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.7.2
        if ws["F81"].value == 'x' or ws["F81"].value == 'X' or ws["F81"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.7.2":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.8
        if ws["F84"].value == 'x' or ws["F84"].value == 'X' or ws["F84"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.8":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
        # 4.8.1
        if ws["F85"].value == 'x' or ws["F85"].value == 'X' or ws["F85"].value == 'CA':
            for counter in range(2,13):
                if driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[3]").get_attribute("innerText") == "4.8.1":
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(counter) + "]/td[1]/a[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[3]/table[1]/tbody[1]/tr[2]/td[2]/input[1]").click()
                    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
                    break
    # Go back to interconnection app
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[1]/div[2]/a[1]").click()
    time.sleep(3)

# Reopen actions
flag = 0
driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[10]/div[1]/div[1]/div[2]/div[1]/a[2]").click()
time.sleep(1)
try:
    for count in range(2,25):
        if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Reopen Action (Review Missing Info)':
             driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
             time.sleep(2)
             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
             flag += 1
             break
except:
    pass
if flag == 0:
    try:
        for count in range(2,25):
            if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/td[2]").get_attribute("innerText") == 'Reopen Action (Initiate Application)':
                 driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[3]/div[1]/div[2]/table[1]/tbody[1]/tr[" + str(count) + "]/th[1]/a[1]").click()
                 time.sleep(2)
                 driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[2]/div[2]/table[1]/tbody[1]/tr[10]/td[2]/div[1]/a[1]").click()
                 break
    except:
        pass
time.sleep(4)
# Move to the correct tab
p = driver.current_window_handle
chwd = driver.window_handles
for w in chwd:
    if(w!=p):
        driver.switch_to.window(w)
        driver.minimize_window()
time.sleep(2)

# Check if actions other than what was caught in checks need to be reopened
flag_2 = 0
if overUnder == "under":
    if ws["C5"].value == 'Rejected':
        # Check system details
        if ws["C7"].value == 'x' or ws["C7"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'System Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Check Application Details
        if ws["C8"].value == 'x' or ws["C8"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Application Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Check Exhibit B Details
        if ws["C9"].value == 'x' or ws["C9"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Exhibit B Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Now we have checked all that are necessary. If flag_2 has been raised, we must stay on 'Yes' and click 'Submit'
        if flag_2 > 0:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/div[1]/button[1]").click()
            time.sleep(5)
            driver.close()
            driver.switch_to.window(p)

        # If flag_2 has not been raised, then we select 'No' and click 'Submit'
        else:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
            time.sleep(5)
            driver.close()
            driver.switch_to.window(p)
            driver.minimize_window()

    # If not getting rejected, just say no and move on
    else:
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
        time.sleep(5)
        driver.close()
        driver.switch_to.window(p)
        driver.minimize_window()
elif overUnder == "over":
    if ws["C4"].value == 'Rejected':
        # Check system details
        if ws["C6"].value == 'x' or ws["C6"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'System Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Check Application Details
        if ws["C7"].value == 'x' or ws["C7"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Application Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Check Exhibit B Details
        if ws["C8"].value == 'x' or ws["C8"].value == 'X':
            try:
                for count in range(1,20):
                    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[2]/lightning-formatted-rich-text[1]/span[1]").get_attribute("innerText") == 'Exhibit B Details':
                         driver.find_element("xpath", f"/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/flowruntime-multi-checkbox-lwc[1]/fieldset[1]/div[1]/span[{count}]/label[1]/span[1]").click()
                         flag_2 += 1
                         break
            except:
                pass
        # Now we have checked all that are necessary. If flag_2 has been raised, we must stay on 'Yes' and click 'Submit'
        if flag_2 > 0:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[4]/div[1]/div[1]/button[1]").click()
            time.sleep(5)
            driver.close()
            driver.switch_to.window(p)

        # If flag_2 has not been raised, then we select 'No' and click 'Submit'
        else:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
            time.sleep(5)
            driver.close()
            driver.switch_to.window(p)
            driver.minimize_window()

    # If not getting rejected, just say no and move on
    else:
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[2]/flowruntime-picklist-input-lwc[1]/div[1]/lightning-select[1]/div[1]/div[1]/select[1]").send_keys(Keys.ENTER, Keys.DOWN, Keys.ENTER)
        driver.find_element("xpath", "/html[1]/body[1]/div[1]/article[1]/div[1]/div[3]/div[1]/div[1]/div[3]/div[1]/div[1]/button[1]").click()
        time.sleep(5)
        driver.close()
        driver.switch_to.window(p)
        driver.minimize_window()

# Go back to the main page so we are ready for the next task
driver.back()
driver.back()
time.sleep(5)

# Go to approve/reject page
for count in range(3,5):
    if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ross Kirby' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == name:
        driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[1]/a[2]").click()
        break
# Paste the comments in
if overUnder == "under":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/textarea[1]").send_keys(ws["D5"].value)
elif overUnder == "over":
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[2]/table[1]/tbody[1]/tr[2]/td[2]/textarea[1]").send_keys(ws["D4"].value)
# Click approve or reject
if ws["C5"].value == 'Rejected' or ws["C4"].value == 'Rejected':
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
else:
    driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()

if status == 'Approved':
    status = 'Appr'
elif status == 'Rejected':
    status = 'Rej'
else:
    status = 'CA'

print("First approval/rejection successfully completed.")

pathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}'.format(reviewer, IA, reviewer)
newPathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}'.format(reviewer, IA, reviewer, status)
v2Path = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v2'.format(reviewer, IA, reviewer, status)
v3Path = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v3'.format(reviewer, IA, reviewer, status)
v4Path = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v4'.format(reviewer, IA, reviewer, status)
v5Path = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v5'.format(reviewer, IA, reviewer, status)
os.rename(pathToFolder, newPathToFolder)

pathToApproved = r'G:\2021\21.00016\Reviews\Complete-Approved'
pathToRejected = r'G:\2021\21.00016\Reviews\Complete-Rejected'
rejectedV2 = os.path.join(pathToRejected, '{0}-{1}-{2}'.format(IA, reviewer, status))
rejectedV3 = os.path.join(pathToRejected, '{0}-{1}-{2}v2'.format(IA, reviewer, status))
rejectedV4 = os.path.join(pathToRejected, '{0}-{1}-{2}v3'.format(IA, reviewer, status))
rejectedV5 = os.path.join(pathToRejected, '{0}-{1}-{2}v4'.format(IA, reviewer, status))
if status == 'Appr' or status == 'CA':
    shutil.move(newPathToFolder, pathToApproved)
else:
    if not os.path.exists(rejectedV2) and not os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
        shutil.move(newPathToFolder, pathToRejected)
    elif os.path.exists(rejectedV2) and not os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
        os.rename(newPathToFolder, v2Path)
        newPathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v2'.format(reviewer, IA, reviewer, status)
        shutil.move(newPathToFolder, pathToRejected)
    elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and not os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
        os.rename(newPathToFolder, v3Path)
        newPathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v3'.format(reviewer, IA, reviewer, status)
        shutil.move(newPathToFolder, pathToRejected)
    elif os.path.exists(rejectedV2) and os.path.exists(rejectedV3) and os.path.exists(rejectedV4) and not os.path.exists(rejectedV5):
        os.rename(newPathToFolder, v4Path)
        newPathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v4'.format(reviewer, IA, reviewer, status)
        shutil.move(newPathToFolder, pathToRejected)
    else:
        os.rename(newPathToFolder, v5Path)
        newPathToFolder = r'G:\2021\21.00016\Reviews\{0}\{1}-{2}-{3}v5'.format(reviewer, IA, reviewer, status)
        shutil.move(newPathToFolder, pathToRejected)

# Wait for page to load again and give it about one minute then refresh and push application through again
if ws["I2"].value != "ESS" and ws["C3"].value != "N/A" and overUnder == "under":
    time.sleep(80)
    driver.refresh()
    time.sleep(5)
    try:
        for count in range(3,5):
            if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ross Kirby' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == name:
                driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[1]/a[2]").click()
                break
        if ws["C5"].value == 'Rejected' or ws["C4"].value == 'Rejected':
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
        else:
            driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
    except:
        if (status == 'Approved' or status == 'CA'):
            print("MADE IT IN!")
            driver.switch_to.frame("0664O000000hsL8")
            driver.find_element("xpath", "/html/body/div[1]/div[1]/div/div[2]/button").click()
            time.sleep(2)
            for count in range(1,3):
                if driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[2]/lightning-formatted-rich-text/span").get_attribute("innerText") == 'The application is deemed complete and is ready for engineer review.':
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/div/div/div/flowruntime-radio-button-input-lwc/fieldset/div/span[" + str(count) + "]/label/span[1]").click()
                    driver.find_element("xpath", "/html/body/div[1]/section/div/div[3]/footer/div[2]/button").click()
                    time.sleep(15)
else:
    pass

print("Second approval/rejection successfully completed (if necessary)")


# # Wait for page to load again and give it about one minute then refresh and push application through again
# if ws["I2"].value != "ESS" and ws["C3"].value != "N/A" and overUnder == "under":
#     time.sleep(60)
#     driver.refresh()
#     time.sleep(5)
#     if (status == 'Approved' or status == 'CA') and (driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[3]/td[2]").get_attribute("innerText") == 'Rejected' and driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[2]").get_attribute("innerText") == 'Submitted'):
#         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[4]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[3]").click()
#         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[2]/div[55]/table[1]/tbody[1]/tr[2]/td[2]/span[1]/input[1]").send_keys(Keys.CONTROL + 'a', Keys.BACKSPACE)
#         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
#         time.sleep(3)
#         driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[1]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
#         time.sleep(5)
#         for count in range(3,5):
#             if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ross Kirby' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == name:
#                 driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[1]/a[2]").click()
#                 break
#         if ws["C5"].value == 'Rejected' or ws["C4"].value == 'Rejected':
#             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
#         else:
#             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
#     else:
#         for count in range(3,5):
#             if driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Josh Guck' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Joseph Nogosek' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == 'Ross Kirby' or driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[4]/a[1]").get_attribute("innerText") == name:
#                 driver.find_element("xpath", f"/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/div[14]/div[1]/div[1]/div[2]/table[1]/tbody[1]/tr[{count}]/td[1]/a[2]").click()
#                 break
#         if ws["C5"].value == 'Rejected' or ws["C4"].value == 'Rejected':
#             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[2]").click()
#         else:
#             driver.find_element("xpath", "/html[1]/body[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/form[1]/div[1]/div[3]/table[1]/tbody[1]/tr[1]/td[2]/input[1]").click()
# else:
#     pass
#
# print("Second approval/rejection successfully completed (if necessary)")


# time.sleep(60)
# # Close the page now that the application has been processed
# driver.close()
